from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import date, timedelta
import json
from .models import *
from .decorators import permission_required, subscription_required

@login_required
@subscription_required
@permission_required('accounts', 'view')
def journal_entries_view(request):
    """عرض القيود اليومية"""
    try:
        entries = JournalEntry.objects.select_related('created_by').order_by('-created_at')[:50]
    except:
        entries = []
    
    context = {
        'entries': entries,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'accounting/journal_entries.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def account_ledger_view(request, account_id):
    """دفتر الأستاذ للحساب"""
    account = get_object_or_404(Account, id=account_id)
    
    try:
        entries = JournalEntryLine.objects.filter(
            account=account
        ).select_related('journal_entry').order_by('-journal_entry__created_at')[:100]
    except:
        entries = []
    
    context = {
        'account': account,
        'entries': entries,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'accounting/account_ledger.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def accounts_list_view(request):
    """قائمة الحسابات المحاسبية"""
    accounts = Account.objects.all().order_by('account_code')
    
    search = request.GET.get('search', '')
    if search:
        accounts = accounts.filter(
            Q(name__icontains=search) |
            Q(account_code__icontains=search)
        )
    
    context = {
        'accounts': accounts,
        'search': search,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'accounting/accounts_list.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def customer_statement_view(request, customer_id):
    """كشف حساب العميل"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    # جلب الفواتير والدفعات
    sales = Sale.objects.filter(customer=customer, status='confirmed').order_by('-created_at')
    try:
        payments = CustomerPayment.objects.filter(customer=customer).order_by('-created_at')
    except:
        payments = []
    
    # دمج المعاملات
    transactions = []
    
    # إضافة الفواتير
    for sale in sales:
        transactions.append({
            'date': sale.created_at.date(),
            'description': f'فاتورة بيع #{sale.invoice_number}',
            'debit': sale.total_amount,
            'credit': 0,
            'type': 'sale'
        })
    
    # إضافة الدفعات
    for payment in payments:
        transactions.append({
            'date': payment.payment_date,
            'description': f'دفعة #{payment.payment_number}',
            'debit': 0,
            'credit': payment.amount,
            'type': 'payment'
        })
    
    # ترتيب المعاملات حسب التاريخ
    transactions.sort(key=lambda x: x['date'], reverse=True)
    
    # حساب الرصيد التراكمي
    running_balance = customer.opening_balance
    for transaction in reversed(transactions):
        running_balance += transaction['debit'] - transaction['credit']
        transaction['balance'] = running_balance
    
    context = {
        'customer': customer,
        'transactions': transactions,
        'opening_balance': customer.opening_balance,
        'final_balance': running_balance,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'accounting/customer_statement.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def supplier_statement_view(request, supplier_id):
    """كشف حساب المورد"""
    supplier = get_object_or_404(Supplier, id=supplier_id)
    
    # جلب المشتريات والدفعات
    purchases = Purchase.objects.filter(supplier=supplier, status='confirmed').order_by('-created_at')
    try:
        payments = SupplierPayment.objects.filter(supplier=supplier).order_by('-created_at')
    except:
        payments = []
    
    # دمج المعاملات
    transactions = []
    
    # إضافة المشتريات
    for purchase in purchases:
        transactions.append({
            'date': purchase.created_at.date(),
            'description': f'فاتورة شراء #{purchase.invoice_number}',
            'debit': 0,
            'credit': purchase.total_amount,
            'type': 'purchase'
        })
    
    # إضافة الدفعات
    for payment in payments:
        transactions.append({
            'date': payment.payment_date,
            'description': f'دفعة #{payment.payment_number}',
            'debit': payment.amount,
            'credit': 0,
            'type': 'payment'
        })
    
    # ترتيب المعاملات حسب التاريخ
    transactions.sort(key=lambda x: x['date'], reverse=True)
    
    # حساب الرصيد التراكمي
    running_balance = supplier.opening_balance
    for transaction in reversed(transactions):
        running_balance += transaction['credit'] - transaction['debit']
        transaction['balance'] = running_balance
    
    context = {
        'supplier': supplier,
        'transactions': transactions,
        'opening_balance': supplier.opening_balance,
        'final_balance': running_balance,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'accounting/supplier_statement.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def trial_balance_view(request):
    """ميزان المراجعة"""
    accounts = Account.objects.all().order_by('account_code')
    total_debits = 0
    total_credits = 0
    
    account_balances = []
    for account in accounts:
        if account.account_type in ['asset', 'expense']:
            debit_balance = account.balance if account.balance > 0 else 0
            credit_balance = 0
        else:
            debit_balance = 0
            credit_balance = account.balance if account.balance > 0 else 0
        
        total_debits += debit_balance
        total_credits += credit_balance
        
        account_balances.append({
            'account': account,
            'debit': debit_balance,
            'credit': credit_balance
        })
    
    context = {
        'account_balances': account_balances,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'is_balanced': total_debits == total_credits,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'accounting/trial_balance.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def income_statement_view(request):
    """قائمة الدخل"""
    # حساب الإيرادات
    revenue_accounts = Account.objects.filter(account_type='revenue')
    total_revenue = revenue_accounts.aggregate(Sum('balance'))['balance__sum'] or 0
    
    # حساب المصروفات
    expense_accounts = Account.objects.filter(account_type='expense')
    total_expenses = expense_accounts.aggregate(Sum('balance'))['balance__sum'] or 0
    
    # صافي الدخل
    net_income = total_revenue - total_expenses
    
    context = {
        'revenue_accounts': revenue_accounts,
        'expense_accounts': expense_accounts,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_income': net_income,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'income_statement.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def balance_sheet_view(request):
    """الميزانية العمومية"""
    # الأصول
    asset_accounts = Account.objects.filter(account_type='asset')
    total_assets = asset_accounts.aggregate(Sum('balance'))['balance__sum'] or 0
    
    # الخصوم
    liability_accounts = Account.objects.filter(account_type='liability')
    total_liabilities = liability_accounts.aggregate(Sum('balance'))['balance__sum'] or 0
    
    # حقوق الملكية
    equity_accounts = Account.objects.filter(account_type='equity')
    total_equity = equity_accounts.aggregate(Sum('balance'))['balance__sum'] or 0
    
    context = {
        'asset_accounts': asset_accounts,
        'liability_accounts': liability_accounts,
        'equity_accounts': equity_accounts,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'total_liabilities_equity': total_liabilities + total_equity,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'accounting/balance_sheet.html', context)

@login_required
def search_accounts_api(request):
    """API البحث في الحسابات"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'accounts': []})
    
    accounts = Account.objects.filter(
        Q(name__icontains=query) |
        Q(account_code__icontains=query)
    )[:10]
    
    accounts_data = []
    for account in accounts:
        accounts_data.append({
            'id': account.id,
            'name': account.name,
            'code': account.account_code,
            'type': account.get_account_type_display(),
            'balance': float(account.balance)
        })
    
    return JsonResponse({'accounts': accounts_data})

@login_required
def search_customers_api(request):
    """API البحث في العملاء"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'customers': []})
    
    customers = Customer.objects.filter(
        Q(name__icontains=query) |
        Q(phone__icontains=query)
    )[:10]
    
    customers_data = []
    for customer in customers:
        customers_data.append({
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone,
            'balance': float(customer.opening_balance)
        })
    
    return JsonResponse({'customers': customers_data})

@login_required
def search_suppliers_api(request):
    """API البحث في الموردين"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'suppliers': []})
    
    suppliers = Supplier.objects.filter(
        Q(name__icontains=query) |
        Q(phone__icontains=query)
    )[:10]
    
    suppliers_data = []
    for supplier in suppliers:
        suppliers_data.append({
            'id': supplier.id,
            'name': supplier.name,
            'phone': supplier.phone,
            'balance': float(supplier.opening_balance)
        })
    
    return JsonResponse({'suppliers': suppliers_data})

@login_required
@subscription_required
def export_journal_entries(request):
    """تصدير القيود اليومية"""
    try:
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "القيود اليومية"
        
        headers = ['رقم القيد', 'التاريخ', 'الوصف', 'المبلغ', 'الحالة']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        try:
            entries = JournalEntry.objects.all()[:100]
            for row, entry in enumerate(entries, 2):
                ws.cell(row=row, column=1, value=entry.entry_number)
                ws.cell(row=row, column=2, value=entry.created_at.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=3, value=entry.description)
                ws.cell(row=row, column=4, value=float(entry.amount))
                ws.cell(row=row, column=5, value='مرحل' if entry.is_posted else 'غير مرحل')
        except:
            pass
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="journal_entries.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('accounting_journal_entries')

@login_required
@subscription_required
def export_trial_balance(request):
    """تصدير ميزان المراجعة"""
    try:
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ميزان المراجعة"
        
        headers = ['رمز الحساب', 'اسم الحساب', 'نوع الحساب', 'مدين', 'دائن']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        accounts = Account.objects.all().order_by('account_code')
        for row, account in enumerate(accounts, 2):
            ws.cell(row=row, column=1, value=account.account_code)
            ws.cell(row=row, column=2, value=account.name)
            ws.cell(row=row, column=3, value=account.get_account_type_display())
            
            if account.account_type in ['asset', 'expense']:
                ws.cell(row=row, column=4, value=float(account.balance) if account.balance > 0 else 0)
                ws.cell(row=row, column=5, value=0)
            else:
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=float(account.balance) if account.balance > 0 else 0)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="trial_balance.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('accounting_trial_balance')

@login_required
@subscription_required
def balance_sheet_export(request):
    """تصدير الميزانية العمومية"""
    messages.success(request, 'سيتم إضافة هذه الميزة قريباً')
    return redirect('accounting_balance_sheet')

@login_required
@subscription_required
def accounting_reports_menu(request):
    """قائمة التقارير المحاسبية"""
    context = {
        'currency_symbol': 'د.ك',
    }
    return render(request, 'accounting/reports_menu.html', context)