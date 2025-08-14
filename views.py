# استيراد Views المحسنة
from .views_fixed import *
from .views_missing import *
# # from .permission_views import *

@login_required
def income_statement_report(request):
    from django.db.models import Sum
    from datetime import date, datetime
    
    # الحصول على فترة التقرير من المعاملات
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # إذا لم يتم تحديد التواريخ، استخدم السنة الحالية
    if not start_date or not end_date:
        current_year = date.today().year
        start_date = f"{current_year}-01-01"
        end_date = f"{current_year}-12-31"
    
    # تحويل التواريخ
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except:
        current_year = date.today().year
        start_date_obj = date(current_year, 1, 1)
        end_date_obj = date(current_year, 12, 31)
        start_date = start_date_obj.strftime('%Y-%m-%d')
        end_date = end_date_obj.strftime('%Y-%m-%d')
    
    # جلب الإيرادات من شجرة الحسابات
    revenue_accounts = Account.objects.filter(account_type='revenue')
    revenues = []
    total_revenue = 0
    
    for account in revenue_accounts:
        # حساب رصيد الحساب من المبيعات المؤكدة في الفترة المحددة
        if account.account_code == '4001':  # حساب المبيعات
            sales_amount = Sale.objects.filter(
                status='confirmed',
                created_at__date__gte=start_date_obj,
                created_at__date__lte=end_date_obj
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            
            if sales_amount > 0:
                revenues.append({
                    'name': account.name,
                    'amount': float(sales_amount),
                    'account_code': account.account_code
                })
                total_revenue += float(sales_amount)
        else:
            # للحسابات الأخرى، استخدم الرصيد المحفوظ
            if account.balance > 0:
                revenues.append({
                    'name': account.name,
                    'amount': float(account.balance),
                    'account_code': account.account_code
                })
                total_revenue += float(account.balance)
    
    # إضافة إيرادات أخرى إذا لم توجد في شجرة الحسابات
    if not any(r['account_code'] == '4001' for r in revenues):
        sales_amount = Sale.objects.filter(
            status='confirmed',
            created_at__date__gte=start_date_obj,
            created_at__date__lte=end_date_obj
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        if sales_amount > 0:
            revenues.append({
                'name': 'مبيعات',
                'amount': float(sales_amount),
                'account_code': '4001'
            })
            total_revenue += float(sales_amount)
    
    # جلب المصروفات من شجرة الحسابات
    expense_accounts = Account.objects.filter(account_type='expense')
    expenses = []
    total_expense = 0
    
    for account in expense_accounts:
        # حساب رصيد الحساب من المشتريات المؤكدة في الفترة المحددة
        if account.account_code == '5001':  # حساب المشتريات
            purchases_amount = Purchase.objects.filter(
                status='confirmed',
                created_at__date__gte=start_date_obj,
                created_at__date__lte=end_date_obj
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            
            if purchases_amount > 0:
                expenses.append({
                    'name': account.name,
                    'amount': float(purchases_amount),
                    'account_code': account.account_code
                })
                total_expense += float(purchases_amount)
        elif account.account_code.startswith('21'):  # حسابات الرواتب
            # حساب الرواتب المؤكدة في الفترة
            salaries_amount = Salary.objects.filter(
                status__in=['confirmed', 'paid'],
                created_at__date__gte=start_date_obj,
                created_at__date__lte=end_date_obj
            ).aggregate(Sum('net_salary'))['net_salary__sum'] or 0
            
            if salaries_amount > 0:
                expenses.append({
                    'name': account.name,
                    'amount': float(salaries_amount),
                    'account_code': account.account_code
                })
                total_expense += float(salaries_amount)
        else:
            # للحسابات الأخرى، استخدم الرصيد المحفوظ
            if account.balance > 0:
                expenses.append({
                    'name': account.name,
                    'amount': float(account.balance),
                    'account_code': account.account_code
                })
                total_expense += float(account.balance)
    
    # إضافة مصروفات أساسية إذا لم توجد في شجرة الحسابات
    if not any(e['account_code'] == '5001' for e in expenses):
        purchases_amount = Purchase.objects.filter(
            status='confirmed',
            created_at__date__gte=start_date_obj,
            created_at__date__lte=end_date_obj
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        if purchases_amount > 0:
            expenses.append({
                'name': 'مشتريات',
                'amount': float(purchases_amount),
                'account_code': '5001'
            })
            total_expense += float(purchases_amount)
    
    # إضافة الرواتب إذا لم توجد
    if not any(e['account_code'].startswith('21') for e in expenses):
        salaries_amount = Salary.objects.filter(
            status__in=['confirmed', 'paid'],
            created_at__date__gte=start_date_obj,
            created_at__date__lte=end_date_obj
        ).aggregate(Sum('net_salary'))['net_salary__sum'] or 0
        
        if salaries_amount > 0:
            expenses.append({
                'name': 'رواتب وأجور',
                'amount': float(salaries_amount),
                'account_code': '2101'
            })
            total_expense += float(salaries_amount)
    
    # حساب صافي الربح
    net_income = total_revenue - total_expense
    
    # إحصائيات إضافية
    stats = {
        'period': f"من {start_date} إلى {end_date}",
        'revenue_accounts_count': len(revenues),
        'expense_accounts_count': len(expenses),
        'gross_profit_margin': (net_income / total_revenue * 100) if total_revenue > 0 else 0,
        'total_transactions': Sale.objects.filter(
            status='confirmed',
            created_at__date__gte=start_date_obj,
            created_at__date__lte=end_date_obj
        ).count() + Purchase.objects.filter(
            status='confirmed',
            created_at__date__gte=start_date_obj,
            created_at__date__lte=end_date_obj
        ).count()
    }
    
    context = {
        'revenues': revenues,
        'expenses': expenses,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_income': net_income,
        'start_date': start_date,
        'end_date': end_date,
        'stats': stats,
        'currency_symbol': get_setting('currency_symbol', 'ر.س'),
        **get_user_context(request)
    }
    
    return render(request, 'income_statement.html', context)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, F, Sum, Count
from django.db import models
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from functools import wraps
from datetime import date, timedelta
import json
import time
import os
from django.conf import settings
from .models import *
from .models import Salary
from .models import POSSession, POSSale, POSSaleItem
from .decorators import permission_required, subscription_required, branch_required, warehouse_required, company_required
from .permission_decorators import enhanced_permission_required
from .permissions_utils import check_user_permission
from .permissions_system import PermissionSystem
from .models import Permission
from .dynamic_settings import DynamicSettingsManager, get_dynamic_setting, set_dynamic_setting
from .inventory_accounting import InventoryAccountingManager

# دوال مساعدة
def get_user_context(request):
    """الحصول على سياق المستخدم مع الصلاحيات"""
    context = {}
    
    # إضافة بيانات الشركة والفرع
    if hasattr(request, 'company'):
        context['company'] = request.company
        context['company_name'] = request.company.name
    if hasattr(request, 'current_branch'):
        context['current_branch'] = request.current_branch
    if hasattr(request, 'current_warehouse'):
        context['current_warehouse'] = request.current_warehouse
    if hasattr(request, 'user_profile'):
        context['user_profile'] = request.user_profile
    
    # إضافة الصلاحيات للسياق
    if request.user.is_authenticated:
        context['user_permissions'] = get_user_permissions(request.user)
    
    # إضافة الإعدادات العامة
    context['current_settings'] = {
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        'company_name': context.get('company_name', 'شركة ERP'),
        'theme_color': get_setting('theme_color', 'blue')
    }
    
    return context

def get_user_permissions(user):
    """الحصول على صلاحيات المستخدم"""
    if user.is_superuser:
        return {'all': True}
    
    # قائمة الشاشات المتاحة
    screens = [
        'dashboard', 'products', 'customers', 'suppliers', 'sales_reps', 'sales', 'purchases',
        'stock', 'accounts', 'reports', 'settings', 'users', 'permissions',
        'companies', 'branches', 'warehouses', 'pos', 'manufacturing',
        'attendance', 'salaries'
    ]
    
    # قائمة الإجراءات المتاحة
    actions = ['view', 'add', 'edit', 'delete', 'confirm', 'print', 'export']
    
    permissions = {}
    for screen in screens:
        permissions[screen] = {}
        for action in actions:
            permissions[screen][action] = user.is_superuser
    
    return permissions

def login_view(request):
    try:
        if request.user.is_authenticated:
            return redirect('dashboard')
    except:
        pass
        
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        company_code = request.POST.get('company_code', 'DEFAULT').strip()
        
        if not username or not password or not company_code:
            companies = get_companies_for_login()
            return render(request, 'login.html', {
                'error': 'يرجى ملء جميع الحقول',
                'companies': companies
            })
        
        # تبديل قاعدة البيانات حسب الشركة
        if not switch_to_company_database(company_code):
            companies = get_companies_for_login()
            return render(request, 'login.html', {
                'error': 'شركة غير موجودة أو غير نشطة',
                'companies': companies
            })
        
        # الحصول على بيانات الشركة من قاعدة البيانات الجديدة
        try:
            company = Company.objects.get(code=company_code, is_active=True)
        except Exception as e:
            # إذا لم توجد الجداول، قم بتشغيل migrations
            if 'no such table' in str(e).lower():
                try:
                    from django.core.management import call_command
                    call_command('migrate', verbosity=0, interactive=False)
                    # إنشاء شركة افتراضية
                    from datetime import date, timedelta
                    company = Company.objects.create(
                        code=company_code,
                        name=f'شركة {company_code}',
                        database_name=f'erp_{company_code.lower()}.db',
                        is_active=True,
                        subscription_end=date.today() + timedelta(days=365)
                    )
                except:
                    companies = get_companies_for_login()
                    return render(request, 'login.html', {
                        'error': 'خطأ في إعداد قاعدة البيانات',
                        'companies': companies
                    })
            else:
                companies = get_companies_for_login()
                return render(request, 'login.html', {
                    'error': 'بيانات الشركة غير صحيحة',
                    'companies': companies
                })
        
        # محاولة تسجيل الدخول
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            # الحصول على أو إنشاء ملف المستخدم
            try:
                profile = UserProfile.objects.get(user=user, is_active=True)
                # إذا كان المستخدم مربوط بشركة أخرى، تحديث الشركة
                if profile.company.code != company_code:
                    profile.company = company
                    profile.save()
            except UserProfile.DoesNotExist:
                profile = UserProfile.objects.create(
                    user=user,
                    company=company,
                    is_active=True
                )
            
            login(request, user)
            request.session['company_id'] = company.id
            request.session['company_code'] = company.code
            request.session['database_name'] = company.database_name
            
            # تبديل إلى قاعدة بيانات الشركة
            switch_to_company_database(company_code)
            
            return redirect('dashboard')
        else:
            companies = get_companies_for_login()
            error_msg = 'حسابك غير نشط' if user and not user.is_active else 'اسم المستخدم أو كلمة المرور غير صحيح'
            return render(request, 'login.html', {
                'error': error_msg,
                'companies': companies
            })
    
    # جلب قائمة الشركات
    companies = get_companies_for_login()
    return render(request, 'login.html', {'companies': companies})

def switch_to_company_database(company_code):
    """تبديل قاعدة البيانات حسب الشركة"""
    try:
        db_name = f"erp_{company_code.lower()}.db"
        db_path = os.path.join(settings.BASE_DIR, 'databases', db_name)
        
        if os.path.exists(db_path):
            settings.DATABASES['default']['NAME'] = db_path
            from django.db import connection
            connection.close()
            return True
        return False
    except:
        return False

def get_companies_for_login():
    """الحصول على قائمة الشركات لتسجيل الدخول"""
    try:
        # جلب الشركات من قاعدة البيانات الحالية
        companies = Company.objects.filter(is_active=True).order_by('name')
        return [{'code': company.code, 'name': company.name} for company in companies]
    except:
        # إذا لم توجد شركات، أرجع قائمة فارغة
        return []

def logout_view(request):
    logout(request)
    request.session.flush()
    return redirect('login')

@login_required
@subscription_required
def dashboard(request):
    try:
        # جمع الإحصائيات الأساسية
        total_products = Product.objects.count()
        total_customers = Customer.objects.count()
        recent_sales = Sale.objects.select_related('customer').order_by('-created_at')[:5]
        
        # إحصائيات المبيعات
        try:
            total_sales = Sale.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        except:
            total_sales = 0
        
        # منتجات بمخزون منخفض (مع معالجة الأخطاء)
        low_stock_items = []
        low_stock_count = 0
        try:
            low_stock_items = ProductStock.objects.filter(
                current_stock__lte=F('min_stock')
            ).select_related('product')[:5]
            low_stock_count = low_stock_items.count()
        except:
            # إذا لم يوجد جدول ProductStock، استخدم المنتجات مباشرة
            try:
                low_stock_items = Product.objects.filter(
                    stock__lte=10  # الحد الأدنى الافتراضي
                )[:5]
                low_stock_count = low_stock_items.count()
            except:
                low_stock_items = []
                low_stock_count = 0
        
        # فواتير غير مؤكدة
        pending_invoices = 0
        try:
            pending_invoices = Sale.objects.filter(status='draft').count()
        except:
            pass
        
        # إشعارات ذكية
        notifications = []
        if low_stock_count > 0:
            notifications.append({
                'type': 'warning',
                'message': f'يوجد {low_stock_count} منتج بمخزون منخفض',
                'action': 'stock'
            })
        
        if pending_invoices > 0:
            notifications.append({
                'type': 'info', 
                'message': f'يوجد {pending_invoices} فاتورة في انتظار التأكيد',
                'action': 'sales'
            })
        
        # الحصول على الإعدادات (مع قيم افتراضية)
        try:
            company_name = get_setting('company_name', 'شركة ERP')
            currency_symbol = get_setting('currency_symbol', 'ر.س')
            theme_color = get_setting('theme_color', 'blue')
        except:
            company_name = 'شركة ERP'
            currency_symbol = 'ر.س'
            theme_color = 'blue'
        
        context = {
            'total_products': total_products,
            'total_customers': total_customers,
            'recent_sales': recent_sales,
            'total_sales': total_sales,
            'low_stock_products': low_stock_items,  # تم تغيير هذا من low_stock_count إلى low_stock_items
            'low_stock_count': low_stock_count,
            'low_stock_items': low_stock_items,
            'notifications': notifications,
            'company_name': company_name,
            'currency_symbol': currency_symbol,
            'theme_color': theme_color,
            **get_user_context(request)
        }
        return render(request, 'dashboard.html', context)
        
    except Exception as e:
        # في حالة أي خطأ، عرض لوحة تحكم بسيطة
        context = {
            'total_products': 0,
            'total_customers': 0,
            'recent_sales': [],
            'total_sales': 0,
            'low_stock_products': [],  # تم تغيير هذا من 0 إلى []
            'low_stock_count': 0,
            'low_stock_items': [],
            'notifications': [{
                'type': 'info',
                'message': 'مرحباً بك في نظام ERP - يرجى تشغيل migrations أولاً',
                'action': 'dashboard'
            }],
            'company_name': 'شركة ERP',
            'currency_symbol': 'ر.س',
            'theme_color': 'blue'
        }
        return render(request, 'dashboard.html', context)

@login_required
@subscription_required
@permission_required('products', 'view')
def products(request):
    products = Product.objects.all().order_by('-created_at')
    
    # فلترة بالبحث
    search = request.GET.get('search', '')
    if search:
        products = products.filter(
            Q(name__icontains=search) | 
            Q(barcode__icontains=search) |
            Q(category__icontains=search)
        )
    
    context = {
        'products': products,
        'search': search,
        **get_user_context(request)
    }
    return render(request, 'products.html', context)

@login_required
@subscription_required
@permission_required('products', 'add')
def add_product(request):
    if request.method == 'POST':
        try:
            # استخدام شركة المستخدم الحالية
            company = getattr(request, 'company', None)
            if not company:
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
            
            # إنشاء المنتج مع جميع البيانات
            product = Product.objects.create(
                company=company,
                name=request.POST.get('name'),
                barcode=request.POST.get('barcode') or f"PRD{int(time.time())}",
                category=request.POST.get('category'),
                unit=request.POST.get('unit', 'قطعة'),
                description=request.POST.get('description', ''),
                price=float(request.POST.get('price', 0)),
                stock=float(request.POST.get('initial_stock', 0)),
                brand=request.POST.get('brand', ''),
                is_active=request.POST.get('is_active') == 'on',
                bom=request.POST.get('bom') == 'on',
                image=request.FILES.get('image'),
                created_by=request.user
            )
            
            # إنشاء سعر ومخزون في جداول منفصلة إذا وجدت
            try:
                if hasattr(request, 'current_warehouse') and request.current_warehouse:
                    ProductPrice.objects.get_or_create(
                        product=product,
                        warehouse=request.current_warehouse,
                        defaults={
                            'cost_price': request.POST.get('cost_price', 0),
                            'selling_price': request.POST.get('selling_price', product.price or 0)
                        }
                    )
                    
                    ProductStock.objects.get_or_create(
                        product=product,
                        warehouse=request.current_warehouse,
                        defaults={
                            'current_stock': request.POST.get('initial_stock', product.stock or 0),
                            'min_stock': request.POST.get('min_stock', 10)
                        }
                    )
            except:
                pass
            
            messages.success(request, 'تم إضافة المنتج بنجاح')
            
            # إرجاع JSON للتحديث اللحظي
            if request.headers.get('Content-Type') == 'application/json' or request.headers.get('Accept') == 'application/json':
                return JsonResponse({'success': True, 'message': 'تم إضافة المنتج بنجاح', 'product_id': product.id})
            
            return redirect('products')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': str(e)})
    
    context = get_user_context(request)
    return render(request, 'add_product.html', context)

@login_required
@subscription_required
@permission_required('products', 'edit')
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            product.name = request.POST.get('name')
            product.barcode = request.POST.get('barcode')
            product.category = request.POST.get('category')
            product.unit = request.POST.get('unit')
            product.description = request.POST.get('description')
            if request.FILES.get('image'):
                product.image = request.FILES.get('image')
            product.save()
            
            # تحديث السعر والمخزون إذا تم إرسالهما
            if hasattr(request, 'current_warehouse') and request.current_warehouse:
                price, created = ProductPrice.objects.get_or_create(
                    product=product,
                    branch=getattr(request, 'current_branch', None),
                    warehouse=request.current_warehouse,
                    defaults={'cost_price': 0, 'selling_price': 0}
                )
                if request.POST.get('cost_price'):
                    price.cost_price = request.POST.get('cost_price')
                if request.POST.get('selling_price'):
                    price.selling_price = request.POST.get('selling_price')
                price.save()
            
            messages.success(request, 'تم تحديث المنتج بنجاح')
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'تم تحديث المنتج بنجاح'})
            
            return redirect('products')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': str(e)})
    
    context = {
        'product': product,
        **get_user_context(request)
    }
    return render(request, 'edit_product.html', context)

@login_required
@subscription_required
@permission_required('products', 'delete')
@csrf_exempt
def delete_product(request, product_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        product = get_object_or_404(Product, id=product_id)
        product_name = product.name
        
        try:
            ProductPrice.objects.filter(product=product).delete()
        except:
            pass
        try:
            ProductStock.objects.filter(product=product).delete()
        except:
            pass
        try:
            StockMovement.objects.filter(product=product).delete()
        except:
            pass
        
        product.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف المنتج "{product_name}" بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@permission_required('customers', 'view')
def customers_list(request):
    customers = Customer.objects.all().order_by('-created_at')
    
    # فلترة بالبحث
    search = request.GET.get('search', '')
    if search:
        customers = customers.filter(
            Q(name__icontains=search) | 
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )
    
    # إضافة البيانات الإضافية لكل عميل
    customers_with_data = []
    for customer in customers:
        # حساب عدد الفواتير
        total_invoices = Sale.objects.filter(customer=customer).count()
        
        # حساب المرتجعات (افتراضي 0 حتى يتم إنشاء نموذج المرتجعات)
        total_returns = 0
        
        # حساب المبلغ المستحق
        confirmed_sales = Sale.objects.filter(customer=customer, status='confirmed')
        total_sales = confirmed_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        paid_amount = confirmed_sales.aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0
        due_amount = total_sales - paid_amount
        
        # حساب الدفعات
        customer_payments = CustomerPayment.objects.filter(customer=customer).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # إضافة البيانات للعميل
        customer.total_invoices = total_invoices
        customer.total_returns = total_returns
        customer.due_amount = due_amount - customer_payments
        customer.total_payments = customer_payments
        
        customers_with_data.append(customer)
    
    context = {
        'customers': customers_with_data,
        'search': search,
        **get_user_context(request)
    }
    context['customer_payments'] = True  # لإظهار رابط الدفعات
    return render(request, 'customers.html', context)

@login_required
@subscription_required
@permission_required('customers', 'view')
def view_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    
    # الحصول على فواتير العميل
    sales = Sale.objects.filter(customer=customer).order_by('-created_at')[:10]
    
    # حساب الإحصائيات
    total_sales = Sale.objects.filter(customer=customer, status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    pending_invoices = Sale.objects.filter(customer=customer, status='draft').count()
    
    context = {
        'customer': customer,
        'sales': sales,
        'total_sales': total_sales,
        'pending_invoices': pending_invoices,
        **get_user_context(request)
    }
    return render(request, 'view_customer.html', context)

@login_required
@subscription_required
@permission_required('customers', 'edit')
def edit_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        try:
            customer.name = request.POST.get('name')
            customer.phone = request.POST.get('phone')
            customer.email = request.POST.get('email')
            customer.address = request.POST.get('address')
            customer.credit_limit = request.POST.get('credit_limit', 0)
            customer.opening_balance = request.POST.get('opening_balance', 0)
            customer.is_active = request.POST.get('is_active') == 'on'
            customer.save()
            
            messages.success(request, 'تم تحديث بيانات العميل بنجاح')
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'تم تحديث بيانات العميل بنجاح'})
            
            return redirect('customers')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': str(e)})
    
    context = {
        'customer': customer,
        **get_user_context(request)
    }
    return render(request, 'edit_customer.html', context)

@login_required
@subscription_required
@permission_required('customers', 'delete')
@csrf_exempt
def delete_customer(request, customer_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        customer_name = customer.name
        
        # التحقق من وجود فواتير مرتبطة
        if Sale.objects.filter(customer=customer).exists():
            return JsonResponse({
                'success': False, 
                'error': 'لا يمكن حذف العميل لوجود فواتير مرتبطة به'
            })
        
        customer.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف العميل "{customer_name}" بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@permission_required('customers', 'export')
def export_customers(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "العملاء"
        
        headers = ['الرقم', 'اسم العميل', 'الهاتف', 'البريد الإلكتروني', 'العنوان', 'الحد الائتماني', 'الرصيد الافتتاحي', 'الحالة']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        customers = Customer.objects.all()
        for row, customer in enumerate(customers, 2):
            ws.cell(row=row, column=1, value=customer.id)
            ws.cell(row=row, column=2, value=customer.name)
            ws.cell(row=row, column=3, value=customer.phone)
            ws.cell(row=row, column=4, value=customer.email or '')
            ws.cell(row=row, column=5, value=customer.address or '')
            ws.cell(row=row, column=6, value=float(customer.credit_limit))
            ws.cell(row=row, column=7, value=float(customer.opening_balance))
            ws.cell(row=row, column=8, value='نشط' if customer.is_active else 'غير نشط')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('customers')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('customers')

@login_required
@subscription_required
@permission_required('customers', 'add')
def import_customers(request):
    """دالة استيراد العملاء من ملف Excel"""
    if request.method == 'POST':
        try:
            import openpyxl
            file = request.FILES.get('file')
            if not file:
                messages.error(request, 'يرجى اختيار ملف')
                return redirect('import_customers')
            
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            imported_count = 0
            errors = []
            
            for row in range(2, ws.max_row + 1):
                try:
                    name = ws.cell(row=row, column=2).value
                    phone = ws.cell(row=row, column=3).value
                    email = ws.cell(row=row, column=4).value
                    address = ws.cell(row=row, column=5).value
                    credit_limit = ws.cell(row=row, column=6).value or 0
                    opening_balance = ws.cell(row=row, column=7).value or 0
                    
                    if not name or not phone:
                        errors.append(f'الصف {row}: اسم العميل ورقم الهاتف مطلوبان')
                        continue
                    
                    Customer.objects.create(
                        name=str(name).strip(),
                        phone=str(phone).strip(),
                        email=str(email).strip() if email else '',
                        address=str(address).strip() if address else '',
                        credit_limit=float(credit_limit),
                        opening_balance=float(opening_balance)
                    )
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f'الصف {row}: {str(e)}')
            
            if imported_count > 0:
                messages.success(request, f'تم استيراد {imported_count} عميل بنجاح')
            
            if errors:
                error_msg = 'أخطاء في الاستيراد:\n' + '\n'.join(errors[:5])
                messages.warning(request, error_msg)
            
            return redirect('customers')
            
        except ImportError:
            messages.error(request, 'مكتبة openpyxl غير مثبتة')
        except Exception as e:
            messages.error(request, f'خطأ في الاستيراد: {str(e)}')
    
    context = get_user_context(request)
    return render(request, 'import_customers.html', context)

@login_required
@subscription_required
@permission_required('customers', 'add')
def add_customer(request):
    if request.method == 'POST':
        try:
            # الحصول على الشركة
            company = getattr(request, 'company', None)
            if not company:
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
            
            # إنشاء العميل مع جميع الحقول
            customer = Customer.objects.create(
                company=company,
                name=request.POST.get('name'),
                phone=request.POST.get('phone'),
                email=request.POST.get('email', ''),
                address=request.POST.get('address', ''),
                credit_limit=float(request.POST.get('credit_limit', 0)),
                opening_balance=float(request.POST.get('opening_balance', 0)),
                is_active=request.POST.get('is_active') == 'on'
            )
            
            # حفظ الحقول الإضافية في جدول منفصل إذا لزم الأمر
            try:
                # يمكن إضافة جدول CustomerDetails لاحقاً
                customer.tax_number = request.POST.get('tax_number', '')
                customer.commercial_register = request.POST.get('commercial_register', '')
                customer.contact_person = request.POST.get('contact_person', '')
                customer.payment_terms = request.POST.get('payment_terms', '')
                customer.notes = request.POST.get('notes', '')
                customer.save()
            except:
                # إذا لم توجد هذه الحقول في النموذج، تجاهلها
                pass
            
            messages.success(request, f'تم إضافة العميل "{customer.name}" بنجاح')
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'تم إضافة العميل بنجاح'})
            
            return redirect('customers')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': str(e)})
    
    context = {
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        **get_user_context(request)
    }
    return render(request, 'add_customer.html', context)

@login_required
@subscription_required
@permission_required('suppliers', 'view')
def suppliers_list(request):
    suppliers = Supplier.objects.all().order_by('-created_at')
    search = request.GET.get('search', '')
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )
    context = {
        'suppliers': suppliers,
        'search': search,
        **get_user_context(request)
    }
    return render(request, 'suppliers.html', context)

@login_required
@subscription_required
@permission_required('suppliers', 'view')
def view_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    context = {
        'supplier': supplier,
        **get_user_context(request)
    }
    return render(request, 'view_supplier.html', context)

@login_required
@subscription_required
@permission_required('suppliers', 'edit')
def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    if request.method == 'POST':
        try:
            supplier.name = request.POST.get('name')
            supplier.phone = request.POST.get('phone')
            supplier.email = request.POST.get('email')
            supplier.address = request.POST.get('address')
            supplier.opening_balance = request.POST.get('opening_balance', 0)
            supplier.save()
            messages.success(request, 'تم تحديث المورد بنجاح')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'تم تحديث المورد بنجاح'})
            return redirect('suppliers')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': str(e)})
    context = {
        'supplier': supplier,
        **get_user_context(request)
    }
    return render(request, 'edit_supplier.html', context)

@login_required
@subscription_required
@permission_required('suppliers', 'export')
def export_suppliers(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الموردين"
        
        headers = ['الرقم', 'اسم المورد', 'الهاتف', 'البريد الإلكتروني', 'العنوان', 'الرصيد الافتتاحي']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        suppliers = Supplier.objects.all()
        for row, supplier in enumerate(suppliers, 2):
            ws.cell(row=row, column=1, value=supplier.id)
            ws.cell(row=row, column=2, value=supplier.name)
            ws.cell(row=row, column=3, value=supplier.phone)
            ws.cell(row=row, column=4, value=supplier.email)
            ws.cell(row=row, column=5, value=supplier.address)
            ws.cell(row=row, column=6, value=float(supplier.opening_balance))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="suppliers.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('suppliers')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('suppliers')

@login_required
@subscription_required
@permission_required('suppliers', 'add')
def add_supplier(request):
    if request.method == 'POST':
        try:
            supplier = Supplier.objects.create(
                name=request.POST.get('name'),
                phone=request.POST.get('phone'),
                email=request.POST.get('email'),
                address=request.POST.get('address'),
                opening_balance=request.POST.get('opening_balance', 0)
            )
            messages.success(request, 'تم إضافة المورد بنجاح')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'تم إضافة المورد بنجاح'})
            return redirect('suppliers')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': str(e)})
    context = get_user_context(request)
    return render(request, 'add_supplier.html', context)

@login_required
@subscription_required
@permission_required('sales', 'view')
def invoices_management(request):
    # جلب جميع المبيعات مع العلاقات
    sales = Sale.objects.select_related(
        'customer', 'created_by', 'branch', 'warehouse', 'sales_rep', 'confirmed_by'
    ).prefetch_related('items__product').all().order_by('-created_at')
    
    # فلترة حسب الحالة
    status = request.GET.get('status', '')
    if status:
        sales = sales.filter(status=status)
    
    # فلترة بالبحث
    search = request.GET.get('search', '')
    if search:
        sales = sales.filter(
            Q(invoice_number__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(created_by__username__icontains=search)
        )
    
    # فلترة بالتاريخ
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        sales = sales.filter(created_at__date__gte=start_date)
    if end_date:
        sales = sales.filter(created_at__date__lte=end_date)
    
    # حساب الإحصائيات
    total_invoices = sales.count()
    confirmed_invoices = sales.filter(status='confirmed').count()
    total_sales = sales.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_paid = sales.aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0
    total_due = total_sales - total_paid
    avg_invoice = total_sales / confirmed_invoices if confirmed_invoices > 0 else 0
    
    # إضافة بيانات إضافية لكل فاتورة
    for sale in sales:
        sale.remaining_amount = (sale.total_amount or 0) - (sale.paid_amount or 0)
        sale.items_count = sale.items.count()
        
        # طريقة الدفع
        if hasattr(sale, 'payment_method') and sale.payment_method:
            sale.payment_method_display = dict([
                ('cash', 'نقدي'),
                ('credit', 'آجل'),
                ('mixed', 'مختلط')
            ]).get(sale.payment_method, sale.payment_method)
        else:
            sale.payment_method_display = 'نقدي'
        
        # اسم المخزن
        sale.warehouse_name = sale.warehouse.name if sale.warehouse else 'غير محدد'
        
        # مندوب المبيعات
        if sale.sales_rep:
            sale.sales_rep_name = sale.sales_rep.get_full_name() or sale.sales_rep.username
        else:
            sale.sales_rep_name = '-'
    
    context = {
        'sales': sales,
        'search': search,
        'status': status,
        'start_date': start_date,
        'end_date': end_date,
        'status_choices': Sale.INVOICE_STATUS,
        'stats': {
            'total_invoices': total_invoices,
            'confirmed_invoices': confirmed_invoices,
            'total_sales': total_sales,
            'total_paid': total_paid,
            'total_due': total_due,
            'avg_invoice': avg_invoice,
        },
        **get_user_context(request)
    }
    return render(request, 'invoices.html', context)

@login_required
@subscription_required
@permission_required('sales', 'view')
def sales_list(request):
    return invoices_management(request)

@login_required
@subscription_required
def pos(request):
    try:
        # POS Sessions
        try:
            sessions = POSSession.objects.all().order_by('-opened_at')[:10]
            active_sessions = POSSession.objects.filter(status='open')
        except:
            sessions = []
            active_sessions = []
        
        # Recent sales
        try:
            recent_sales = Sale.objects.select_related('customer').order_by('-created_at')[:5]
        except:
            recent_sales = []
        
        # Recent purchases
        try:
            recent_purchases = Purchase.objects.select_related('supplier').order_by('-created_at')[:5]
        except:
            recent_purchases = []
        
        # Recent stock movements
        try:
            recent_stock_movements = StockMovement.objects.select_related('product').order_by('-created_at')[:5]
        except:
            recent_stock_movements = []
        
        # Today stats
        from datetime import date
        today = date.today()
        try:
            today_sales = Sale.objects.filter(created_at__date=today, status='confirmed')
            today_pos_sales = POSSale.objects.filter(created_at__date=today)
        except:
            today_sales = Sale.objects.none()
            today_pos_sales = []
        
        try:
            stats = {
                'active_sessions': len(active_sessions) if active_sessions else 0,
                'today_sessions': POSSession.objects.filter(opened_at__date=today).count() if POSSession.objects.exists() else 0,
                'today_sales_count': today_sales.count() + len(today_pos_sales),
                'today_sales_total': (today_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0),
                'today_cash_total': 0,
                'today_card_total': 0,
            }
        except:
            stats = {
                'active_sessions': 0,
                'today_sessions': 0,
                'today_sales_count': 0,
                'today_sales_total': 0,
                'today_cash_total': 0,
                'today_card_total': 0,
            }
        
        # Check if user can open new session
        try:
            user_active_session = POSSession.objects.filter(cashier=request.user, status='open').first()
            can_open_session = not user_active_session
        except:
            can_open_session = True
        
    except Exception as e:
        # Default values on error
        sessions = []
        active_sessions = []
        recent_sales = []
        recent_purchases = []
        recent_stock_movements = []
        stats = {
            'active_sessions': 0,
            'today_sessions': 0,
            'today_sales_count': 0,
            'today_sales_total': 0,
            'today_cash_total': 0,
            'today_card_total': 0,
        }
        can_open_session = True
    
    context = {
        'sessions': sessions,
        'active_sessions': active_sessions,
        'stats': stats,
        'recent_sales': recent_sales,
        'recent_purchases': recent_purchases,
        'recent_stock_movements': recent_stock_movements,
        'can_open_session': can_open_session,
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        **get_user_context(request)
    }
    return render(request, 'pos.html', context)

@login_required
@subscription_required
def pos_open_session(request):
    """فتح جلسة نقاط بيع جديدة"""
    # التحقق من عدم وجود جلسة مفتوحة للمستخدم
    active_session = POSSession.objects.filter(cashier=request.user, status='open').first()
    if active_session:
        messages.error(request, 'لديك جلسة مفتوحة بالفعل')
        return redirect('pos')
    
    if request.method == 'POST':
        try:
            opening_balance = float(request.POST.get('opening_balance', 0))
            branch_id = request.POST.get('branch_id')
            warehouse_id = request.POST.get('warehouse_id')
            notes = request.POST.get('notes', '')
            
            if opening_balance < 0:
                messages.error(request, 'المبلغ الافتتاحي لا يمكن أن يكون سالباً')
                return redirect('pos_open_session')
            
            # الحصول على الشركة الافتراضية
            company = getattr(request, 'company', None)
            if not company:
                try:
                    company = Company.objects.first()
                    if not company:
                        company = Company.objects.create(
                            code='DEFAULT',
                            name='الشركة الافتراضية',
                            database_name='erp_default'
                        )
                except:
                    company = None
            
            # الحصول على الفرع الافتراضي
            if not branch_id:
                try:
                    default_branch = Branch.objects.first()
                    branch_id = default_branch.id if default_branch else None
                except:
                    branch_id = None
            
            if not warehouse_id:
                try:
                    default_warehouse = Warehouse.objects.first()
                    warehouse_id = default_warehouse.id if default_warehouse else None
                except:
                    warehouse_id = None
            
            # إنشاء الجلسة
            session = POSSession.objects.create(
                company=company,
                cashier=request.user,
                branch_id=branch_id,
                warehouse_id=warehouse_id,
                opening_balance=opening_balance,
                notes=notes
            )
            
            messages.success(request, f'تم فتح الجلسة #{session.session_number} بنجاح')
            return redirect('pos')
            
        except Exception as e:
            messages.error(request, f'خطأ في فتح الجلسة: {str(e)}')
    
    context = {
        'branches': Branch.objects.all(),
        'warehouses': Warehouse.objects.all(),
        'current_branch': getattr(request, 'current_branch', None),
        'current_warehouse': getattr(request, 'current_warehouse', None),
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        **get_user_context(request)
    }
    return render(request, 'pos_open_session.html', context)

@login_required
@subscription_required
def pos_close_session(request, session_id):
    """إغلاق جلسة نقاط البيع"""
    session = get_object_or_404(POSSession, id=session_id, cashier=request.user, status='open')
    
    if request.method == 'POST':
        try:
            closing_balance = float(request.POST.get('closing_balance', 0))
            notes = request.POST.get('notes', '')
            
            # حساب إجماليات الجلسة
            pos_sales = POSSale.objects.filter(session=session)
            total_sales = pos_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            
            # حساب المبالغ النقدية والكي نت بشكل صحيح
            total_cash = 0
            total_card = 0
            for sale in pos_sales:
                if sale.payment_method == 'cash':
                    total_cash += sale.cash_amount or 0
                elif sale.payment_method == 'card':
                    total_card += sale.card_amount or 0
                elif sale.payment_method == 'mixed':
                    total_cash += sale.cash_amount or 0
                    total_card += sale.card_amount or 0
            
            # حساب الفرق
            expected_balance = float(session.opening_balance) + float(total_cash)
            difference = closing_balance - expected_balance
            
            # تحديث الجلسة
            session.closing_balance = closing_balance
            session.total_sales = total_sales
            session.total_cash = total_cash
            session.total_card = total_card
            session.status = 'closed'
            session.closed_at = timezone.now()
            session.notes = f"{notes}\nالفرق: {difference:.2f} د.ك" if notes else f"الفرق: {difference:.2f} د.ك"
            session.save()
            
            messages.success(request, f'تم إغلاق الجلسة #{session.session_number} بنجاح')
            return redirect('pos_session_report', session_id=session.id)
            
        except Exception as e:
            messages.error(request, f'خطأ في إغلاق الجلسة: {str(e)}')
    
    # حساب الإحصائيات للعرض
    pos_sales = POSSale.objects.filter(session=session)
    total_cash = 0
    total_card = 0
    
    # حساب المبالغ النقدية والكي نت
    for sale in pos_sales:
        if sale.payment_method == 'cash':
            total_cash += float(sale.cash_amount or 0)
        elif sale.payment_method == 'card':
            total_card += float(sale.card_amount or 0)
        elif sale.payment_method == 'mixed':
            total_cash += float(sale.cash_amount or 0)
            total_card += float(sale.card_amount or 0)
    
    stats = {
        'total_sales': pos_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_cash': total_cash,
        'total_card': total_card,
        'sales_count': pos_sales.count(),
        'expected_balance': float(session.opening_balance) + total_cash
    }
    
    context = {
        'session': session,
        'stats': stats,
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        **get_user_context(request)
    }
    return render(request, 'pos_close_session.html', context)

@login_required
@subscription_required
def pos_session_report(request, session_id):
    """تقرير جلسة نقاط البيع"""
    session = get_object_or_404(POSSession, id=session_id)
    
    # جلب مبيعات الجلسة
    sales = POSSale.objects.filter(session=session).order_by('-created_at')
    
    # إحصائيات المبيعات
    sales_count = sales.count()
    items_count = POSSaleItem.objects.filter(pos_sale__session=session).count()
    avg_sale = session.total_sales / sales_count if sales_count > 0 else 0
    
    # حساب مدة الجلسة
    if session.closed_at and session.opened_at:
        duration = session.closed_at - session.opened_at
        hours = duration.seconds // 3600
        minutes = (duration.seconds % 3600) // 60
        session_duration = f"{hours}:{minutes:02d}"
    else:
        session_duration = "--:--"
    
    # أكثر المنتجات مبيعاً
    top_products = POSSaleItem.objects.filter(
        pos_sale__session=session
    ).values(
        'product__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_sales=Sum('total_price')
    ).order_by('-total_sales')[:5]
    
    context = {
        'session': session,
        'sales': sales,
        'sales_count': sales_count,
        'items_count': items_count,
        'avg_sale': avg_sale,
        'session_duration': session_duration,
        'top_products': top_products,
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        'company_name': get_setting('company_name', 'شركة ERP'),
        **get_user_context(request)
    }
    return render(request, 'pos_session_report.html', context)

@login_required
@subscription_required
def pos_sale(request):
    """إجراء عملية بيع في نقاط البيع"""
    # التحقق من وجود جلسة مفتوحة
    active_session = POSSession.objects.filter(cashier=request.user, status='open').first()
    if not active_session:
        messages.error(request, 'يجب فتح جلسة أولاً')
        return redirect('pos_open_session')
    
    if request.method == 'POST':
        try:
            customer_name = request.POST.get('customer_name', 'عميل نقدي')
            payment_method = request.POST.get('payment_method', 'cash')
            cash_amount = float(request.POST.get('cash_amount', 0))
            card_amount = float(request.POST.get('card_amount', 0))
            discount_amount = float(request.POST.get('discount_amount', 0))
            
            # الحصول على الشركة الحالية
            company = getattr(request, 'company', None)
            if not company:
                company = getattr(active_session, 'company', None)
                if not company:
                    company, created = Company.objects.get_or_create(
                        code='DEFAULT',
                        defaults={
                            'name': 'الشركة الافتراضية',
                            'database_name': 'erp_default',
                            'subscription_end': date.today() + timedelta(days=365)
                        }
                    )
            
            # إنشاء البيع
            pos_sale = POSSale.objects.create(
                company=company,
                session=active_session,
                customer_name=customer_name,
                payment_method=payment_method,
                cash_amount=cash_amount,
                card_amount=card_amount,
                discount_amount=discount_amount,
                subtotal=0,
                total_amount=0,
                created_by=request.user
            )
            
            # إضافة العناصر
            items_data = json.loads(request.POST.get('items', '[]'))
            subtotal = 0
            
            for item_data in items_data:
                product = get_object_or_404(Product, id=item_data['product_id'])
                quantity = float(item_data['quantity'])
                unit_price = float(item_data['unit_price'])
                
                # خصم الكمية من المخزون
                if hasattr(product, 'stock'):
                    if product.stock < quantity:
                        messages.error(request, f'المخزون غير كافي للمنتج {product.name}')
                        pos_sale.delete()
                        return redirect('pos_sale')
                    
                    from decimal import Decimal
                    product.stock = Decimal(str(product.stock)) - Decimal(str(quantity))
                    product.save()
                    
                    # تسجيل حركة المخزون
                    StockMovement.objects.create(
                        company=company,
                        product=product,
                        movement_type='out',
                        quantity=quantity,
                        reference=f'بيع نقاط البيع #{pos_sale.receipt_number}',
                        created_by=request.user
                    )
                
                # إنشاء عنصر البيع
                item_total = quantity * unit_price
                POSSaleItem.objects.create(
                    company=company,
                    pos_sale=pos_sale,
                    product=product,
                    quantity=quantity,
                    unit_price=unit_price,
                    total_price=item_total
                )
                
                subtotal += item_total
            
            # حساب الإجمالي
            total_amount = subtotal - discount_amount
            change_amount = (cash_amount + card_amount) - total_amount
            
            pos_sale.subtotal = subtotal
            pos_sale.total_amount = total_amount
            pos_sale.change_amount = change_amount
            pos_sale.save()
            
            messages.success(request, f'تم إتمام البيع #{pos_sale.receipt_number} بنجاح')
            return redirect('pos')
            
        except Exception as e:
            messages.error(request, f'خطأ في البيع: {str(e)}')
    
    # جلب المنتجات النشطة
    products = Product.objects.filter(is_active=True).order_by('name')
    
    context = {
        'active_session': active_session,
        'products': products,
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        **get_user_context(request)
    }
    return render(request, 'pos_sale.html', context)

@login_required
@subscription_required
@permission_required('sales', 'add')
def add_sale(request):
    if request.method == 'POST':
        try:
            customer_id = request.POST.get('customer_id')
            if not customer_id:
                messages.error(request, 'يرجى اختيار عميل')
                return redirect('add_sale')
            
            # الحصول على الشركة
            company = getattr(request, 'company', None)
            if not company:
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
            
            # جلب البيانات الجديدة
            sales_rep_id = request.POST.get('sales_rep_id')
            warehouse_id = request.POST.get('warehouse_id')
            payment_method = request.POST.get('payment_method', 'cash')
            created_at = request.POST.get('created_at')
            due_date = request.POST.get('due_date')
            
            # تحويل التواريخ
            if created_at:
                from datetime import datetime
                created_at = datetime.fromisoformat(created_at.replace('T', ' '))
            else:
                created_at = timezone.now()
                
            if due_date:
                from datetime import datetime
                due_date = datetime.strptime(due_date, '%Y-%m-%d').date()
            
            sale = Sale.objects.create(
                company=company,
                customer_id=customer_id,
                sales_rep_id=sales_rep_id if sales_rep_id else None,
                warehouse_id=warehouse_id if warehouse_id else None,
                payment_method=payment_method,
                subtotal=float(request.POST.get('subtotal', 0)),
                discount_amount=float(request.POST.get('discount_amount', 0)),
                tax_amount=float(request.POST.get('tax_amount', 0)),
                total_amount=float(request.POST.get('total_amount', 0)),
                notes=request.POST.get('notes', ''),
                due_date=due_date,
                status='draft',
                is_confirmed=False,
                created_at=created_at,
                created_by=request.user
            )
            
            # إضافة عناصر الفاتورة
            items_json = request.POST.get('items', '[]')
            
            if not items_json or items_json == '[]':
                sale.delete()
                messages.error(request, 'يرجى إضافة عناصر للفاتورة')
                return redirect('add_sale')
            
            items_data = json.loads(items_json)
            if not items_data:
                sale.delete()
                messages.error(request, 'يرجى إضافة عناصر للفاتورة')
                return redirect('add_sale')
            
            for item_data in items_data:
                SaleItem.objects.create(
                    company=company,
                    sale=sale,
                    product_id=item_data['product_id'],
                    quantity=float(item_data['quantity']),
                    unit_price=float(item_data['unit_price']),
                    discount_percent=float(item_data.get('discount_percent', 0)),
                    tax_rate=float(item_data.get('tax_rate', 0))
                )
            
            messages.success(request, f'تم إنشاء الفاتورة #{sale.invoice_number} بنجاح وهي في حالة مسودة')
            return redirect('invoices_management')
            
        except Exception as e:
            print("Error:", str(e))
            messages.error(request, f'خطأ: {str(e)}')
            return redirect('add_sale')
    
    # تحضير بيانات المنتجات للـ DataGrid
    products_data = []
    for product in Product.objects.filter(is_active=True):
        try:
            # الحصول على السعر
            price = 0
            # محاولة الحصول على السعر من ProductPrice
            try:
                price_obj = ProductPrice.objects.filter(product=product).first()
                if price_obj and price_obj.selling_price:
                    price = float(price_obj.selling_price)
            except:
                pass
            
            # إذا لم يوجد سعر، جرب الحقل المباشر
            if price == 0:
                try:
                    if hasattr(product, 'price') and product.price:
                        price = float(product.price)
                except:
                    price = 0
            
            # الحصول على المخزون
            stock = 0
            try:
                stock_obj = ProductStock.objects.filter(product=product).first()
                if stock_obj:
                    stock = float(stock_obj.current_stock)
                elif hasattr(product, 'stock') and product.stock:
                    stock = float(product.stock)
            except:
                stock = 0
            
            # الحصول على سعر التكلفة
            cost_price = 0
            try:
                if price_obj and price_obj.cost_price:
                    cost_price = float(price_obj.cost_price)
                elif hasattr(product, 'cost_price') and product.cost_price:
                    cost_price = float(product.cost_price)
            except:
                cost_price = 0
            
            products_data.append({
                'id': product.id,
                'name': product.name,
                'barcode': product.barcode or '',
                'category': product.category or '',
                'unit': product.unit or 'قطعة',
                'price': price,
                'cost_price': cost_price,
                'stock': stock,
                'brand': getattr(product, 'brand', '') or '',
                'description': product.description or '',
                'image_url': product.image.url if product.image else '',
                'is_active': product.is_active,
                'bom': getattr(product, 'bom', False)
            })
        except Exception as e:
            continue
    
    from datetime import datetime
    
    # جلب مناديب المبيعات
    try:
        sales_reps = SalesRep.objects.filter(is_active=True).select_related('employee__user')
    except:
        sales_reps = []
    
    # جلب المخازن
    try:
        warehouses = Warehouse.objects.filter(is_active=True)
    except:
        warehouses = []
    
    context = {
        'customers': Customer.objects.all(),
        'sales_reps': sales_reps,
        'warehouses': warehouses,
        'products': json.dumps(products_data),
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        'today': date.today().strftime('%Y-%m-%d'),
        'current_time': datetime.now().strftime('%H:%M'),
        **get_user_context(request)
    }
    return render(request, 'add_invoice.html', context)

@login_required
@subscription_required
@permission_required('purchases', 'view')
def purchases_list(request):
    purchases = Purchase.objects.select_related('supplier', 'created_by').all().order_by('-created_at')
    status = request.GET.get('status', '')
    if status:
        purchases = purchases.filter(status=status)
    search = request.GET.get('search', '')
    if search:
        purchases = purchases.filter(
            Q(invoice_number__icontains=search) |
            Q(supplier__name__icontains=search)
        )
    
    # حساب الإحصائيات
    total_purchases = purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    confirmed_count = purchases.filter(status='confirmed').count()
    paid_count = purchases.filter(status='paid').count()
    
    context = {
        'purchases': purchases,
        'search': search,
        'status': status,
        'status_choices': Purchase.INVOICE_STATUS,
        'total_purchases': total_purchases,
        'confirmed_count': confirmed_count,
        'paid_count': paid_count,
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        'company_name': get_setting('company_name', 'شركة ERP'),
        **get_user_context(request)
    }
    return render(request, 'purchases.html', context)

@login_required
@subscription_required
@permission_required('purchases', 'add')
def add_purchase(request):
    if request.method == 'POST':
        try:
            purchase = Purchase.objects.create(
                supplier_id=request.POST.get('supplier_id'),
                subtotal=request.POST.get('subtotal', 0),
                discount_amount=request.POST.get('discount_amount', 0),
                tax_amount=request.POST.get('tax_amount', 0),
                total_amount=request.POST.get('total_amount', 0),
                notes=request.POST.get('notes'),
                due_date=request.POST.get('due_date') or None,
                created_by=request.user
            )
            items_data = json.loads(request.POST.get('items', '[]'))
            for item_data in items_data:
                PurchaseItem.objects.create(
                    purchase=purchase,
                    product_id=item_data['product_id'],
                    quantity=item_data['quantity'],
                    unit_price=item_data['unit_price'],
                    discount_percent=item_data.get('discount_percent', 0),
                    tax_rate=item_data.get('tax_rate', 0)
                )
            messages.success(request, f'تم إنشاء فاتورة الشراء #{purchase.invoice_number} بنجاح')
            return redirect('purchases')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    # تحضير بيانات المنتجات للـ DataGrid
    products_data = []
    for product in Product.objects.filter(is_active=True):
        # الحصول على السعر من ProductPrice أو Product
        price = 0
        try:
            product_price = ProductPrice.objects.filter(product=product).first()
            if product_price and product_price.cost_price:
                price = float(product_price.cost_price)
            elif hasattr(product, 'cost_price') and product.cost_price:
                price = float(product.cost_price)
            elif hasattr(product, 'price') and product.price:
                price = float(product.price)
        except:
            price = float(getattr(product, 'price', 0) or 0)
        
        products_data.append({
            'id': product.id,
            'name': product.name,
            'price': price
        })
    
    context = {
        'suppliers': Supplier.objects.all(),
        'products': json.dumps(products_data),
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        'company_name': get_setting('company_name', 'شركة ERP'),
        'default_tax_rate': get_setting('default_tax_rate', '15'),
        **get_user_context(request)
    }
    return render(request, 'add_purchase.html', context)

# تم نقل هذه الدالة إلى أعلى في الملف

# تم نقل هذه الدوال إلى أعلى في الملف

# تم نقل هذه الدوال إلى أعلى في الملف

# تم نقل هذه الدوال إلى أعلى في الملف

# تم نقل هذه الدوال إلى أعلى في الملف

# تم نقل هذه الدوال إلى أعلى في الملف

@login_required
@subscription_required
@permission_required('settings', 'view')
def settings_page(request):
    from .settings_helper import save_setting, get_setting as get_file_setting, save_all_settings, init_default_settings
    
    if request.method == 'POST':
        try:
            # حفظ الإعدادات في ملف JSON
            settings_to_save = {
                'company_name': request.POST.get('company_name', 'الفنار'),
                'company_code': request.POST.get('company_code', ''),
                'company_phone': request.POST.get('company_phone', ''),
                'company_email': request.POST.get('company_email', ''),
                'company_address': request.POST.get('company_address', ''),
                'currency_symbol': request.POST.get('currency_symbol', 'د.ك'),
                'default_tax_rate': request.POST.get('default_tax_rate', '15'),
                'decimal_places': request.POST.get('decimal_places', '2'),
                'allow_negative_stock': 'true' if request.POST.get('allow_negative_stock') == 'on' else 'false',
                'theme_color': request.POST.get('theme_color', 'blue'),
                'items_per_page': request.POST.get('items_per_page', '25'),
                'enable_notifications': 'true' if request.POST.get('enable_notifications') == 'on' else 'false',
                'enable_realtime': 'true' if request.POST.get('enable_realtime') == 'on' else 'false',
                'pos_receipt_width': request.POST.get('pos_receipt_width', '40'),
                'pos_auto_print': 'true' if request.POST.get('pos_auto_print') == 'on' else 'false',
                'pos_receipt_header': request.POST.get('pos_receipt_header', ''),
                'pos_receipt_footer': request.POST.get('pos_receipt_footer', 'شكراً لتعاملكم معنا'),
                'report_date_format': request.POST.get('report_date_format', 'd/m/Y'),
                'export_max_rows': request.POST.get('export_max_rows', '1000'),
                'show_company_logo': 'true' if request.POST.get('show_company_logo') == 'on' else 'false',
                'auto_backup_reports': 'true' if request.POST.get('auto_backup_reports') == 'on' else 'false',
                'session_timeout': request.POST.get('session_timeout', '30'),
                'backup_count': request.POST.get('backup_count', '5'),
                'debug_mode': 'true' if request.POST.get('debug_mode') == 'on' else 'false',
                'maintenance_mode': 'true' if request.POST.get('maintenance_mode') == 'on' else 'false',
            }
            
            # حفظ جميع الإعدادات
            if save_all_settings(settings_to_save):
                messages.success(request, 'تم حفظ الإعدادات بنجاح')
            else:
                messages.error(request, 'خطأ في حفظ الإعدادات')
            
            return redirect('settings')
            
        except Exception as e:
            messages.error(request, f'خطأ في حفظ الإعدادات: {str(e)}')
            return redirect('settings')
    
    # جلب الإعدادات من ملف JSON
    current_settings = init_default_settings()
    
    # تحويل القيم النصية إلى قيم مناسبة للعرض
    for key, value in current_settings.items():
        if key in ['default_tax_rate', 'decimal_places', 'items_per_page', 'pos_receipt_width', 'export_max_rows', 'session_timeout', 'backup_count']:
            try:
                current_settings[key] = int(value)
            except:
                current_settings[key] = 0
        elif key in ['allow_negative_stock', 'enable_notifications', 'enable_realtime', 'pos_auto_print', 'show_company_logo', 'auto_backup_reports', 'debug_mode', 'maintenance_mode']:
            current_settings[key] = str(value).lower() == 'true'
    
    context = {
        'current_settings': current_settings,
        **get_user_context(request)
    }
    return render(request, 'settings.html', context)

from functools import lru_cache

def get_setting(key, default=None, company=None):
    try:
        from .settings_helper import get_setting as get_file_setting
        return get_file_setting(key, default)
    except Exception as e:
        return default

def set_setting(key, value, setting_type='string', category='general'):
    """حفظ إعداد في قاعدة البيانات"""
    try:
        current_branch = getattr(request, 'current_branch', None)
        setting, created = Setting.objects.update_or_create(
            key=key,
            branch=current_branch,
            defaults={
                'value': str(value),
                'setting_type': setting_type,
                'category': category
            }
        )
        
        # مسح الكاش
        get_setting.cache_clear()
        return setting
    except Exception as e:
        return None

# API البحث عن المنتجات
@login_required
@csrf_exempt
def search_products_api(request):
    """API للبحث عن المنتجات بالاسم، الكود، أو الباركود"""
    try:
        query = request.GET.get('q', '').strip()
        
        if not query:
            return JsonResponse({'success': False, 'message': 'يرجى إدخال نص البحث'})
        
        # البحث في المنتجات
        products = Product.objects.filter(
            Q(name__icontains=query) |
            Q(barcode__icontains=query) |
            Q(category__icontains=query),
            is_active=True
        ).order_by('name')[:20]
        
        products_data = []
        for product in products:
            try:
                # الحصول على السعر
                price = 0
                cost_price = 0
                try:
                    price_obj = ProductPrice.objects.filter(product=product).first()
                    if price_obj:
                        price = float(price_obj.selling_price or 0)
                        cost_price = float(price_obj.cost_price or 0)
                except:
                    pass
                
                if price == 0 and hasattr(product, 'price') and product.price:
                    price = float(product.price)
                if cost_price == 0 and hasattr(product, 'cost_price') and product.cost_price:
                    cost_price = float(product.cost_price)
                
                # الحصول على المخزون
                stock = 0
                try:
                    stock_obj = ProductStock.objects.filter(product=product).first()
                    if stock_obj:
                        stock = float(stock_obj.current_stock)
                    elif hasattr(product, 'stock') and product.stock:
                        stock = float(product.stock)
                except:
                    stock = 0
                
                products_data.append({
                    'id': product.id,
                    'name': product.name,
                    'barcode': product.barcode or '',
                    'category': product.category or '',
                    'unit': product.unit or 'قطعة',
                    'price': price,
                    'cost_price': cost_price,
                    'stock': stock,
                    'brand': getattr(product, 'brand', '') or '',
                    'description': product.description or '',
                    'image_url': product.image.url if product.image else ''
                })
            except Exception as e:
                continue
        
        return JsonResponse({
            'success': True,
            'products': products_data,
            'count': len(products_data)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def get_product_details_api(request, product_id):
    """API للحصول على تفاصيل منتج محدد"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        
        # جمع جميع البيانات
        price = 0
        cost_price = 0
        try:
            price_obj = ProductPrice.objects.filter(product=product).first()
            if price_obj:
                price = float(price_obj.selling_price or 0)
                cost_price = float(price_obj.cost_price or 0)
        except:
            pass
        
        if price == 0 and hasattr(product, 'price') and product.price:
            price = float(product.price)
        if cost_price == 0 and hasattr(product, 'cost_price') and product.cost_price:
            cost_price = float(product.cost_price)
        
        stock = 0
        try:
            stock_obj = ProductStock.objects.filter(product=product).first()
            if stock_obj:
                stock = float(stock_obj.current_stock)
            elif hasattr(product, 'stock') and product.stock:
                stock = float(product.stock)
        except:
            stock = 0
        
        product_data = {
            'id': product.id,
            'name': product.name,
            'barcode': product.barcode or '',
            'category': product.category or '',
            'unit': product.unit or 'قطعة',
            'price': price,
            'cost_price': cost_price,
            'stock': stock,
            'brand': getattr(product, 'brand', '') or '',
            'description': product.description or '',
            'image_url': product.image.url if product.image else '',
            'is_active': product.is_active,
            'created_at': product.created_at.strftime('%Y-%m-%d') if product.created_at else '',
            'bom': getattr(product, 'bom', False)
        }
        
        return JsonResponse({
            'success': True,
            'product': product_data
        })
        
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'المنتج غير موجود'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@csrf_exempt
def quick_add_product_api(request):
    """API لإضافة منتج سريع"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'طريقة غير مسموحة'})
    
    try:
        data = json.loads(request.body)
        
        name = data.get('name', '').strip()
        barcode = data.get('barcode', '').strip()
        price = float(data.get('price', 0))
        category = data.get('category', '').strip()
        unit = data.get('unit', 'قطعة').strip()
        
        if not name:
            return JsonResponse({'success': False, 'message': 'اسم المنتج مطلوب'})
        
        if barcode and Product.objects.filter(barcode=barcode).exists():
            return JsonResponse({'success': False, 'message': 'الباركود موجود بالفعل'})
        
        # الحصول على الشركة
        company = None
        try:
            if hasattr(request, 'company') and request.company:
                company = request.company
            else:
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
        except:
            company, created = Company.objects.get_or_create(
                code='DEFAULT',
                defaults={
                    'name': 'الشركة الافتراضية', 
                    'database_name': 'erp_default',
                    'subscription_end': date.today() + timedelta(days=365)
                }
            )
        
        # إنشاء المنتج
        product = Product.objects.create(
            company=company,
            name=name,
            barcode=barcode or f"PRD{int(time.time())}",
            category=category or 'أخرى',
            unit=unit,
            price=price,
            is_active=True,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'تم إضافة المنتج بنجاح',
            'product': {
                'id': product.id,
                'name': product.name,
                'barcode': product.barcode,
                'category': product.category,
                'unit': product.unit,
                'price': float(product.price),
                'stock': 0
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# API لتحديث الإعدادات فورياً
@login_required
@csrf_exempt
def update_setting_ajax(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        # دعم كل من JSON و form data
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
            
        key = data.get('key')
        value = data.get('value')
        setting_type = data.get('type', 'string')
        
        if not key:
            return JsonResponse({'success': False, 'error': 'مفتاح الإعداد مطلوب'})
        
        # تحديد الفئة حسب المفتاح
        category_mapping = {
            'company_name': 'الشركة',
            'company_code': 'الشركة', 
            'company_phone': 'الشركة',
            'company_email': 'الشركة',
            'company_address': 'الشركة',
            'currency_symbol': 'عملة',
            'default_tax_rate': 'فواتير',
            'decimal_places': 'عملة',
            'allow_negative_stock': 'المخزون',
            'theme_color': 'شاشات',
            'items_per_page': 'شاشات',
            'enable_notifications': 'شاشات',
            'enable_realtime': 'شاشات',
            'pos_receipt_width': 'نقاط البيع',
            'pos_auto_print': 'نقاط البيع',
            'pos_receipt_header': 'نقاط البيع',
            'pos_receipt_footer': 'نقاط البيع',
            'report_date_format': 'تقارير',
            'export_max_rows': 'تقارير',
            'show_company_logo': 'تقارير',
            'auto_backup_reports': 'تقارير',
            'session_timeout': 'أمان',
            'backup_count': 'نسخ احتياطي',
            'debug_mode': 'أمان',
            'maintenance_mode': 'أمان'
        }
        
        category = category_mapping.get(key, 'عامة')
        
        # الحصول على الشركة الحالية
        company = getattr(request, 'company', None)
        if not company:
            try:
                company = Company.objects.first()
            except:
                company = None
        
        # حفظ أو تحديث الإعداد
        setting, created = Setting.objects.update_or_create(
            key=key,
            company=company,
            branch=getattr(request, 'current_branch', None),
            defaults={
                'value': str(value), 
                'setting_type': setting_type, 
                'category': category,
                'created_by': request.user
            }
        )
        
        # مسح الكاش
        try:
            get_setting.cache_clear()
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'تم حفظ الإعداد {key} بنجاح',
            'created': created
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_setting_ajax(request, key):
    """جلب قيمة إعداد معين عبر AJAX"""
    try:
        value = DynamicSettingsManager.get(
            key=key,
            branch=getattr(request, 'current_branch', None),
            default=None
        )
        
        return JsonResponse({
            'success': True,
            'key': key,
            'value': value
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_all_settings_ajax(request):
    try:
        current_branch = getattr(request, 'current_branch', None)
        settings = {}
        
        settings_queryset = Setting.objects.filter(
            models.Q(branch=current_branch) | models.Q(branch__isnull=True)
        ).order_by('branch')
        
        for setting in settings_queryset:
            settings[setting.key] = setting.value
        
        return JsonResponse({'success': True, 'settings': settings})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_product_movements(request, product_id):
    """جلب حركات المخزون لمنتج معين"""
    try:
        product = get_object_or_404(Product, id=product_id)
        movements = StockMovement.objects.filter(product=product).order_by('-created_at')[:20]
        
        movements_data = []
        for movement in movements:
            movements_data.append({
                'movement_type': movement.movement_type,
                'quantity': float(movement.quantity),
                'reference': movement.reference,
                'created_at': movement.created_at.strftime('%Y/%m/%d %H:%M')
            })
        
        return JsonResponse({
            'success': True,
            'movements': movements_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@csrf_exempt
def adjust_stock_api(request):
    """تسوية المخزون عبر API"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'طريقة غير مسموحة'})
    
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        adjustment_type = data.get('adjustment_type')
        quantity = float(data.get('quantity', 0))
        reason = data.get('reason', '')
        
        product = get_object_or_404(Product, id=product_id)
        
        # تحديث المخزون حسب نوع التسوية
        current_stock = float(product.stock or 0)
        
        if adjustment_type == 'add':
            new_stock = current_stock + quantity
            movement_type = 'in'
        elif adjustment_type == 'subtract':
            new_stock = current_stock - quantity
            movement_type = 'out'
        elif adjustment_type == 'set':
            new_stock = quantity
            movement_type = 'adjustment'
        else:
            return JsonResponse({'success': False, 'message': 'نوع تسوية غير صحيح'})
        
        # التأكد من عدم السماح بالمخزون السالب
        if new_stock < 0:
            allow_negative = get_setting('allow_negative_stock', 'false')
            if allow_negative.lower() != 'true':
                return JsonResponse({'success': False, 'message': 'لا يمكن أن يكون المخزون سالباً'})
        
        # تحديث المخزون
        product.stock = new_stock
        product.save()
        
        # تسجيل حركة المخزون
        company = getattr(request, 'company', None) or Company.objects.first()
        StockMovement.objects.create(
            company=company,
            product=product,
            movement_type=movement_type,
            quantity=abs(quantity),
            reference=f'تسوية مخزون - {reason}' if reason else 'تسوية مخزون',
            notes=reason,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'تم تسوية المخزون بنجاح',
            'new_stock': float(new_stock)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def export_stock(request):
    """تصدير بيانات المخزون"""
    try:
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير المخزون"
        
        headers = ['اسم المنتج', 'الكود', 'الباركود', 'المخزون الحالي', 'الوحدة', 'الحد الأدنى', 'الحالة', 'قيمة المخزون']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # جلب بيانات المخزون
        products = Product.objects.all()
        for row, product in enumerate(products, 2):
            current_stock = float(product.stock or 0)
            min_stock = 10  # افتراضي
            cost_price = float(getattr(product, 'cost_price', 0) or 0)
            stock_value = current_stock * cost_price
            
            # تحديد الحالة
            if current_stock == 0:
                status = 'نفد المخزون'
            elif current_stock <= min_stock:
                status = 'مخزون منخفض'
            else:
                status = 'مخزون جيد'
            
            ws.cell(row=row, column=1, value=product.name)
            ws.cell(row=row, column=2, value=getattr(product, 'code', product.barcode))
            ws.cell(row=row, column=3, value=product.barcode or '')
            ws.cell(row=row, column=4, value=current_stock)
            ws.cell(row=row, column=5, value=product.unit or 'قطعة')
            ws.cell(row=row, column=6, value=min_stock)
            ws.cell(row=row, column=7, value=status)
            ws.cell(row=row, column=8, value=stock_value)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="stock_report.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('stock')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('stock')

@login_required
@subscription_required
@permission_required('products', 'add')
def import_products_page(request):
    if request.method == 'POST':
        try:
            import openpyxl
            file = request.FILES.get('excel_file')
            if not file:
                messages.error(request, 'يرجى اختيار ملف Excel')
                return redirect('import_products_page')
            
            # خيارات الاستيراد
            update_existing = request.POST.get('update_existing') == 'on'
            skip_errors = request.POST.get('skip_errors') == 'on'
            create_backup = request.POST.get('create_backup') == 'on'
            default_warehouse = request.POST.get('default_warehouse')
            default_category = request.POST.get('default_category')
            
            # إنشاء نسخة احتياطية إذا طُلب ذلك
            if create_backup:
                try:
                    from .backup_system import create_backup
                    create_backup('products_backup', include_media=False)
                except:
                    pass
            
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            success_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            for row in range(2, ws.max_row + 1):
                try:
                    # قراءة البيانات مع معالجة القيم الفارغة
                    name = ws.cell(row=row, column=1).value
                    code = ws.cell(row=row, column=2).value
                    barcode = ws.cell(row=row, column=3).value
                    price = ws.cell(row=row, column=4).value
                    description = ws.cell(row=row, column=5).value or ''
                    category = ws.cell(row=row, column=6).value or default_category or ''
                    unit = ws.cell(row=row, column=7).value or 'قطعة'
                    initial_stock = ws.cell(row=row, column=8).value or 0
                    
                    # تنظيف البيانات
                    if name:
                        name = str(name).strip()
                    if code:
                        code = str(code).strip()
                    if price:
                        try:
                            price = float(price)
                        except:
                            price = 0
                    else:
                        price = 0
                    
                    try:
                        initial_stock = float(initial_stock)
                    except:
                        initial_stock = 0
                    
                    if not name or not code:
                        if not skip_errors:
                            errors.append({'row': row, 'message': 'اسم المنتج والكود مطلوبان'})
                            error_count += 1
                        continue
                    
                    # التحقق من وجود المنتج
                    if hasattr(Product, 'code'):
                        existing_product = Product.objects.filter(code=code).first()
                    else:
                        existing_product = Product.objects.filter(name=str(name).strip()).first()
                    
                    if existing_product:
                        if update_existing:
                            existing_product.name = str(name).strip()
                            existing_product.barcode = str(barcode).strip() if barcode else ''
                            existing_product.price = float(price)
                            existing_product.description = str(description).strip()
                            existing_product.category = str(category).strip()
                            existing_product.unit = str(unit).strip()
                            
                            # تحديث المخزون إذا كان موجود
                            if hasattr(existing_product, 'stock'):
                                existing_product.stock = float(initial_stock)
                            
                            existing_product.save()
                            updated_count += 1
                        else:
                            if not skip_errors:
                                errors.append({'row': row, 'message': f'المنتج بالاسم "{name}" موجود بالفعل'})
                                error_count += 1
                            continue
                    else:
                        # إنشاء منتج جديد
                        company = getattr(request, 'company', None)
                        if not company:
                            company, created = Company.objects.get_or_create(
                                code='DEFAULT',
                                defaults={
                                    'name': 'الشركة الافتراضية',
                                    'database_name': 'erp_default',
                                    'subscription_end': date.today() + timedelta(days=365)
                                }
                            )
                        
                        product = Product(
                            company=company,
                            name=str(name).strip(),
                            barcode=str(barcode).strip() if barcode else '',
                            price=float(price),
                            description=str(description).strip(),
                            category=str(category).strip(),
                            unit=str(unit).strip(),
                            created_by=request.user
                        )
                        
                        # إضافة الكود إذا كان موجود في النموذج
                        if hasattr(product, 'code'):
                            product.code = str(code).strip()
                        
                        # إضافة المخزون إذا كان موجود في النموذج
                        if hasattr(product, 'stock'):
                            product.stock = float(initial_stock)
                        
                        product.save()
                        success_count += 1
                        
                except Exception as e:
                    if not skip_errors:
                        errors.append({'row': row, 'message': str(e)})
                        error_count += 1
            
            # إعداد النتائج
            results = {
                'success_count': success_count,
                'updated_count': updated_count,
                'error_count': error_count,
                'total_rows': ws.max_row - 1,
                'errors': errors[:10]  # أول 10 أخطاء فقط
            }
            
            if success_count > 0 or updated_count > 0:
                messages.success(request, f'تم استيراد {success_count} منتج جديد وتحديث {updated_count} منتج')
            
            if error_count > 0:
                messages.warning(request, f'تم تجاهل {error_count} صف بسبب أخطاء')
            
            context = {
                'results': results,
                **get_user_context(request)
            }
            return render(request, 'import_products.html', context)
            
        except ImportError:
            messages.error(request, 'مكتبة openpyxl غير مثبتة')
        except Exception as e:
            messages.error(request, f'خطأ في الاستيراد: {str(e)}')
    
    context = get_user_context(request)
    return render(request, 'import_products.html', context)

@login_required
@subscription_required
@permission_required('settings', 'add')
def add_setting(request):
    if request.method == 'POST':
        try:
            key = request.POST.get('key')
            value = request.POST.get('value')
            setting_type = request.POST.get('setting_type', 'string')
            category = request.POST.get('category', 'general')
            description = request.POST.get('description', '')
            
            if not key or not value:
                messages.error(request, 'المفتاح والقيمة مطلوبان')
                return redirect('add_setting')
            
            # التحقق من عدم وجود الإعداد مسبقاً
            if Setting.objects.filter(key=key).exists():
                messages.error(request, f'الإعداد "{key}" موجود بالفعل')
                return redirect('add_setting')
            
            # حفظ الإعداد الجديد
            DynamicSettingsManager.set(
                key=key,
                value=value,
                branch=getattr(request, 'current_branch', None),
                setting_type=setting_type,
                category=category
            )
            
            messages.success(request, f'تم إضافة الإعداد "{key}" بنجاح')
            return redirect('settings')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'setting_types': Setting.SETTING_TYPES,
        'categories': Setting.CATEGORIES,
        **get_user_context(request)
    }
    return render(request, 'add_setting.html', context)



def add_permission(request):
    if not request.session.get('user_id'):
        return redirect('login')
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        screen = request.POST.get('screen')
        permissions = {
            'can_add': 'can_add' in request.POST,
            'can_edit': 'can_edit' in request.POST,
            'can_delete': 'can_delete' in request.POST,
            'can_print': 'can_print' in request.POST,
            'can_confirm': 'can_confirm' in request.POST,
            'can_cancel': 'can_cancel' in request.POST,
        }
        return redirect('permissions')
    return render(request, 'add_permission.html')

@login_required
@subscription_required
@permission_required('reports', 'view')
def reports_center(request):
    from django.db.models import Sum
    from datetime import datetime, timedelta
    
    # الحصول على التواريخ
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # فلترة حسب التاريخ
    sales_filter = {'status': 'confirmed'}
    purchases_filter = {'status': 'confirmed'}
    
    if start_date:
        sales_filter['created_at__gte'] = start_date
        purchases_filter['created_at__gte'] = start_date
    if end_date:
        sales_filter['created_at__lte'] = end_date
        purchases_filter['created_at__lte'] = end_date
    
    # إحصائيات رئيسية
    total_sales = Sale.objects.filter(**sales_filter).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_purchases = Purchase.objects.filter(**purchases_filter).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_expenses = 0  # يمكن إضافة نموذج المصروفات لاحقاً
    net_profit = total_sales - total_purchases - total_expenses
    
    # تقرير المبيعات
    sales_data = Sale.objects.filter(**sales_filter).select_related('customer').order_by('-created_at')[:10]
    
    # حركة المخزون
    stock_movements = StockMovement.objects.select_related('product').order_by('-created_at')[:10]
    
    # تنبيهات المخزون
    low_stock_products = []
    for product in Product.objects.all():
        current_stock = getattr(product, 'stock', 0) or 0
        min_stock = 10  # الحد الأدنى الافتراضي
        if current_stock <= min_stock:
            low_stock_products.append({
                'product': product,
                'current_stock': current_stock,
                'min_stock': min_stock,
                'status': 'منخفض' if current_stock <= min_stock else 'جيد'
            })
    
    context = {
        'total_sales': total_sales,
        'total_purchases': total_purchases,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'sales_data': sales_data,
        'stock_movements': stock_movements,
        'low_stock_products': low_stock_products,
        'start_date': start_date,
        'end_date': end_date,
        **get_user_context(request)
    }
    
    return render(request, 'reports.html', context)

def add_report(request):
    if not request.session.get('user_id'):
        return redirect('login')
    return render(request, 'add_report.html')

@login_required
@subscription_required
@permission_required('products', 'view')
def view_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    
    # الحصول على معلومات إضافية
    try:
        product_price = ProductPrice.objects.filter(product=product).first()
        product_stock = ProductStock.objects.filter(product=product).first()
        recent_movements = StockMovement.objects.filter(product=product).order_by('-created_at')[:10]
    except:
        product_price = None
        product_stock = None
        recent_movements = []
    
    context = {
        'product': product,
        'product_price': product_price,
        'product_stock': product_stock,
        'recent_movements': recent_movements,
        **get_user_context(request)
    }
    return render(request, 'view_product.html', context)

def view_purchase(request, purchase_id):
    if not request.session.get('user_id'):
        return redirect('login')
    from .models import Purchase
    purchase = Purchase.objects.get(id=purchase_id)
    return render(request, 'view_purchase.html', {'purchase': purchase})

def edit_purchase(request, purchase_id):
    if not request.session.get('user_id'):
        return redirect('login')
    from .models import Purchase
    purchase = Purchase.objects.get(id=purchase_id)
    if request.method == 'POST':
        purchase.notes = request.POST.get('notes', purchase.notes)
        purchase.save()
        return redirect('purchases')
    return render(request, 'edit_purchase.html', {'purchase': purchase})

def print_purchase(request, purchase_id):
    if not request.session.get('user_id'):
        return redirect('login')
    from .models import Purchase
    purchase = Purchase.objects.get(id=purchase_id)
    return render(request, 'print_purchase.html', {'purchase': purchase})

def export_purchases(request):
    if not request.session.get('user_id'):
        return redirect('login')
    from django.http import HttpResponse
    return HttpResponse('تصدير المشتريات')

@login_required
@subscription_required
@permission_required('products', 'add')
@csrf_exempt
def duplicate_product(request, product_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        original = get_object_or_404(Product, id=product_id)
        
        company = getattr(request, 'company', None)
        if not company:
            company, created = Company.objects.get_or_create(
                code='DEFAULT',
                defaults={
                    'name': 'الشركة الافتراضية',
                    'database_name': 'erp_default',
                    'subscription_end': date.today() + timedelta(days=365)
                }
            )
        
        new_product = Product.objects.create(
            company=company,
            name=f"{original.name} - نسخة",
            category=original.category,
            unit=original.unit,
            description=original.description,
            bom=original.bom,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'تم إنشاء نسخة من المنتج بنجاح',
            'product_id': new_product.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@permission_required('stock', 'add')
@csrf_exempt
def add_stock(request, product_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        import json
        data = json.loads(request.body)
        quantity = int(data.get('quantity', 0))
        
        if quantity <= 0:
            return JsonResponse({'success': False, 'error': 'الكمية يجب أن تكون أكبر من صفر'})
        
        product = get_object_or_404(Product, id=product_id)
        product.stock = (product.stock or 0) + quantity
        product.save()
        
        StockMovement.objects.create(
            company=getattr(request, 'company', None) or Company.objects.first(),
            product=product,
            movement_type='in',
            quantity=quantity,
            reference=f'إضافة مخزون يدوية',
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'تم إضافة {quantity} وحدة بنجاح',
            'new_stock': product.stock
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# تم نقل هذه الدالة إلى أعلى في الملف

@login_required
@subscription_required
@permission_required('products', 'print')
def print_barcode(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    context = {
        'product': product,
        **get_user_context(request)
    }
    return render(request, 'print_barcode.html', context)

@login_required
@subscription_required
@permission_required('products', 'export')
def export_products(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "المنتجات"
        
        headers = ['اسم المنتج', 'الكود', 'الباركود', 'السعر', 'الوصف', 'الفئة', 'الوحدة', 'المخزون']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        products = Product.objects.all()
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product.name)
            ws.cell(row=row, column=2, value=getattr(product, 'code', ''))
            ws.cell(row=row, column=3, value=product.barcode or '')
            ws.cell(row=row, column=4, value=float(getattr(product, 'price', 0) or 0))
            ws.cell(row=row, column=5, value=product.description or '')
            ws.cell(row=row, column=6, value=product.category or '')
            ws.cell(row=row, column=7, value=product.unit or '')
            ws.cell(row=row, column=8, value=float(getattr(product, 'stock', 0) or 0))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('products')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('products')

@login_required
@subscription_required
@permission_required('products', 'view')
def download_products_template(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "قالب المنتجات"
        
        # عناوين الأعمدة
        headers = ['اسم المنتج *', 'الكود *', 'الباركود', 'السعر *', 'الوصف', 'الفئة', 'الوحدة', 'المخزون الأولي']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        # بيانات تجريبية
        sample_data = [
            ['لابتوب Dell', 'DELL001', '1234567890123', '1500.000', 'لابتوب Dell Inspiron', 'إلكترونيات', 'قطعة', '10'],
            ['ماوس لاسلكي', 'MOUSE001', '9876543210987', '25.500', 'ماوس لاسلكي أسود', 'إكسسوارات', 'قطعة', '50']
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # تعديل عرض الأعمدة
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="products_template.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('import_products_page')
    except Exception as e:
        messages.error(request, f'خطأ في تحميل القالب: {str(e)}')
        return redirect('import_products_page')

@login_required
@subscription_required
@permission_required('products', 'add')
def import_products(request):
    return redirect('import_products_page')

def print_barcodes(request):
    if not request.session.get('user_id'):
        return redirect('login')
    from django.http import HttpResponse
    return HttpResponse('طباعة الباركود')

def delete_customer(request, customer_id):
    if not request.session.get('user_id'):
        return redirect('login')
    from .models import Customer
    customer = Customer.objects.get(id=customer_id)
    customer.delete()
    return redirect('customers')

@login_required
@subscription_required
@permission_required('suppliers', 'delete')
@csrf_exempt
def delete_supplier(request, supplier_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        supplier = get_object_or_404(Supplier, id=supplier_id)
        supplier_name = supplier.name
        supplier.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف المورد "{supplier_name}" بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
    return redirect('suppliers')

@login_required
@subscription_required
@permission_required('sales', 'delete')
def delete_sale(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)
    # عكس تأثير الفاتورة على المخزون
    for item in sale.items.all():
        product = item.product
        product.stock += item.quantity
        product.save()
        # تسجيل حركة مخزون عكسية
        StockMovement.objects.create(
            company=getattr(request, 'company', None) or Company.objects.first(),
            product=product,
            movement_type='in',
            quantity=item.quantity,
            reference=f'إلغاء فاتورة بيع #{sale.invoice_number}',
            created_by=request.user
        )
    # حذف قيود المحاسبة المرتبطة
    AccountingEngine.delete_sale_entries(sale)
    sale.delete()
    messages.success(request, 'تم حذف الفاتورة وكل تأثيراتها بنجاح')
    return redirect('sales')

@login_required
@subscription_required
@permission_required('purchases', 'delete')
def delete_purchase(request, purchase_id):
    purchase = get_object_or_404(Purchase, id=purchase_id)
    # عكس تأثير الفاتورة على المخزون
    for item in purchase.items.all():
        product = item.product
        product.stock -= item.quantity
        product.save()
        # تسجيل حركة مخزون عكسية
        StockMovement.objects.create(
            company=getattr(request, 'company', None) or Company.objects.first(),
            product=product,
            movement_type='out',
            quantity=item.quantity,
            reference=f'إلغاء فاتورة شراء #{purchase.invoice_number}',
            created_by=request.user
        )
    # حذف قيود المحاسبة المرتبطة
    AccountingEngine.delete_purchase_entries(purchase)
    purchase.delete()
    messages.success(request, 'تم حذف فاتورة الشراء وكل تأثيراتها بنجاح')
    return redirect('purchases')

# إدارة الشركات والفروع والمخازن
# from .company_views import companies_list, add_company, delete_company, company_details

@login_required
@subscription_required
@permission_required('branches', 'view')
def branches_list(request):
    branches = Branch.objects.select_related('company').all().order_by('-id')
    context = {
        'branches': branches,
        **get_user_context(request)
    }
    return render(request, 'branches.html', context)

@login_required
@subscription_required
@permission_required('branches', 'add')
def add_branch(request):
    if request.method == 'POST':
        try:
            company_id = request.POST.get('company_id')
            name = request.POST.get('name')
            code = request.POST.get('code')
            address = request.POST.get('address', '')
            manager_id = request.POST.get('manager_id') or None
            
            if not company_id or not name or not code:
                messages.error(request, 'يرجى إدخال جميع البيانات المطلوبة')
                return redirect('add_branch')
            
            # التحقق من عدم تكرار كود الفرع في نفس الشركة
            if Branch.objects.filter(company_id=company_id, code=code).exists():
                messages.error(request, f'كود الفرع "{code}" موجود بالفعل في هذه الشركة')
                return redirect('add_branch')
            
            branch = Branch.objects.create(
                company_id=company_id,
                name=name,
                code=code,
                address=address,
                manager_id=manager_id
            )
            messages.success(request, f'تم إضافة الفرع "{name}" بنجاح')
            return redirect('branches')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            return redirect('add_branch')
    
    context = {
        'companies': Company.objects.all(),
        'users': User.objects.all(),
        **get_user_context(request)
    }
    return render(request, 'add_branch.html', context)

@login_required
@subscription_required
@permission_required('warehouses', 'view')
def warehouses_list(request):
    warehouses = Warehouse.objects.select_related('branch', 'branch__company').all().order_by('-id')
    context = {
        'warehouses': warehouses,
        **get_user_context(request)
    }
    return render(request, 'warehouses.html', context)

@login_required
@subscription_required
@permission_required('warehouses', 'add')
def add_warehouse(request):
    if request.method == 'POST':
        try:
            branch_id = request.POST.get('branch_id')
            name = request.POST.get('name')
            code = request.POST.get('code')
            
            if not branch_id or not name or not code:
                messages.error(request, 'يرجى إدخال جميع البيانات المطلوبة')
                return redirect('add_warehouse')
            
            # التحقق من عدم تكرار كود المخزن في نفس الفرع
            if Warehouse.objects.filter(branch_id=branch_id, code=code).exists():
                messages.error(request, f'كود المخزن "{code}" موجود بالفعل في هذا الفرع')
                return redirect('add_warehouse')
            
            warehouse = Warehouse.objects.create(
                branch_id=branch_id,
                name=name,
                code=code
            )
            messages.success(request, f'تم إضافة المخزن "{name}" بنجاح')
            return redirect('warehouses')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            return redirect('add_warehouse')
    
    context = {
        'branches': Branch.objects.select_related('company').all(),
        **get_user_context(request)
    }
    return render(request, 'add_warehouse.html', context)

@login_required
@subscription_required
@permission_required('users', 'view')
def users_list(request):
    users = User.objects.select_related('userprofile').all().order_by('-date_joined')
    
    # حساب الإحصائيات
    active_users_count = users.filter(is_active=True).count()
    inactive_users_count = users.filter(is_active=False).count()
    admin_users_count = users.filter(is_superuser=True).count()
    
    context = {
        'users': users,
        'active_users_count': active_users_count,
        'inactive_users_count': inactive_users_count,
        'admin_users_count': admin_users_count,
        **get_user_context(request)
    }
    return render(request, 'users.html', context)

@login_required
@subscription_required
@permission_required('users', 'add')
def add_user(request):
    if request.method == 'POST':
        try:
            # التحقق من البيانات المطلوبة
            username = request.POST.get('username')
            password = request.POST.get('password')
            password_confirm = request.POST.get('password_confirm')
            first_name = request.POST.get('first_name')
            
            if not username or not password or not first_name:
                messages.error(request, 'يرجى ملء جميع الحقول المطلوبة')
                return redirect('add_user')
            
            if password != password_confirm:
                messages.error(request, 'كلمة المرور وتأكيد كلمة المرور غير متطابقتين')
                return redirect('add_user')
            
            if len(password) < 6:
                messages.error(request, 'كلمة المرور يجب أن تكون 6 أحرف على الأقل')
                return redirect('add_user')
            
            # التحقق من عدم وجود اسم مستخدم مشابه
            if User.objects.filter(username=username).exists():
                messages.error(request, 'اسم المستخدم موجود بالفعل')
                return redirect('add_user')
            
            # إنشاء المستخدم
            user = User.objects.create_user(
                username=username,
                email=request.POST.get('email', ''),
                password=password,
                first_name=first_name,
                last_name=request.POST.get('last_name', ''),
                is_active=request.POST.get('is_active') == 'on',
                is_staff=request.POST.get('is_staff') == 'on',
                is_superuser=request.POST.get('is_superuser') == 'on'
            )
            
            # إنشاء ملف المستخدم في نفس شركة المستخدم الحالي
            company = getattr(request, 'company', None)
            if not company:
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
            
            branch_id = request.POST.get('branch_id') or None
            warehouse_id = request.POST.get('warehouse_id') or None
            
            UserProfile.objects.create(
                user=user,
                company=company,
                default_branch_id=branch_id,
                default_warehouse_id=warehouse_id,
                is_active=True
            )
            
            messages.success(request, f'تم إضافة المستخدم "{user.get_full_name() or user.username}" بنجاح')
            return redirect('users')
            
        except Exception as e:
            messages.error(request, f'خطأ في إضافة المستخدم: {str(e)}')
    
    context = {
        'companies': Company.objects.all(),
        'branches': Branch.objects.all(),
        'warehouses': Warehouse.objects.all(),
        **get_user_context(request)
    }
    return render(request, 'add_user.html', context)

@login_required
@subscription_required
@permission_required('users', 'edit')
def edit_user(request, user_id):
    user_obj = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        try:
            # تحديث بيانات المستخدم الأساسية
            user_obj.username = request.POST.get('username')
            user_obj.email = request.POST.get('email')
            user_obj.first_name = request.POST.get('first_name')
            user_obj.last_name = request.POST.get('last_name')
            user_obj.is_active = request.POST.get('is_active') == 'on'
            user_obj.is_staff = request.POST.get('is_staff') == 'on'
            user_obj.is_superuser = request.POST.get('is_superuser') == 'on'
            
            # تحديث كلمة المرور إذا تم إدخالها
            password = request.POST.get('password')
            password_confirm = request.POST.get('password_confirm')
            
            if password:
                if password != password_confirm:
                    messages.error(request, 'كلمة المرور وتأكيد كلمة المرور غير متطابقتين')
                    return redirect('edit_user', user_id=user_id)
                user_obj.set_password(password)
            
            user_obj.save()
            
            # تحديث ملف المستخدم
            profile, created = UserProfile.objects.get_or_create(user=user_obj)
            
            branch_id = request.POST.get('branch_id')
            warehouse_id = request.POST.get('warehouse_id')
            
            if branch_id:
                profile.default_branch_id = branch_id
            if warehouse_id:
                profile.default_warehouse_id = warehouse_id
            
            profile.save()
            
            messages.success(request, 'تم تحديث بيانات المستخدم بنجاح')
            return redirect('users')
            
        except Exception as e:
            messages.error(request, f'خطأ في تحديث البيانات: {str(e)}')
    
    context = {
        'user': user_obj,
        'branches': Branch.objects.all(),
        'warehouses': Warehouse.objects.all(),
        **get_user_context(request)
    }
    return render(request, 'edit_user.html', context)

@login_required
@subscription_required
@permission_required('users', 'delete')
def delete_user(request, user_id):
    try:
        user = get_object_or_404(User, id=user_id)
        
        # منع حذف المستخدم لنفسه
        if user.id == request.user.id:
            messages.error(request, 'لا يمكنك حذف نفسك')
            return redirect('users')
        
        if request.method == 'POST':
            username = user.username
            user.delete()
            messages.success(request, f'تم حذف المستخدم "{username}" بنجاح')
            return redirect('users')
        
        context = {
            'user_to_delete': user,
            **get_user_context(request)
        }
        return render(request, 'delete_user_confirm.html', context)
        
    except Exception as e:
        messages.error(request, 'المستخدم غير موجود')
        return redirect('users')

@login_required
@subscription_required
@permission_required('sales', 'confirm')
def confirm_invoice(request, sale_id):
    try:
        sale = get_object_or_404(Sale, id=sale_id)
        if sale.status == 'confirmed' or getattr(sale, 'is_confirmed', False):
            messages.error(request, 'الفاتورة مؤكدة بالفعل')
        else:
            # تأكيد الفاتورة
            sale.status = 'confirmed'
            if hasattr(sale, 'is_confirmed'):
                sale.is_confirmed = True
            if hasattr(sale, 'confirmed_by'):
                sale.confirmed_by = request.user
            if hasattr(sale, 'confirmed_at'):
                sale.confirmed_at = timezone.now()
            sale.save()
            
            # تحديث المخزون
            try:
                for item in sale.items.all():
                    product = item.product
                    if hasattr(product, 'stock'):
                        current_stock = getattr(product, 'stock', 0) or 0
                        product.stock = current_stock - item.quantity
                        product.save()
                        
                        # تسجيل حركة مخزون
                        StockMovement.objects.create(
                            company=getattr(request, 'company', None) or Company.objects.first(),
                            product=product,
                            movement_type='out',
                            quantity=item.quantity,
                            reference=f'فاتورة بيع #{sale.invoice_number}',
                            created_by=request.user
                        )
            except Exception as e:
                print(f"خطأ في تحديث المخزون: {e}")
            
            # إنشاء القيد المحاسبي
            try:
                create_sale_journal_entry(sale)
                messages.success(request, f'تم تأكيد الفاتورة #{sale.invoice_number} وتسجيل القيد المحاسبي بنجاح')
            except Exception as e:
                messages.warning(request, f'تم تأكيد الفاتورة ولكن حدث خطأ في المحاسبة: {str(e)}')
                
    except Exception as e:
        messages.error(request, f'خطأ غير متوقع: {str(e)}')
    return redirect('invoices_management')

@login_required
@subscription_required
@permission_required('purchases', 'confirm')
def confirm_purchase(request, purchase_id):
    try:
        purchase = get_object_or_404(Purchase, id=purchase_id)
        if purchase.status == 'confirmed':
            messages.error(request, 'الفاتورة مؤكدة بالفعل')
        else:
            purchase.status = 'confirmed'
            purchase.confirmed_by = request.user
            purchase.confirmed_at = timezone.now()
            purchase.save()
            
            # معالجة شاملة: مخزون + محاسبة
            try:
                InventoryAccountingManager.process_purchase(purchase.items.all(), request.user)
                InventoryAccountingManager.update_account_balances()
                messages.success(request, f'تم تأكيد فاتورة الشراء #{purchase.invoice_number} وتسجيل جميع القيود المحاسبية')
            except Exception as e:
                messages.warning(request, f'تم تأكيد الفاتورة ولكن حدث خطأ في المحاسبة: {str(e)}')
    except Exception as e:
        messages.error(request, f'خطأ غير متوقع: {str(e)}')
    return redirect('purchases')

@login_required
@subscription_required
@permission_required('sales', 'delete')
def cancel_invoice(request, sale_id):
    try:
        sale = get_object_or_404(Sale, id=sale_id)
        if sale.status == 'confirmed':
            messages.error(request, 'لا يمكن إلغاء فاتورة مؤكدة')
        else:
            sale.status = 'cancelled'
            sale.save()
            messages.success(request, f'تم إلغاء فاتورة البيع #{sale.invoice_number}')
    except Exception as e:
        messages.error(request, f'خطأ: {str(e)}')
    return redirect('invoices_management')

@login_required
@subscription_required
@permission_required('purchases', 'delete')
def cancel_purchase(request, purchase_id):
    try:
        purchase = get_object_or_404(Purchase, id=purchase_id)
        if purchase.status == 'confirmed':
            messages.error(request, 'لا يمكن إلغاء فاتورة مؤكدة')
        else:
            purchase.status = 'cancelled'
            purchase.save()
            messages.success(request, f'تم إلغاء فاتورة الشراء #{purchase.invoice_number}')
    except Exception as e:
        messages.error(request, f'خطأ: {str(e)}')
    return redirect('purchases')

@login_required
@subscription_required
def invoices_management(request):
    # فلترة حسب الحالة
    status_filter = request.GET.get('status', 'all')
    invoice_type = request.GET.get('type', 'sales')
    
    if invoice_type == 'sales':
        base_query = Sale.objects.select_related(
            'customer', 'created_by', 'branch', 'warehouse', 'sales_rep', 'confirmed_by'
        ).prefetch_related('items__product')
        
        if status_filter == 'confirmed':
            invoices = base_query.filter(status='confirmed').order_by('-created_at')
        elif status_filter == 'pending':
            invoices = base_query.filter(status='draft').order_by('-created_at')
        else:
            invoices = base_query.all().order_by('-created_at')
        invoice_model = 'sale'
        all_invoices = Sale.objects.all()
    else:
        base_query = Purchase.objects.select_related(
            'supplier', 'created_by', 'branch', 'warehouse', 'confirmed_by'
        ).prefetch_related('items__product')
        
        if status_filter == 'confirmed':
            invoices = base_query.filter(status='confirmed').order_by('-created_at')
        elif status_filter == 'pending':
            invoices = base_query.filter(status='draft').order_by('-created_at')
        else:
            invoices = base_query.all().order_by('-created_at')
        invoice_model = 'purchase'
        all_invoices = Purchase.objects.all()
    
    # إضافة بيانات محسوبة لكل فاتورة
    for invoice in invoices:
        invoice.remaining_amount = (invoice.total_amount or 0) - (invoice.paid_amount or 0)
        invoice.items_count = invoice.items.count() if hasattr(invoice, 'items') else 0
        invoice.payment_status = 'مدفوع' if invoice.remaining_amount <= 0 else 'مستحق جزئي' if (invoice.paid_amount or 0) > 0 else 'غير مدفوع'
    
    # حساب الإحصائيات المحسنة
    stats = {
        'total_invoices': all_invoices.count(),
        'draft_count': all_invoices.filter(status='draft').count(),
        'confirmed_count': all_invoices.filter(status='confirmed').count(),
        'cancelled_count': all_invoices.filter(status='cancelled').count(),
        'total_amount': all_invoices.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_paid': all_invoices.aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0,
    }
    stats['total_due'] = stats['total_amount'] - stats['total_paid']
    
    context = {
        'invoices': invoices,
        'status_filter': status_filter,
        'invoice_type': invoice_type,
        'invoice_model': invoice_model,
        'stats': stats,
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        **get_user_context(request)
    }
    return render(request, 'invoices_management.html', context)

@login_required
@subscription_required
@permission_required('attendance', 'view')
def attendance_list(request):
    from datetime import date, timedelta
    from django.db.models import Q, Count
    
    # فلترة البيانات
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    employee_id = request.GET.get('employee')
    status = request.GET.get('status')
    
    # تحديد التاريخ الافتراضي (آخر 30 يوم)
    if not start_date:
        start_date = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = date.today().strftime('%Y-%m-%d')
    
    # جلب سجلات الحضور
    attendance_records = Attendance.objects.select_related('employee').filter(
        date__range=[start_date, end_date]
    ).order_by('-date', '-created_at')
    
    if employee_id:
        attendance_records = attendance_records.filter(employee_id=employee_id)
    if status:
        attendance_records = attendance_records.filter(status=status)
    
    # إحصائيات اليوم
    today = date.today()
    today_attendance = Attendance.objects.filter(date=today)
    
    stats = {
        'present_count': today_attendance.filter(status='present').count(),
        'absent_count': today_attendance.filter(status='absent').count(),
        'late_count': today_attendance.filter(status='late').count(),
        'early_leave_count': today_attendance.filter(status='early_leave').count(),
    }
    
    # حساب ساعات العمل لكل سجل
    for record in attendance_records:
        if record.check_in and record.check_out:
            work_duration = record.check_out.hour * 60 + record.check_out.minute - (record.check_in.hour * 60 + record.check_in.minute)
            hours = work_duration // 60
            minutes = work_duration % 60
            record.work_hours = f"{hours}:{minutes:02d}"
        else:
            record.work_hours = "--:--"
    
    context = {
        'attendance_records': attendance_records,
        'employees': User.objects.filter(is_active=True).order_by('first_name'),
        'stats': stats,
        **get_user_context(request)
    }
    return render(request, 'attendance.html', context)

@login_required
@subscription_required
@permission_required('attendance', 'add')
def add_attendance(request):
    if request.method == 'POST':
        try:
            from datetime import datetime
            
            employee_id = request.POST.get('employee_id')
            date_str = request.POST.get('date')
            check_in_str = request.POST.get('check_in')
            check_out_str = request.POST.get('check_out')
            status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            
            # تحويل التواريخ والأوقات
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            check_in = datetime.strptime(check_in_str, '%H:%M').time() if check_in_str else None
            check_out = datetime.strptime(check_out_str, '%H:%M').time() if check_out_str else None
            
            # التحقق من عدم وجود سجل مسبق
            if Attendance.objects.filter(employee_id=employee_id, date=attendance_date).exists():
                messages.error(request, 'يوجد سجل حضور لهذا الموظف في هذا التاريخ')
                return redirect('add_attendance')
            
            # حساب الساعات الإضافية
            overtime_hours = 0
            if check_in and check_out:
                work_minutes = (check_out.hour * 60 + check_out.minute) - (check_in.hour * 60 + check_in.minute)
                if work_minutes > 480:  # أكثر من 8 ساعات
                    overtime_hours = (work_minutes - 480) / 60
            
            Attendance.objects.create(
                employee_id=employee_id,
                date=attendance_date,
                check_in=check_in,
                check_out=check_out,
                status=status,
                notes=notes,
                overtime_hours=overtime_hours,
                created_by=request.user
            )
            
            messages.success(request, 'تم تسجيل الحضور بنجاح')
            return redirect('attendance')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'employees': User.objects.filter(is_active=True).order_by('first_name'),
        **get_user_context(request)
    }
    return render(request, 'add_attendance.html', context)

# دوال جهاز البصمة
@login_required
@csrf_exempt
def device_status(request):
    """فحص حالة جهاز البصمة"""
    try:
        # محاولة الاتصال بجهاز البصمة
        # هذا مثال - يجب تعديله حسب نوع الجهاز
        import socket
        device_ip = get_setting('fingerprint_device_ip', '192.168.1.100')
        device_port = int(get_setting('fingerprint_device_port', '4370'))
        
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(3)
        result = sock.connect_ex((device_ip, device_port))
        sock.close()
        
        if result == 0:
            return JsonResponse({
                'connected': True,
                'device_name': get_setting('fingerprint_device_name', 'جهاز البصمة'),
                'ip': device_ip
            })
        else:
            return JsonResponse({'connected': False})
            
    except Exception as e:
        return JsonResponse({'connected': False, 'error': str(e)})

@login_required
@csrf_exempt
def connect_device(request):
    """إعادة الاتصال بجهاز البصمة"""
    try:
        # محاولة إعادة الاتصال
        return JsonResponse({'success': True, 'message': 'تم الاتصال بنجاح'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@csrf_exempt
def start_fingerprint_scan(request):
    """بدء مسح البصمة"""
    try:
        data = json.loads(request.body)
        action = data.get('action')  # 'checkin' أو 'checkout'
        
        # حفظ العملية في الجلسة
        request.session['fingerprint_action'] = action
        request.session['fingerprint_scanning'] = True
        
        # بدء مسح البصمة على الجهاز
        # هذا مثال - يجب تعديله حسب SDK الجهاز
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@csrf_exempt
def stop_fingerprint_scan(request):
    """إيقاف مسح البصمة"""
    request.session['fingerprint_scanning'] = False
    return JsonResponse({'success': True})

@login_required
def fingerprint_result(request):
    """التحقق من نتيجة مسح البصمة"""
    try:
        if not request.session.get('fingerprint_scanning'):
            return JsonResponse({'completed': False})
        
        # فحص نتيجة البصمة من الجهاز
        # هذا مثال - يجب تعديله حسب SDK الجهاز
        
        # محاكاة نجاح البصمة
        import random
        if random.choice([True, False]):  # محاكاة
            # العثور على الموظف من البصمة
            employee = request.user  # في الواقع سيتم البحث في قاعدة بيانات البصمات
            action = request.session.get('fingerprint_action')
            
            from datetime import date, datetime
            today = date.today()
            now = datetime.now().time()
            
            # الحصول على أو إنشاء سجل الحضور
            attendance, created = Attendance.objects.get_or_create(
                employee=employee,
                date=today,
                defaults={
                    'status': 'present',
                    'created_by': request.user
                }
            )
            
            if action == 'checkin':
                attendance.check_in = now
                # تحديد حالة التأخير
                work_start = datetime.strptime('09:00', '%H:%M').time()
                if now > work_start:
                    attendance.status = 'late'
            else:  # checkout
                attendance.check_out = now
                # حساب الساعات الإضافية
                if attendance.check_in:
                    work_minutes = (now.hour * 60 + now.minute) - (attendance.check_in.hour * 60 + attendance.check_in.minute)
                    if work_minutes > 480:  # أكثر من 8 ساعات
                        attendance.overtime_hours = (work_minutes - 480) / 60
            
            attendance.save()
            request.session['fingerprint_scanning'] = False
            
            return JsonResponse({
                'completed': True,
                'success': True,
                'employee': employee.get_full_name() or employee.username
            })
        else:
            return JsonResponse({'completed': False})
            
    except Exception as e:
        request.session['fingerprint_scanning'] = False
        return JsonResponse({
            'completed': True,
            'success': False,
            'message': str(e)
        })

@login_required
@csrf_exempt
def sync_attendance(request):
    """مزامنة بيانات الحضور من جهاز البصمة"""
    try:
        # جلب البيانات من جهاز البصمة
        # هذا مثال - يجب تعديله حسب SDK الجهاز
        
        synced_count = 0
        # محاكاة مزامنة بعض السجلات
        
        return JsonResponse({
            'success': True,
            'synced_count': synced_count,
            'message': f'تم مزامنة {synced_count} سجل'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def attendance_stats(request):
    """إحصائيات الحضور اليومية"""
    from datetime import date
    today = date.today()
    today_attendance = Attendance.objects.filter(date=today)
    
    stats = {
        'present_count': today_attendance.filter(status='present').count(),
        'absent_count': today_attendance.filter(status='absent').count(),
        'late_count': today_attendance.filter(status='late').count(),
        'early_leave_count': today_attendance.filter(status='early_leave').count(),
    }
    
    return JsonResponse(stats)

@login_required
@permission_required('attendance', 'export')
def export_attendance(request):
    """تصدير بيانات الحضور"""
    try:
        import openpyxl
        from django.http import HttpResponse
        from datetime import date, timedelta
        
        # فلترة البيانات
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        employee_id = request.GET.get('employee')
        
        if not start_date:
            start_date = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = date.today().strftime('%Y-%m-%d')
        
        attendance_records = Attendance.objects.select_related('employee').filter(
            date__range=[start_date, end_date]
        ).order_by('-date')
        
        if employee_id:
            attendance_records = attendance_records.filter(employee_id=employee_id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الحضور والانصراف"
        
        headers = ['الموظف', 'التاريخ', 'وقت الحضور', 'وقت الانصراف', 'ساعات العمل', 'الساعات الإضافية', 'الحالة', 'ملاحظات']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row, record in enumerate(attendance_records, 2):
            ws.cell(row=row, column=1, value=record.employee.get_full_name() or record.employee.username)
            ws.cell(row=row, column=2, value=record.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=3, value=record.check_in.strftime('%H:%M') if record.check_in else '')
            ws.cell(row=row, column=4, value=record.check_out.strftime('%H:%M') if record.check_out else '')
            
            # حساب ساعات العمل
            if record.check_in and record.check_out:
                work_duration = (record.check_out.hour * 60 + record.check_out.minute) - (record.check_in.hour * 60 + record.check_in.minute)
                hours = work_duration // 60
                minutes = work_duration % 60
                ws.cell(row=row, column=5, value=f"{hours}:{minutes:02d}")
            else:
                ws.cell(row=row, column=5, value='')
            
            ws.cell(row=row, column=6, value=float(record.overtime_hours))
            ws.cell(row=row, column=7, value=dict(Attendance.ATTENDANCE_STATUS).get(record.status, record.status))
            ws.cell(row=row, column=8, value=record.notes or '')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="attendance.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('attendance')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('attendance')

@login_required
@permission_required('attendance', 'edit')
def edit_attendance(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    
    if request.method == 'POST':
        try:
            from datetime import datetime
            
            check_in_str = request.POST.get('check_in')
            check_out_str = request.POST.get('check_out')
            status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            
            attendance.check_in = datetime.strptime(check_in_str, '%H:%M').time() if check_in_str else None
            attendance.check_out = datetime.strptime(check_out_str, '%H:%M').time() if check_out_str else None
            attendance.status = status
            attendance.notes = notes
            
            # حساب الساعات الإضافية
            if attendance.check_in and attendance.check_out:
                work_minutes = (attendance.check_out.hour * 60 + attendance.check_out.minute) - (attendance.check_in.hour * 60 + attendance.check_in.minute)
                if work_minutes > 480:  # أكثر من 8 ساعات
                    attendance.overtime_hours = (work_minutes - 480) / 60
                else:
                    attendance.overtime_hours = 0
            
            attendance.save()
            messages.success(request, 'تم تحديث سجل الحضور بنجاح')
            return redirect('attendance')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'attendance': attendance,
        **get_user_context(request)
    }
    return render(request, 'edit_attendance.html', context)

@login_required
@permission_required('attendance', 'delete')
@csrf_exempt
def delete_attendance(request, attendance_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        attendance = get_object_or_404(Attendance, id=attendance_id)
        employee_name = attendance.employee.get_full_name() or attendance.employee.username
        attendance.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف سجل حضور {employee_name} بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@permission_required('salaries', 'view')
def salaries_list(request):
    from datetime import date
    from django.db.models import Sum, Count
    
    # فلترة البيانات
    month = request.GET.get('month')
    year = request.GET.get('year')
    employee_id = request.GET.get('employee')
    status = request.GET.get('status')
    
    # جلب الرواتب
    salaries = Salary.objects.select_related('employee').all().order_by('-year', '-month', 'employee__first_name')
    
    if month:
        salaries = salaries.filter(month=month)
    if year:
        salaries = salaries.filter(year=year)
    if employee_id:
        salaries = salaries.filter(employee_id=employee_id)
    if status:
        salaries = salaries.filter(status=status)
    
    # إحصائيات
    current_month = date.today().month
    current_year = date.today().year
    
    current_month_salaries = Salary.objects.filter(month=current_month, year=current_year)
    
    stats = {
        'total_salaries': current_month_salaries.aggregate(Sum('net_salary'))['net_salary__sum'] or 0,
        'pending_count': current_month_salaries.filter(status='draft').count(),
        'paid_count': current_month_salaries.filter(status='paid').count(),
        'employees_count': User.objects.filter(is_active=True).count(),
    }
    
    # جلب الإعدادات الحالية
    current_settings = {}
    for setting in Setting.objects.all():
        current_settings[setting.key] = setting.get_value()
    
    # قائمة السنوات المتاحة
    years = list(range(current_year - 2, current_year + 2))
    
    context = {
        'salaries': salaries,
        'employees': User.objects.filter(is_active=True).order_by('first_name'),
        'stats': stats,
        'years': years,
        'current_settings': current_settings,
        **get_user_context(request)
    }
    return render(request, 'salaries.html', context)

@login_required
@subscription_required
@permission_required('salaries', 'add')
def add_salary(request):
    if request.method == 'POST':
        try:
            employee_id = request.POST.get('employee_id')
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))
            basic_salary = float(request.POST.get('basic_salary', 0))
            allowances = float(request.POST.get('allowances', 0))
            deductions = float(request.POST.get('deductions', 0))
            overtime_rate = float(request.POST.get('overtime_rate', 0))
            notes = request.POST.get('notes', '')
            
            # التحقق من عدم وجود راتب مسبق
            if Salary.objects.filter(employee_id=employee_id, month=month, year=year).exists():
                messages.error(request, 'يوجد راتب لهذا الموظف في هذه الفترة')
                return redirect('add_salary')
            
            # حساب الساعات الإضافية من الحضور
            overtime_hours = calculate_overtime_hours(employee_id, month, year)
            overtime_amount = overtime_hours * overtime_rate
            
            Salary.objects.create(
                employee_id=employee_id,
                month=month,
                year=year,
                basic_salary=basic_salary,
                allowances=allowances,
                deductions=deductions,
                overtime_hours=overtime_hours,
                overtime_rate=overtime_rate,
                overtime_amount=overtime_amount,
                notes=notes,
                created_by=request.user
            )
            
            messages.success(request, 'تم إضافة الراتب بنجاح')
            return redirect('salaries')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    from datetime import date
    context = {
        'employees': User.objects.filter(is_active=True).order_by('first_name'),
        'current_month': date.today().month,
        'current_year': date.today().year,
        **get_user_context(request)
    }
    return render(request, 'add_salary.html', context)

def calculate_overtime_hours(employee_id, month, year):
    """حساب الساعات الإضافية من سجلات الحضور"""
    from django.db.models import Sum
    
    total_overtime = Attendance.objects.filter(
        employee_id=employee_id,
        date__month=month,
        date__year=year
    ).aggregate(Sum('overtime_hours'))['overtime_hours__sum'] or 0
    
    return float(total_overtime)

@login_required
@csrf_exempt
def generate_salaries(request):
    try:
        data = json.loads(request.body)
        month = int(data.get('month'))
        year = int(data.get('year'))
        
        generated_count = 0
        employees = User.objects.filter(is_active=True)
        
        for employee in employees:
            if Salary.objects.filter(employee=employee, month=month, year=year).exists():
                continue
            
            basic_salary = 5000
            overtime_hours = calculate_overtime_hours(employee.id, month, year)
            overtime_rate = 50
            overtime_amount = overtime_hours * overtime_rate
            
            attendance_days = Attendance.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year,
                status__in=['present', 'late']
            ).count()
            
            working_days = 30
            if attendance_days < working_days:
                absence_deduction = (basic_salary / working_days) * (working_days - attendance_days)
            else:
                absence_deduction = 0
            
            Salary.objects.create(
                employee=employee,
                month=month,
                year=year,
                basic_salary=basic_salary,
                allowances=0,
                deductions=absence_deduction,
                overtime_hours=overtime_hours,
                overtime_rate=overtime_rate,
                overtime_amount=overtime_amount,
                created_by=request.user
            )
            generated_count += 1
        
        return JsonResponse({
            'success': True,
            'generated_count': generated_count,
            'message': f'تم توليد {generated_count} راتب'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def salary_details(request, salary_id):
    try:
        salary = get_object_or_404(Salary, id=salary_id)
        
        attendance_days = Attendance.objects.filter(
            employee=salary.employee,
            date__month=salary.month,
            date__year=salary.year,
            status__in=['present', 'late']
        ).count()
        
        return JsonResponse({
            'success': True,
            'salary': {
                'employee_name': salary.employee.get_full_name() or salary.employee.username,
                'month': salary.month,
                'year': salary.year,
                'basic_salary': float(salary.basic_salary),
                'allowances': float(salary.allowances),
                'deductions': float(salary.deductions),
                'overtime_hours': float(salary.overtime_hours),
                'overtime_amount': float(salary.overtime_amount),
                'net_salary': float(salary.net_salary),
                'attendance_days': attendance_days,
                'status': salary.status
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@csrf_exempt
def confirm_salary(request, salary_id):
    try:
        salary = get_object_or_404(Salary, id=salary_id)
        if salary.status != 'draft':
            return JsonResponse({'success': False, 'message': 'الراتب مؤكد بالفعل'})
        
        salary.status = 'confirmed'
        salary.save()
        
        # تسجيل القيد المحاسبي للراتب
        try:
            InventoryAccountingManager.process_salary(salary, request.user)
            InventoryAccountingManager.update_account_balances()
            message = 'تم تأكيد الراتب وتسجيل القيد المحاسبي بنجاح'
        except Exception as e:
            message = f'تم تأكيد الراتب ولكن حدث خطأ في المحاسبة: {str(e)}'
        
        return JsonResponse({'success': True, 'message': message})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@csrf_exempt
def pay_salary(request, salary_id):
    try:
        salary = get_object_or_404(Salary, id=salary_id)
        if salary.status != 'confirmed':
            return JsonResponse({'success': False, 'message': 'يجب تأكيد الراتب أولاً'})
        
        salary.status = 'paid'
        salary.save()
        
        return JsonResponse({'success': True, 'message': 'تم تسجيل دفع الراتب بنجاح'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@csrf_exempt
def delete_salary(request, salary_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        salary = get_object_or_404(Salary, id=salary_id)
        employee_name = salary.employee.get_full_name() or salary.employee.username
        salary.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف راتب {employee_name} بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@csrf_exempt
def bulk_pay_salaries(request):
    try:
        confirmed_salaries = Salary.objects.filter(status='confirmed')
        paid_count = confirmed_salaries.update(status='paid')
        
        return JsonResponse({
            'success': True,
            'paid_count': paid_count,
            'message': f'تم دفع {paid_count} راتب'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def print_salary_slip(request, salary_id):
    salary = get_object_or_404(Salary, id=salary_id)
    
    attendance_days = Attendance.objects.filter(
        employee=salary.employee,
        date__month=salary.month,
        date__year=salary.year,
        status__in=['present', 'late']
    ).count()
    
    context = {
        'salary': salary,
        'attendance_days': attendance_days,
        'company_name': get_setting('company_name', 'شركة ERP'),
        'currency_symbol': get_setting('currency_symbol', 'ر.س'),
        **get_user_context(request)
    }
    return render(request, 'print_salary_slip.html', context)

@login_required
def edit_salary(request, salary_id):
    salary = get_object_or_404(Salary, id=salary_id)
    
    if request.method == 'POST':
        try:
            salary.basic_salary = float(request.POST.get('basic_salary', 0))
            salary.allowances = float(request.POST.get('allowances', 0))
            salary.deductions = float(request.POST.get('deductions', 0))
            salary.overtime_rate = float(request.POST.get('overtime_rate', 0))
            salary.notes = request.POST.get('notes', '')
            
            salary.overtime_hours = calculate_overtime_hours(salary.employee.id, salary.month, salary.year)
            salary.overtime_amount = salary.overtime_hours * salary.overtime_rate
            
            salary.save()
            
            messages.success(request, 'تم تحديث الراتب بنجاح')
            return redirect('salaries')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'salary': salary,
        **get_user_context(request)
    }
    return render(request, 'edit_salary.html', context)

@login_required
@permission_required('salaries', 'export')
def export_salaries(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        
        month = request.GET.get('month')
        year = request.GET.get('year')
        employee_id = request.GET.get('employee')
        status = request.GET.get('status')
        
        salaries = Salary.objects.select_related('employee').all()
        
        if month:
            salaries = salaries.filter(month=month)
        if year:
            salaries = salaries.filter(year=year)
        if employee_id:
            salaries = salaries.filter(employee_id=employee_id)
        if status:
            salaries = salaries.filter(status=status)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الرواتب"
        
        headers = ['الموظف', 'الشهر', 'السنة', 'الراتب الأساسي', 'البدلات', 'الخصومات', 'الساعات الإضافية', 'مبلغ الساعات الإضافية', 'صافي الراتب', 'الحالة']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row, salary in enumerate(salaries, 2):
            ws.cell(row=row, column=1, value=salary.employee.get_full_name() or salary.employee.username)
            ws.cell(row=row, column=2, value=salary.month)
            ws.cell(row=row, column=3, value=salary.year)
            ws.cell(row=row, column=4, value=float(salary.basic_salary))
            ws.cell(row=row, column=5, value=float(salary.allowances))
            ws.cell(row=row, column=6, value=float(salary.deductions))
            ws.cell(row=row, column=7, value=float(salary.overtime_hours))
            ws.cell(row=row, column=8, value=float(salary.overtime_amount))
            ws.cell(row=row, column=9, value=float(salary.net_salary))
            ws.cell(row=row, column=10, value=dict(Salary.SALARY_STATUS).get(salary.status, salary.status))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="salaries.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('salaries')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('salaries')

@login_required
@subscription_required
@permission_required('users', 'add')
def add_employee(request):
    if request.method == 'POST':
        try:
            from django.contrib.auth.hashers import make_password
            from datetime import datetime, date, timedelta
            
            # بيانات المستخدم
            username = request.POST.get('username')
            email = request.POST.get('email', '')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            password = request.POST.get('password')
            can_login = request.POST.get('can_login') == 'on'
            
            # بيانات الموظف
            national_id = request.POST.get('national_id')
            phone = request.POST.get('phone')
            address = request.POST.get('address', '')
            birth_date = request.POST.get('birth_date')
            hire_date = request.POST.get('hire_date')
            department = request.POST.get('department')
            position = request.POST.get('position')
            basic_salary = float(request.POST.get('basic_salary', 0))
            overtime_rate = float(request.POST.get('overtime_rate', 0))
            employment_type = request.POST.get('employment_type', 'full_time')
            branch_id = request.POST.get('branch_id')
            fingerprint_id = request.POST.get('fingerprint_id', '')
            
            # إعدادات إضافية
            is_sales_rep = request.POST.get('is_sales_rep') == 'on'
            commission_rate = float(request.POST.get('commission_rate', 0)) if is_sales_rep else 0
            create_account = request.POST.get('create_account') == 'on'
            
            # التحقق من عدم تكرار البيانات
            if User.objects.filter(username=username).exists():
                messages.error(request, 'اسم المستخدم موجود بالفعل')
                return redirect('add_employee')
            
            if Employee.objects.filter(national_id=national_id).exists():
                messages.error(request, 'رقم الهوية موجود بالفعل')
                return redirect('add_employee')
            
            # إنشاء المستخدم
            user = User.objects.create(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                password=make_password(password),
                is_active=can_login
            )
            
            # الحصول على الشركة الحالية
            company = getattr(request, 'company', None)
            if not company:
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
            
            # إنشاء الموظف
            employee = Employee.objects.create(
                company=company,
                user=user,
                national_id=national_id,
                phone=phone,
                address=address,
                birth_date=datetime.strptime(birth_date, '%Y-%m-%d').date() if birth_date else None,
                hire_date=datetime.strptime(hire_date, '%Y-%m-%d').date(),
                department=department,
                position=position,
                basic_salary=basic_salary,
                overtime_rate=overtime_rate,
                employment_type=employment_type,
                branch_id=branch_id if branch_id else None,
                fingerprint_id=fingerprint_id,
                created_by=request.user
            )
            
            # إنشاء ملف المستخدم
            company = getattr(request, 'company', None)
            if not company:
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
            
            UserProfile.objects.create(
                user=user,
                company=company,
                default_branch_id=branch_id if branch_id else None
            )
            
            # إنشاء حساب محاسبي
            if create_account:
                account_code = f"2101{employee.employee_id[-3:]}"
                Account.objects.create(
                    account_code=account_code,
                    name=f"راتب {user.get_full_name()}",
                    account_type='expense',
                    balance=0
                )
            
            # إضافة مندوب مبيعات
            if is_sales_rep:
                SalesRep.objects.create(
                    employee=employee,
                    commission_rate=commission_rate,
                    target_amount=0
                )
            
            messages.success(request, f'تم إضافة الموظف {user.get_full_name()} برقم {employee.employee_id} بنجاح')
            return redirect('employees')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    # جلب الإعدادات الحالية
    current_settings = {}
    for setting in Setting.objects.all():
        current_settings[setting.key] = setting.get_value()
    
    context = {
        'branches': Branch.objects.all(),
        'current_settings': current_settings,
        **get_user_context(request)
    }
    return render(request, 'add_employee.html', context)

@login_required
@subscription_required
def employees_list(request):
    employees = Employee.objects.select_related('user', 'branch').all().order_by('-created_at')
    
    search = request.GET.get('search', '')
    if search:
        employees = employees.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(national_id__icontains=search)
        )
    
    context = {
        'employees': employees,
        'search': search,
        **get_user_context(request)
    }
    return render(request, 'employees.html', context)



@login_required
@subscription_required
@permission_required('stock', 'view')
def stock_list(request):
    products = Product.objects.all().order_by('name')
    stock_data = []
    total_stock_value = 0
    
    for product in products:
        try:
            # تحديث المخزون من المبيعات والمشتريات
            confirmed_sales = SaleItem.objects.filter(
                product=product, 
                sale__status='confirmed'
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            confirmed_purchases = PurchaseItem.objects.filter(
                product=product, 
                purchase__status='confirmed'
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            # المخزون الحالي = المشتريات - المبيعات + المخزون الأولي
            initial_stock = getattr(product, 'stock', 0) or 0
            current_stock = initial_stock + confirmed_purchases - confirmed_sales
            
            # تحديث المخزون في المنتج
            product.stock = current_stock
            product.save()
            
            # حساب قيمة المخزون
            cost_price = getattr(product, 'cost_price', 0) or 0
            stock_value = current_stock * cost_price
            total_stock_value += stock_value
            
            stock = ProductStock.objects.filter(product=product).first()
            if stock:
                stock.current_stock = current_stock
                stock.save()
                min_stock = stock.min_stock
                warehouse = stock.warehouse
            else:
                min_stock = 10  # الحد الأدنى الافتراضي
                warehouse = None
            
            stock_data.append({
                'product': product,
                'current_stock': current_stock,
                'min_stock': min_stock,
                'warehouse': warehouse,
                'is_low_stock': current_stock <= min_stock,
                'stock_value': stock_value,
                'sales_count': confirmed_sales,
                'purchases_count': confirmed_purchases
            })
        except:
            current_stock = getattr(product, 'stock', 0) or 0
            stock_data.append({
                'product': product,
                'current_stock': current_stock,
                'min_stock': 0,
                'warehouse': None,
                'is_low_stock': False,
                'stock_value': 0,
                'sales_count': 0,
                'purchases_count': 0
            })
    
    recent_movements = StockMovement.objects.select_related('product').order_by('-created_at')[:10]
    
    # إحصائيات المخازن
    warehouses_stats = []
    for warehouse in Warehouse.objects.all():
        warehouse_products = ProductStock.objects.filter(warehouse=warehouse).count()
        warehouse_value = sum(
            (stock.current_stock * (stock.product.cost_price or 0))
            for stock in ProductStock.objects.filter(warehouse=warehouse)
        )
        warehouses_stats.append({
            'warehouse': warehouse,
            'products_count': warehouse_products,
            'total_value': warehouse_value
        })
    
    context = {
        'stock_data': stock_data,
        'recent_movements': recent_movements,
        'total_products': len(stock_data),
        'low_stock_count': sum(1 for item in stock_data if item['is_low_stock']),
        'total_stock_value': total_stock_value,
        'warehouses': Warehouse.objects.all(),
        'warehouses_stats': warehouses_stats,
        **get_user_context(request)
    }
    return render(request, 'stock.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def accounts(request):
    # جلب العملة من الإعدادات
    current_branch = getattr(request, 'current_branch', None)
    currency_symbol = DynamicSettingsManager.get('currency_symbol', current_branch, 'ر.س')
    
    # جلب الحسابات
    accounts_list = Account.objects.all().order_by('account_code')
    
    # إحصائيات الحسابات
    stats = {
        'total_accounts': Account.objects.count(),
        'customers_count': Customer.objects.count(),
        'suppliers_count': Supplier.objects.count(),
        'total_sales': Sale.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_purchases': Purchase.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
    }
    
    context = {
        'accounts': accounts_list,
        'stats': stats,
        'currency_symbol': currency_symbol,
        **get_user_context(request)
    }
    return render(request, 'accounts.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'add')
def add_account(request):
    if request.method == 'POST':
        try:
            # التحقق من عدم تكرار رمز الحساب
            account_code = request.POST.get('account_code')
            if Account.objects.filter(account_code=account_code).exists():
                messages.error(request, f'رمز الحساب "{account_code}" موجود بالفعل')
                return redirect('add_account')
            
            # الحصول على الشركة
            company = getattr(request, 'company', None)
            if not company:
                from datetime import date, timedelta
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
            
            Account.objects.create(
                company=company,
                account_code=account_code,
                name=request.POST.get('name'),
                account_type=request.POST.get('account_type'),
                parent_account_id=request.POST.get('parent_account_id') or None,
                balance=float(request.POST.get('balance', 0)),
                opening_balance=float(request.POST.get('balance', 0)),
                debit_balance=0,
                credit_balance=0,
                auto_update=True
            )
            messages.success(request, 'تم إضافة الحساب بنجاح')
            return redirect('accounts')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'account_types': Account.ACCOUNT_TYPES,
        'parent_accounts': Account.objects.all().order_by('account_code'),
        **get_user_context(request)
    }
    return render(request, 'add_account.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'edit')
def edit_account(request, account_id):
    account = get_object_or_404(Account, id=account_id)
    
    # جلب العملة من الإعدادات
    current_branch = getattr(request, 'current_branch', None)
    currency_symbol = DynamicSettingsManager.get('currency_symbol', current_branch, 'ر.س')
    
    if request.method == 'POST':
        try:
            # التحقق من البيانات المطلوبة
            account_code = request.POST.get('account_code', '').strip()
            name = request.POST.get('name', '').strip()
            account_type = request.POST.get('account_type', '')
            
            if not account_code or not name or not account_type:
                messages.error(request, 'يرجى ملء جميع الحقول المطلوبة')
                return render(request, 'edit_account.html', {
                    'account': account,
                    'account_types': Account.ACCOUNT_TYPES,
                    'parent_accounts': Account.objects.exclude(id=account.id).order_by('account_code'),
                    'currency_symbol': currency_symbol,
                    **get_user_context(request)
                })
            
            # التحقق من عدم تكرار رمز الحساب
            if account_code != account.account_code:
                if Account.objects.filter(account_code=account_code).exists():
                    messages.error(request, f'رمز الحساب "{account_code}" موجود بالفعل')
                    return render(request, 'edit_account.html', {
                        'account': account,
                        'account_types': Account.ACCOUNT_TYPES,
                        'parent_accounts': Account.objects.exclude(id=account.id).order_by('account_code'),
                        'currency_symbol': currency_symbol,
                        **get_user_context(request)
                    })
            
            # تحديث بيانات الحساب
            account.account_code = account_code
            account.name = name
            account.account_type = account_type
            
            # الحساب الأب (اختياري)
            parent_account_id = request.POST.get('parent_account_id')
            if parent_account_id and parent_account_id.strip():
                try:
                    parent_account = Account.objects.get(id=int(parent_account_id))
                    # التأكد من عدم جعل الحساب أب لنفسه
                    if parent_account.id == account.id:
                        messages.error(request, 'لا يمكن جعل الحساب أب لنفسه')
                        return render(request, 'edit_account.html', {
                            'account': account,
                            'account_types': Account.ACCOUNT_TYPES,
                            'parent_accounts': Account.objects.exclude(id=account.id).order_by('account_code'),
                            'currency_symbol': currency_symbol,
                            **get_user_context(request)
                        })
                    account.parent_account = parent_account
                except (ValueError, Account.DoesNotExist):
                    account.parent_account = None
            else:
                account.parent_account = None
            
            # الرصيد
            try:
                balance = float(request.POST.get('balance', 0))
                account.balance = balance
                # تحديث الرصيد الافتتاحي إذا لم يكن محدد مسبقاً
                if not account.opening_balance:
                    account.opening_balance = balance
            except (ValueError, TypeError):
                account.balance = 0
            
            # حالة النشاط
            account.is_active = request.POST.get('is_active') == 'on'
            
            # حفظ التغييرات
            account.save()
            
            messages.success(request, f'تم تحديث الحساب "{account.name}" بنجاح')
            
            # إذا كان الطلب AJAX، أرجع JSON
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'success': True, 
                    'message': f'تم تحديث الحساب "{account.name}" بنجاح'
                })
            
            return redirect('accounts')
            
        except Exception as e:
            error_msg = f'خطأ في تحديث الحساب: {str(e)}'
            messages.error(request, error_msg)
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': error_msg})
    
    # جلب الحسابات الأب (استبعاد الحساب الحالي لمنع التكرار)
    parent_accounts = Account.objects.exclude(id=account.id).order_by('account_code')
    
    context = {
        'account': account,
        'account_types': Account.ACCOUNT_TYPES,
        'parent_accounts': parent_accounts,
        'currency_symbol': currency_symbol,
        **get_user_context(request)
    }
    return render(request, 'edit_account.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'delete')
@csrf_exempt
def delete_account(request, account_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        account = get_object_or_404(Account, id=account_id)
        account_name = account.name
        account.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف الحساب "{account_name}" بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# حسابات العملاء
@login_required
@subscription_required
def customer_accounts(request):
    customers = Customer.objects.all().order_by('name')
    
    customer_data = []
    for customer in customers:
        total_sales = Sale.objects.filter(customer=customer, status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        customer_data.append({
            'customer': customer,
            'total_sales': total_sales,
            'balance': customer.opening_balance + total_sales
        })
    
    context = {
        'customer_data': customer_data,
        **get_user_context(request)
    }
    return render(request, 'customer_accounts.html', context)

# حسابات الموردين
@login_required
@subscription_required
def supplier_accounts(request):
    suppliers = Supplier.objects.all().order_by('name')
    
    supplier_data = []
    for supplier in suppliers:
        total_purchases = Purchase.objects.filter(supplier=supplier, status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        supplier_payments = SupplierPayment.objects.filter(supplier=supplier).aggregate(Sum('amount'))['amount__sum'] or 0
        supplier_data.append({
            'supplier': supplier,
            'total_purchases': total_purchases,
            'total_payments': supplier_payments,
            'balance': supplier.opening_balance + total_purchases - supplier_payments
        })
    
    context = {
        'supplier_data': supplier_data,
        **get_user_context(request)
    }
    return render(request, 'supplier_accounts.html', context)

@login_required
@subscription_required
def accounting_tree(request):
    from django.db.models import Sum
    
    accounts_by_type = {
        'assets': Account.objects.filter(account_type='asset').order_by('account_code'),
        'liabilities': Account.objects.filter(account_type='liability').order_by('account_code'),
        'equity': Account.objects.filter(account_type='equity').order_by('account_code'),
        'revenue': Account.objects.filter(account_type='revenue').order_by('account_code'),
        'expenses': Account.objects.filter(account_type='expense').order_by('account_code'),
    }
    
    # حساب الأرصدة مع تحديث من المبيعات والمشتريات
    # تحديث أرصدة العملاء
    for customer in Customer.objects.all():
        account = get_or_create_customer_account(customer)
        total_sales = Sale.objects.filter(customer=customer, status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        payments = CustomerPayment.objects.filter(customer=customer).aggregate(Sum('amount'))['amount__sum'] or 0
        account.balance = customer.opening_balance + total_sales - payments
        account.save()
    
    # تحديث أرصدة الموردين
    for supplier in Supplier.objects.all():
        account = get_or_create_supplier_account(supplier)
        total_purchases = Purchase.objects.filter(supplier=supplier, status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        payments = SupplierPayment.objects.filter(supplier=supplier).aggregate(Sum('amount'))['amount__sum'] or 0
        account.balance = supplier.opening_balance + total_purchases - payments
        account.save()
    
    # الحصول على الشركة الحالية
    company = getattr(request, 'company', None)
    if not company:
        try:
            company = Company.objects.first()
        except:
            from datetime import date, timedelta
            company = Company.objects.create(
                code='DEFAULT',
                name='الشركة الافتراضية',
                database_name='erp_default',
                subscription_end=date.today() + timedelta(days=365)
            )
    
    # تحديث حساب المبيعات
    sales_account = get_or_create_account('4001', 'مبيعات', 'revenue', company)
    total_sales = Sale.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    sales_account.balance = total_sales
    sales_account.save()
    
    # تحديث حساب المشتريات
    purchases_account = get_or_create_account('5001', 'مشتريات', 'expense', company)
    total_purchases = Purchase.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    purchases_account.balance = total_purchases
    purchases_account.save()
    
    # تحديث حساب المخزون
    inventory_account = get_or_create_account('1301', 'مخزون بضاعة', 'asset', company)
    total_inventory_value = 0
    for product in Product.objects.all():
        stock = getattr(product, 'stock', 0) or 0
        cost_price = getattr(product, 'cost_price', 0) or 0
        total_inventory_value += stock * cost_price
    inventory_account.balance = total_inventory_value
    inventory_account.save()
    
    # إحصائيات محاسبية
    stats = {
        'total_accounts': Account.objects.count(),
        'assets_total': accounts_by_type['assets'].aggregate(Sum('balance'))['balance__sum'] or 0,
        'liabilities_total': accounts_by_type['liabilities'].aggregate(Sum('balance'))['balance__sum'] or 0,
        'equity_total': accounts_by_type['equity'].aggregate(Sum('balance'))['balance__sum'] or 0,
        'revenue_total': accounts_by_type['revenue'].aggregate(Sum('balance'))['balance__sum'] or 0,
        'expenses_total': accounts_by_type['expenses'].aggregate(Sum('balance'))['balance__sum'] or 0,
        'total_sales': total_sales,
        'total_purchases': total_purchases,
        'customers_count': Customer.objects.count(),
        'suppliers_count': Supplier.objects.count(),
        'employees_count': User.objects.filter(is_active=True).count(),
        'products_count': Product.objects.count(),
    }
    
    # آخر القيود
    recent_entries = []
    
    context = {
        'accounts_by_type': accounts_by_type,
        'stats': stats,
        'recent_entries': recent_entries,
        **get_user_context(request)
    }
    
    return render(request, 'accounting_tree.html', context)

# ملخص الحسابات
@login_required
@subscription_required
def accounts_summary(request):
    from django.db.models import Sum
    from datetime import date
    
    # الإحصائيات الأساسية
    summary = {
        'total_customers': Customer.objects.count(),
        'total_suppliers': Supplier.objects.count(),
        'total_sales': Sale.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_purchases': Purchase.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_accounts': Account.objects.count(),
    }
    
    # حساب صافي الربح
    net_profit = summary['total_sales'] - summary['total_purchases']
    
    # أرصدة شجرة الحسابات
    try:
        assets_total = Account.objects.filter(account_type='asset').aggregate(Sum('balance'))['balance__sum'] or 0
        liabilities_total = Account.objects.filter(account_type='liability').aggregate(Sum('balance'))['balance__sum'] or 0
        equity_total = Account.objects.filter(account_type='equity').aggregate(Sum('balance'))['balance__sum'] or 0
        revenue_total = Account.objects.filter(account_type='revenue').aggregate(Sum('balance'))['balance__sum'] or 0
        expenses_total = Account.objects.filter(account_type='expense').aggregate(Sum('balance'))['balance__sum'] or 0
    except:
        assets_total = liabilities_total = equity_total = revenue_total = expenses_total = 0
    
    # إحصائيات إضافية
    sales_count = Sale.objects.filter(status='confirmed').count()
    purchases_count = Purchase.objects.filter(status='confirmed').count()
    avg_sale = summary['total_sales'] / sales_count if sales_count > 0 else 0
    avg_purchase = summary['total_purchases'] / purchases_count if purchases_count > 0 else 0
    
    # حسابات المناديب
    try:
        sales_reps_count = SalesRep.objects.filter(is_active=True).count()
        total_commissions = 0  # يمكن حسابها لاحقاً
        reps_sales = Sale.objects.filter(sales_rep__isnull=False, status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    except:
        sales_reps_count = total_commissions = reps_sales = 0
    
    # المبالغ المستحقة
    customers_due = summary['total_sales'] * 0.3  # تقدير 30%
    suppliers_due = summary['total_purchases'] * 0.2  # تقدير 20%
    
    # الحسابات النشطة
    active_accounts = Account.objects.filter(is_active=True).count() if hasattr(Account, 'is_active') else summary['total_accounts']
    
    context = {
        'summary': summary,
        'net_profit': net_profit,
        'assets_total': assets_total,
        'liabilities_total': liabilities_total,
        'equity_total': equity_total,
        'revenue_total': revenue_total,
        'expenses_total': expenses_total,
        'sales_count': sales_count,
        'purchases_count': purchases_count,
        'avg_sale': avg_sale,
        'avg_purchase': avg_purchase,
        'sales_reps_count': sales_reps_count,
        'total_commissions': total_commissions,
        'reps_sales': reps_sales,
        'customers_due': customers_due,
        'suppliers_due': suppliers_due,
        'active_accounts': active_accounts,
        'last_update': date.today(),
        'currency_symbol': get_setting('currency_symbol', 'ر.س'),
        **get_user_context(request)
    }
    return render(request, 'accounts_summary.html', context)

# حسابات المبيعات
@login_required
@subscription_required
def sales_accounts(request):
    sales = Sale.objects.filter(status='confirmed').order_by('-created_at')[:50]
    stats = {
        'total_sales': Sale.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_invoices': Sale.objects.filter(status='confirmed').count()
    }
    
    context = {
        'sales': sales,
        'stats': stats,
        **get_user_context(request)
    }
    return render(request, 'sales_accounts.html', context)

# حسابات المشتريات
@login_required
@subscription_required
def purchase_accounts(request):
    purchases = Purchase.objects.filter(status='confirmed').order_by('-created_at')[:50]
    stats = {
        'total_purchases': Purchase.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_invoices': Purchase.objects.filter(status='confirmed').count()
    }
    
    context = {
        'purchases': purchases,
        'stats': stats,
        **get_user_context(request)
    }
    return render(request, 'purchase_accounts.html', context)

# حسابات المناديب
@login_required
@subscription_required
def sales_rep_accounts(request):
    sales_reps = SalesRep.objects.filter(is_active=True).select_related('employee__user')
    
    context = {
        'sales_reps': sales_reps,
        **get_user_context(request)
    }
    return render(request, 'sales_rep_accounts.html', context)

@login_required
@subscription_required
def journal_entries_list(request):
    context = {
        'entries': [],
        **get_user_context(request)
    }
    return render(request, 'journal_entries.html', context)

@login_required
@subscription_required
def add_journal_entry(request):
    if request.method == 'POST':
        messages.success(request, 'سيتم إضافة هذه الميزة قريباً')
        return redirect('journal_entries')
    
    context = {
        'accounts': Account.objects.all().order_by('account_code'),
        **get_user_context(request)
    }
    return render(request, 'add_journal_entry.html', context)

@login_required
@subscription_required
def vouchers_list(request):
    context = {
        'vouchers': [],
        **get_user_context(request)
    }
    return render(request, 'vouchers.html', context)

@login_required
@subscription_required
def add_voucher(request):
    if request.method == 'POST':
        messages.success(request, 'سيتم إضافة هذه الميزة قريباً')
        return redirect('vouchers')
    
    context = get_user_context(request)
    return render(request, 'add_voucher.html', context)

@login_required
@subscription_required
@permission_required('accounts', 'view')
def trial_balance(request):
    from django.db.models import Sum
    
    accounts = Account.objects.all().order_by('account_code')
    total_debits = 0
    total_credits = 0
    
    account_balances = []
    for account in accounts:
        # تحديد المدين والدائن حسب نوع الحساب
        if account.account_type in ['asset', 'expense']:
            debit_balance = float(account.balance) if account.balance > 0 else 0
            credit_balance = 0
        else:  # liability, equity, revenue
            debit_balance = 0
            credit_balance = float(account.balance) if account.balance > 0 else 0
        
        total_debits += debit_balance
        total_credits += credit_balance
        
        # إضافة الحساب فقط إذا كان له رصيد
        if debit_balance > 0 or credit_balance > 0:
            account_balances.append({
                'account': account,
                'debit': debit_balance,
                'credit': credit_balance
            })
    
    # فحص التوازن
    is_balanced = abs(total_debits - total_credits) < 0.01
    balance_difference = total_debits - total_credits
    
    # الحصول على العملة
    currency_symbol = get_setting('currency_symbol', 'د.ك')
    
    context = {
        'account_balances': account_balances,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'is_balanced': is_balanced,
        'balance_difference': balance_difference,
        'currency_symbol': currency_symbol,
        **get_user_context(request)
    }
    return render(request, 'trial_balance.html', context)

@login_required
@subscription_required
def account_ledger(request, account_id):
    account = get_object_or_404(Account, id=account_id)
    
    context = {
        'account': account,
        'entries': [],
        **get_user_context(request)
    }
    return render(request, 'account_ledger.html', context)

@login_required
@subscription_required
def customer_statement(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    
    # جلب الفواتير والدفعات
    sales = Sale.objects.filter(customer=customer, status='confirmed').order_by('-created_at')
    payments = CustomerPayment.objects.filter(customer=customer).order_by('-created_at')
    
    # دمج الفواتير والدفعات في قائمة واحدة
    transactions = []
    
    # إضافة الفواتير
    for sale in sales:
        transactions.append({
            'date': sale.created_at.date(),
            'description': f'فاتورة بيع #{sale.invoice_number}',
            'debit': sale.total_amount,
            'credit': 0,
            'type': 'sale',
            'payment_method': '-'
        })
    
    # إضافة الدفعات
    for payment in payments:
        transactions.append({
            'date': payment.payment_date,
            'description': f'دفعة #{payment.payment_number}',
            'debit': 0,
            'credit': payment.amount,
            'type': 'payment',
            'payment_method': payment.get_payment_method_display()
        })
    
    # ترتيب المعاملات حسب التاريخ
    transactions.sort(key=lambda x: x['date'], reverse=True)
    
    # حساب الرصيد التراكمي
    running_balance = customer.opening_balance
    for transaction in reversed(transactions):
        running_balance += transaction['debit'] - transaction['credit']
        transaction['balance'] = running_balance
    
    context = {
        'customer': customer,
        'transactions': transactions,
        'opening_balance': customer.opening_balance,
        'final_balance': running_balance,
        **get_user_context(request)
    }
    return render(request, 'customer_statement.html', context)

@login_required
@subscription_required
def supplier_statement(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    
    # جلب المشتريات والدفعات
    purchases = Purchase.objects.filter(supplier=supplier, status='confirmed').order_by('-created_at')
    payments = SupplierPayment.objects.filter(supplier=supplier).order_by('-created_at')
    
    # دمج المشتريات والدفعات في قائمة واحدة
    transactions = []
    
    # إضافة المشتريات
    for purchase in purchases:
        transactions.append({
            'date': purchase.created_at.date(),
            'description': f'فاتورة شراء #{purchase.invoice_number}',
            'debit': 0,
            'credit': purchase.total_amount,
            'type': 'purchase',
            'payment_method': '-'
        })
    
    # إضافة الدفعات
    for payment in payments:
        transactions.append({
            'date': payment.payment_date,
            'description': f'دفعة #{payment.payment_number}',
            'debit': payment.amount,
            'credit': 0,
            'type': 'payment',
            'payment_method': payment.get_payment_method_display()
        })
    
    # ترتيب المعاملات حسب التاريخ
    transactions.sort(key=lambda x: x['date'], reverse=True)
    
    # حساب الرصيد التراكمي
    running_balance = supplier.opening_balance
    for transaction in reversed(transactions):
        running_balance += transaction['credit'] - transaction['debit']
        transaction['balance'] = running_balance
    
    context = {
        'supplier': supplier,
        'transactions': transactions,
        'opening_balance': supplier.opening_balance,
        'final_balance': running_balance,
        **get_user_context(request)
    }
    return render(request, 'supplier_statement.html', context)

@login_required
@subscription_required
def balance_sheet(request):
    from django.db.models import Sum
    
    assets = Account.objects.filter(account_type='asset').aggregate(Sum('balance'))['balance__sum'] or 0
    liabilities = Account.objects.filter(account_type='liability').aggregate(Sum('balance'))['balance__sum'] or 0
    equity = Account.objects.filter(account_type='equity').aggregate(Sum('balance'))['balance__sum'] or 0
    
    context = {
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity,
        'total_liabilities_equity': liabilities + equity,
        **get_user_context(request)
    }
    return render(request, 'balance_sheet.html', context)

@login_required
@subscription_required
def manufacturing_list(request):
    context = get_user_context(request)
    return render(request, 'manufacturing.html', context)

@login_required
@subscription_required
@permission_required('permissions', 'view')
@login_required
@subscription_required
@permission_required('permissions', 'view')
def permissions_list(request):
    users = User.objects.all().order_by('username')
    
    # تحضير بيانات الصلاحيات لكل مستخدم
    users_with_permissions = []
    for user in users:
        user_perms = {}
        permissions = Permission.objects.filter(user=user)
        for perm in permissions:
            user_perms[perm.screen] = {
                'can_view': perm.can_view,
                'can_add': perm.can_add,
                'can_edit': perm.can_edit,
                'can_delete': perm.can_delete,
                'can_confirm': perm.can_confirm,
                'can_print': perm.can_print,
                'can_export': perm.can_export,
            }
        users_with_permissions.append({
            'user': user,
            'permissions': user_perms
        })
    
    # حساب عدد المستخدمين النشطين
    active_users_count = users.filter(is_active=True).count()
    
    context = {
        'users': users,
        'users_with_permissions': users_with_permissions,
        'screen_choices': Permission.SCREEN_CHOICES,
        'branches': Branch.objects.all(),
        'warehouses': Warehouse.objects.all(),
        'active_users_count': active_users_count,
        **get_user_context(request)
    }
    return render(request, 'permissions.html', context)

@login_required
@subscription_required
@permission_required('stock', 'edit')
@csrf_exempt
def adjust_stock(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        product_id = request.POST.get('product_id')
        actual_quantity = float(request.POST.get('actual_quantity', 0))
        reason = request.POST.get('reason', 'جرد مخزون')
        
        product = get_object_or_404(Product, id=product_id)
        
        # استخدام نظام ربط المخزون بالمحاسبة للجرد
        try:
            entry = InventoryAccountingManager.process_inventory_adjustment(
                product, actual_quantity, request.user, reason
            )
            
            if entry:
                return JsonResponse({
                    'success': True,
                    'message': f'تم جرد مخزون {product.name} وتسجيل القيود المحاسبية',
                    'new_stock': float(product.stock or 0)
                })
            else:
                return JsonResponse({
                    'success': True,
                    'message': f'لا يوجد فرق في مخزون {product.name}',
                    'new_stock': float(product.stock or 0)
                })
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'خطأ في الجرد: {str(e)}'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@permission_required('stock', 'export')
def export_stock(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "المخزون"
        
        headers = ['الرقم', 'اسم المنتج', 'الباركود', 'الفئة', 'الوحدة', 'المخزون الحالي', 'الحد الأدنى', 'الحالة']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        products = Product.objects.all()
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product.id)
            ws.cell(row=row, column=2, value=product.name)
            ws.cell(row=row, column=3, value=product.barcode or '')
            ws.cell(row=row, column=4, value=product.category or '')
            ws.cell(row=row, column=5, value=product.unit or '')
            ws.cell(row=row, column=6, value=float(getattr(product, 'stock', 0) or 0))
            ws.cell(row=row, column=7, value=10)  # الحد الأدنى الافتراضي
            ws.cell(row=row, column=8, value='نشط' if product.is_active else 'غير نشط')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="stock.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('stock')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('stock')

def get_product_info(request, product_id):
    return JsonResponse({'product': {}})

def get_realtime_updates(request):
    return JsonResponse({'updates': []})

def refresh_data(request):
    return JsonResponse({'success': True})

# API لجلب تفاصيل العميل
@login_required
def get_customer_invoices(request, customer_id):
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        invoices = Sale.objects.filter(customer=customer).order_by('-created_at')[:10]
        
        invoices_data = []
        for invoice in invoices:
            invoices_data.append({
                'invoice_number': invoice.invoice_number,
                'date': invoice.created_at.strftime('%Y-%m-%d'),
                'total_amount': float(invoice.total_amount),
                'paid_amount': float(getattr(invoice, 'paid_amount', 0)),
                'remaining_amount': float(invoice.total_amount - getattr(invoice, 'paid_amount', 0)),
                'status': invoice.status,
                'status_display': dict(Sale.INVOICE_STATUS).get(invoice.status, invoice.status)
            })
        
        return JsonResponse({
            'success': True,
            'invoices': invoices_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_customer_returns(request, customer_id):
    try:
        # مؤقتاً - بيانات وهمية حتى يتم إنشاء نموذج المرتجعات
        returns_data = [
            {
                'return_number': 'RET-001',
                'date': '2024-01-15',
                'amount': 150.00,
                'reason': 'عيب في المنتج'
            },
            {
                'return_number': 'RET-002', 
                'date': '2024-01-10',
                'amount': 75.50,
                'reason': 'منتج غير مطابق'
            }
        ]
        
        return JsonResponse({
            'success': True,
            'returns': returns_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_customer_payments(request, customer_id):
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        payments = CustomerPayment.objects.filter(customer=customer).order_by('-created_at')[:10]
        
        payments_data = []
        for payment in payments:
            payments_data.append({
                'payment_number': payment.payment_number,
                'date': payment.payment_date.strftime('%Y-%m-%d'),
                'amount': float(payment.amount),
                'method': payment.get_payment_method_display(),
                'reference_number': payment.reference_number or '',
                'notes': payment.notes or ''
            })
        
        return JsonResponse({
            'success': True,
            'payments': payments_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_customer_summary(request, customer_id):
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        
        # حساب الإحصائيات
        total_invoices = Sale.objects.filter(customer=customer).count()
        confirmed_sales = Sale.objects.filter(customer=customer, status='confirmed')
        total_sales = confirmed_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        paid_amount = confirmed_sales.aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0
        
        # حساب الدفعات
        customer_payments = CustomerPayment.objects.filter(customer=customer).aggregate(Sum('amount'))['amount__sum'] or 0
        
        due_amount = total_sales - paid_amount - customer_payments
        
        summary_data = {
            'total_invoices': total_invoices,
            'total_sales': float(total_sales),
            'total_returns': 0,  # مؤقت
            'total_payments': float(paid_amount + customer_payments),
            'due_amount': float(due_amount),
            'current_balance': float(customer.opening_balance + total_sales - paid_amount - customer_payments)
        }
        
        return JsonResponse({
            'success': True,
            'summary': summary_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@permission_required('permissions', 'edit')
@csrf_exempt
def update_permission(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        user_id = request.POST.get('user_id')
        screen = request.POST.get('screen')
        action = request.POST.get('action')
        value = request.POST.get('value') == 'true'
        
        user = get_object_or_404(User, id=user_id)
        permission, created = Permission.objects.get_or_create(
            user=user,
            screen=screen,
            defaults={'can_view': True}
        )
        
        setattr(permission, f'can_{action}', value)
        permission.save()
        
        return JsonResponse({
            'success': True,
            'message': f'تم تحديث صلاحية {action} للمستخدم {user.username}'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# إدارة المناديب
@login_required
@subscription_required
@permission_required('sales_reps', 'view')
def sales_reps_list(request):
    sales_reps = SalesRep.objects.select_related('employee__user').filter(is_active=True).order_by('-created_at')
    
    search = request.GET.get('search', '')
    if search:
        sales_reps = sales_reps.filter(
            Q(employee__user__first_name__icontains=search) |
            Q(employee__user__last_name__icontains=search) |
            Q(employee__employee_id__icontains=search)
        )
    
    # جلب الإعدادات الحالية
    current_settings = {}
    for setting in Setting.objects.all():
        current_settings[setting.key] = setting.get_value()
    
    context = {
        'sales_reps': sales_reps,
        'search': search,
        'current_settings': current_settings,
        **get_user_context(request)
    }
    return render(request, 'sales_reps.html', context)

@login_required
@subscription_required
@permission_required('sales_reps', 'add')
def add_sales_rep(request):
    if request.method == 'POST':
        try:
            employee_id = request.POST.get('employee')
            commission_rate = request.POST.get('commission_rate', 0)
            target_amount = request.POST.get('target_amount', 0)
            
            # الحصول على الموظف
            try:
                employee = Employee.objects.get(id=employee_id)
            except Employee.DoesNotExist:
                messages.error(request, 'الموظف غير موجود')
                return redirect('add_sales_rep')
            
            if SalesRep.objects.filter(employee=employee).exists():
                messages.error(request, 'هذا الموظف مندوب مبيعات بالفعل')
                return redirect('add_sales_rep')
            
            # إنشاء المندوب
            sales_rep = SalesRep(
                employee=employee,
                commission_rate=commission_rate,
                target_amount=target_amount
            )
            # إضافة employee_code يدوياً إذا كان مطلوباً
            if hasattr(sales_rep, 'employee_code') and 'employee_code' in [f.name for f in SalesRep._meta.fields]:
                sales_rep.employee_code = employee.employee_id
            sales_rep.save()
            
            messages.success(request, f'تم إضافة المندوب بنجاح')
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'success': True, 
                    'message': f'تم إضافة المندوب بنجاح'
                })
            
            return redirect('sales_reps')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': str(e)})
    
    # جلب جميع الموظفين بدون فلترة
    available_employees = Employee.objects.select_related('user').all().order_by('user__first_name')
    
    # توليد رمز الموظف التالي للعرض
    try:
        last_rep = SalesRep.objects.order_by('-id').first()
        if last_rep and hasattr(last_rep, 'employee') and last_rep.employee:
            next_employee_code = f"REP{SalesRep.objects.count() + 1:03d}"
        else:
            next_employee_code = "REP001"
    except:
        next_employee_code = "REP001"
    
    # جلب الإعدادات الحالية
    current_settings = {}
    for setting in Setting.objects.all():
        current_settings[setting.key] = setting.get_value()
    
    context = {
        'available_employees': available_employees,
        'employees': available_employees,  # إضافة للتوافق مع القالب
        'next_employee_code': next_employee_code,
        'current_settings': current_settings,
        'total_employees': available_employees.count(),
        **get_user_context(request)
    }
    return render(request, 'add_sales_rep.html', context)

@login_required
@subscription_required
@permission_required('sales_reps', 'edit')
def edit_sales_rep(request, rep_id):
    sales_rep = get_object_or_404(SalesRep, id=rep_id)
    
    if request.method == 'POST':
        try:
            commission_rate = request.POST.get('commission_rate', 0)
            target_amount = request.POST.get('target_amount', 0)
            is_active = request.POST.get('is_active') == 'on'
            
            # تحديث البيانات بدون رمز الموظف (غير قابل للتعديل)
            sales_rep.commission_rate = commission_rate
            sales_rep.target_amount = target_amount
            sales_rep.is_active = is_active
            sales_rep.save()
            
            messages.success(request, 'تم تحديث بيانات المندوب بنجاح')
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'تم تحديث بيانات المندوب بنجاح'})
            
            return redirect('sales_reps')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'message': str(e)})
    
    # جلب الإعدادات الحالية
    current_settings = {}
    for setting in Setting.objects.all():
        current_settings[setting.key] = setting.get_value()
    
    context = {
        'sales_rep': sales_rep,
        'current_settings': current_settings,
        **get_user_context(request)
    }
    return render(request, 'edit_sales_rep.html', context)

@login_required
@subscription_required
@permission_required('sales_reps', 'delete')
@csrf_exempt
def delete_sales_rep(request, rep_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        sales_rep = get_object_or_404(SalesRep, id=rep_id)
        rep_name = sales_rep.employee.user.get_full_name() or sales_rep.employee.user.username
        
        # بدلاً من الحذف، نقوم بإلغاء التفعيل
        sales_rep.is_active = False
        sales_rep.save()
        
        return JsonResponse({
            'success': True,
            'message': f'تم إلغاء تفعيل المندوب "{rep_name}" بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@permission_required('sales_reps', 'export')
def export_sales_reps(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "مناديب المبيعات"
        
        headers = ['الرقم', 'اسم المندوب', 'رمز الموظف', 'نسبة العمولة %', 'الهدف الشهري', 'الحالة', 'تاريخ الإضافة']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        sales_reps = SalesRep.objects.select_related('employee__user').all()
        for row, rep in enumerate(sales_reps, 2):
            ws.cell(row=row, column=1, value=rep.id)
            ws.cell(row=row, column=2, value=rep.employee.user.get_full_name() or rep.employee.user.username)
            ws.cell(row=row, column=3, value=getattr(rep.employee, 'employee_id', 'N/A'))
            ws.cell(row=row, column=4, value=float(rep.commission_rate))
            ws.cell(row=row, column=5, value=float(rep.target_amount))
            ws.cell(row=row, column=6, value='نشط' if rep.is_active else 'غير نشط')
            ws.cell(row=row, column=7, value=rep.created_at.strftime('%Y-%m-%d'))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sales_reps.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('sales_reps')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('sales_reps')

# دفعات العملاء
@login_required
@subscription_required
def customer_payments_list(request):
    payments = CustomerPayment.objects.select_related('customer').all().order_by('-created_at')
    
    search = request.GET.get('search', '')
    if search:
        payments = payments.filter(
            Q(customer__name__icontains=search) |
            Q(payment_number__icontains=search)
        )
    
    context = {
        'payments': payments,
        'search': search,
        **get_user_context(request)
    }
    return render(request, 'customer_payments.html', context)

@login_required
@subscription_required
def add_customer_payment(request):
    if request.method == 'POST':
        try:
            customer_id = request.POST.get('customer_id')
            amount = float(request.POST.get('amount', 0))
            payment_method = request.POST.get('payment_method', 'cash')
            reference_number = request.POST.get('reference_number', '')
            notes = request.POST.get('notes', '')
            
            if amount <= 0:
                messages.error(request, 'يجب أن يكون المبلغ أكبر من صفر')
                return redirect('add_customer_payment')
            
            payment = CustomerPayment.objects.create(
                customer_id=customer_id,
                amount=amount,
                payment_method=payment_method,
                reference_number=reference_number,
                notes=notes,
                created_by=request.user
            )
            
            # تسجيل القيد المحاسبي تلقائياً
            try:
                InventoryAccountingManager.process_customer_payment(payment, request.user)
                InventoryAccountingManager.update_account_balances()
                messages.success(request, 'تم تسجيل الدفعة والقيد المحاسبي بنجاح')
            except Exception as e:
                messages.warning(request, f'تم تسجيل الدفعة ولكن حدث خطأ في المحاسبة: {str(e)}')
            
            return redirect('customer_payments')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'customers': Customer.objects.all().order_by('name'),
        **get_user_context(request)
    }
    return render(request, 'add_customer_payment.html', context)

@login_required
@csrf_exempt
def delete_customer_payment(request, payment_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        payment = get_object_or_404(CustomerPayment, id=payment_id)
        payment_number = payment.payment_number
        payment.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف الدفعة "{payment_number}" بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# دفعات الموردين
@login_required
@subscription_required
def supplier_payments_list(request):
    payments = SupplierPayment.objects.select_related('supplier').all().order_by('-created_at')
    
    search = request.GET.get('search', '')
    if search:
        payments = payments.filter(
            Q(supplier__name__icontains=search) |
            Q(payment_number__icontains=search)
        )
    
    context = {
        'payments': payments,
        'search': search,
        **get_user_context(request)
    }
    return render(request, 'supplier_payments.html', context)

@login_required
@subscription_required
def add_supplier_payment(request):
    if request.method == 'POST':
        try:
            supplier_id = request.POST.get('supplier_id')
            amount = float(request.POST.get('amount', 0))
            payment_method = request.POST.get('payment_method', 'cash')
            reference_number = request.POST.get('reference_number', '')
            notes = request.POST.get('notes', '')
            
            if amount <= 0:
                messages.error(request, 'يجب أن يكون المبلغ أكبر من صفر')
                return redirect('add_supplier_payment')
            
            payment = SupplierPayment.objects.create(
                supplier_id=supplier_id,
                amount=amount,
                payment_method=payment_method,
                reference_number=reference_number,
                notes=notes,
                created_by=request.user
            )
            
            # تسجيل القيد المحاسبي تلقائياً
            try:
                InventoryAccountingManager.process_supplier_payment(payment, request.user)
                InventoryAccountingManager.update_account_balances()
                messages.success(request, 'تم تسجيل الدفعة والقيد المحاسبي بنجاح')
            except Exception as e:
                messages.warning(request, f'تم تسجيل الدفعة ولكن حدث خطأ في المحاسبة: {str(e)}')
            
            return redirect('supplier_payments')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'suppliers': Supplier.objects.all().order_by('name'),
        **get_user_context(request)
    }
    return render(request, 'add_supplier_payment.html', context)

@login_required
@csrf_exempt
def delete_supplier_payment(request, payment_id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        payment = get_object_or_404(SupplierPayment, id=payment_id)
        payment_number = payment.payment_number
        payment.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف الدفعة "{payment_number}" بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# مرتجعات البيع
@login_required
@subscription_required
def sale_returns_list(request):
    returns = SaleReturn.objects.select_related('customer').order_by('-created_at')
    
    status = request.GET.get('status', '')
    if status:
        returns = returns.filter(status=status)
    
    search = request.GET.get('search', '')
    if search:
        returns = returns.filter(
            Q(return_number__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(customer_return_number__icontains=search)
        )
    
    context = {
        'returns': returns,
        'search': search,
        'status': status,
        'status_choices': [('draft', 'مسودة'), ('confirmed', 'مؤكد'), ('cancelled', 'ملغي')],
        **get_user_context(request)
    }
    return render(request, 'sale_returns.html', context)

# دالة تأكيد مرتجع البيع
@login_required
@subscription_required
def confirm_sale_return(request, return_id):
    try:
        sale_return = get_object_or_404(SaleReturn, id=return_id)
        if sale_return.status == 'confirmed':
            messages.error(request, 'المرتجع مؤكد بالفعل')
        else:
            sale_return.status = 'confirmed'
            sale_return.confirmed_by = request.user
            sale_return.confirmed_at = timezone.now()
            sale_return.save()
            
            # تحديث المخزون
            for item in sale_return.items.all():
                product = item.product
                if hasattr(product, 'stock'):
                    product.stock = (product.stock or 0) + item.quantity
                    product.save()
            
            messages.success(request, f'تم تأكيد مرتجع البيع #{sale_return.return_number} وتحديث المخزون')
    except Exception as e:
        messages.error(request, f'خطأ غير متوقع: {str(e)}')
    return redirect('sale_returns')

@login_required
@subscription_required
def add_sale_return(request):
    if request.method == 'POST':
        try:
            customer_id = request.POST.get('customer_id')
            sales_rep_id = request.POST.get('sales_rep_id')
            warehouse_id = request.POST.get('warehouse_id')
            customer_return_number = request.POST.get('customer_return_number', '')
            notes = request.POST.get('notes', '')
            
            if not customer_id:
                messages.error(request, 'يرجى اختيار عميل')
                return redirect('add_sale_return')
            
            # إنشاء مرتجع البيع
            # نحتاج فاتورة بيع وهمية مؤقتاً
            dummy_sale = Sale.objects.filter(customer_id=customer_id).first()
            if not dummy_sale:
                # إنشاء فاتورة وهمية
                dummy_sale = Sale.objects.create(
                    customer_id=customer_id,
                    invoice_number=f"TEMP-{timezone.now().timestamp()}",
                    total_amount=0,
                    created_by=request.user
                )
            
            sale_return = SaleReturn.objects.create(
                original_sale=dummy_sale,
                customer_id=customer_id,
                reason=notes,
                total_amount=0,
                created_by=request.user
            )
            
            # إضافة عناصر المرتجع
            items_data = json.loads(request.POST.get('items', '[]'))
            if not items_data:
                sale_return.delete()
                messages.error(request, 'يرجى إضافة منتجات للمرتجع')
                return redirect('add_sale_return')
            
            total_amount = 0
            for item_data in items_data:
                product_id = item_data['product_id']
                quantity = float(item_data['quantity'])
                price = float(item_data['price'])
                discount = float(item_data.get('discount', 0))
                
                # حساب الإجمالي مع الخصم
                subtotal = quantity * price
                discount_amount = subtotal * (discount / 100)
                item_total = subtotal - discount_amount
                
                # إنشاء عنصر المرتجع
                # نحتاج عنصر فاتورة وهمي
                dummy_sale_item = dummy_sale.items.first()
                if not dummy_sale_item:
                    dummy_sale_item = SaleItem.objects.create(
                        sale=dummy_sale,
                        product_id=product_id,
                        quantity=quantity,
                        unit_price=price,
                        total_price=item_total
                    )
                
                SaleReturnItem.objects.create(
                    sale_return=sale_return,
                    original_sale_item=dummy_sale_item,
                    product_id=product_id,
                    quantity=quantity,
                    unit_price=price,
                    total_price=item_total
                )
                
                total_amount += item_total
            
            # حساب الضريبة والإجمالي النهائي
            tax_rate = 0.15  # 15% ضريبة
            tax_amount = total_amount * tax_rate
            final_total = total_amount + tax_amount
            
            sale_return.subtotal = total_amount
            sale_return.tax_amount = tax_amount
            sale_return.total_amount = final_total
            sale_return.save()
            
            messages.success(request, f'تم إنشاء فاتورة مرتجع البيع #{sale_return.return_number} بنجاح')
            return redirect('sale_returns')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    # تحضير البيانات للصفحة
    from datetime import date
    
    # جلب العملة من الإعدادات
    currency_symbol = DynamicSettingsManager.get('currency_symbol', default='د.ك')
    
    context = {
        'customers': Customer.objects.all().order_by('name'),
        'sales_reps': SalesRep.objects.filter(is_active=True),
        'warehouses': Warehouse.objects.all().order_by('name'),
        'products': Product.objects.filter(is_active=True).order_by('name'),
        'today': date.today(),
        'currency_symbol': currency_symbol,
        **get_user_context(request)
    }
    return render(request, 'add_sale_return.html', context)

@login_required
@subscription_required
def confirm_sale_return_view(request, return_id):
    sale_return = get_object_or_404(SaleReturn, id=return_id)
    
    try:
        sale_return.status = 'confirmed'
        sale_return.confirmed_by = request.user
        sale_return.confirmed_at = timezone.now()
        sale_return.save()
        
        # ربط المخزون بالمحاسبة - قيد عكسي
        try:
            InventoryAccountingManager.process_sale_return(sale_return.items.all(), request.user)
            messages.success(request, f'تم تأكيد مرتجع البيع #{sale_return.return_number} وتحديث المخزون والقيود المحاسبية')
        except Exception as e:
            messages.warning(request, f'تم تأكيد المرتجع ولكن حدث خطأ في المحاسبة: {str(e)}')
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'خطأ: {str(e)}')
    
    return redirect('sale_returns')

# مرتجعات الشراء
@login_required
@subscription_required
def purchase_returns_list(request):
    returns = PurchaseReturn.objects.select_related('supplier', 'original_purchase').order_by('-created_at')
    
    status = request.GET.get('status', '')
    if status:
        returns = returns.filter(status=status)
    
    search = request.GET.get('search', '')
    if search:
        returns = returns.filter(
            Q(return_number__icontains=search) |
            Q(supplier__name__icontains=search) |
            Q(original_purchase__invoice_number__icontains=search)
        )
    
    context = {
        'returns': returns,
        'search': search,
        'status': status,
        'status_choices': PurchaseReturn.RETURN_STATUS,
        **get_user_context(request)
    }
    return render(request, 'purchase_returns.html', context)

@login_required
@subscription_required
def add_purchase_return(request):
    if request.method == 'POST':
        try:
            original_purchase_id = request.POST.get('original_purchase_id')
            reason = request.POST.get('reason', '')
            
            original_purchase = get_object_or_404(Purchase, id=original_purchase_id, status='confirmed')
            
            purchase_return = PurchaseReturn.objects.create(
                original_purchase=original_purchase,
                supplier=original_purchase.supplier,
                reason=reason,
                total_amount=0,
                created_by=request.user
            )
            
            items_data = json.loads(request.POST.get('items', '[]'))
            total_amount = 0
            
            for item_data in items_data:
                original_item = get_object_or_404(PurchaseItem, id=item_data['original_item_id'])
                quantity = float(item_data['quantity'])
                
                if quantity > original_item.quantity:
                    raise ValidationError(f'الكمية المرتجعة أكبر من الكمية الأصلية للمنتج {original_item.product.name}')
                
                return_item = PurchaseReturnItem.objects.create(
                    purchase_return=purchase_return,
                    original_purchase_item=original_item,
                    product=original_item.product,
                    quantity=quantity,
                    unit_price=original_item.unit_price
                )
                
                total_amount += return_item.total_price
            
            purchase_return.total_amount = total_amount
            purchase_return.save()
            
            messages.success(request, f'تم إنشاء مرتجع الشراء #{purchase_return.return_number} بنجاح')
            return redirect('purchase_returns')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    confirmed_purchases = Purchase.objects.filter(status='confirmed').select_related('supplier').order_by('-created_at')[:50]
    
    context = {
        'confirmed_purchases': confirmed_purchases,
        **get_user_context(request)
    }
    return render(request, 'add_purchase_return.html', context)

@login_required
@subscription_required
def confirm_purchase_return_view(request, return_id):
    purchase_return = get_object_or_404(PurchaseReturn, id=return_id)
    
    try:
        purchase_return.confirm_return(request.user)
        messages.success(request, f'تم تأكيد مرتجع الشراء #{purchase_return.return_number} وتحديث المخزون')
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'خطأ: {str(e)}')
    
    return redirect('purchase_returns')

# API لجلب عناصر الفاتورة
@login_required
def get_sale_items(request, sale_id):
    try:
        sale = get_object_or_404(Sale, id=sale_id, status='confirmed')
        items = []
        
        for item in sale.items.all():
            items.append({
                'id': item.id,
                'product_name': item.product.name,
                'quantity': float(item.quantity),
                'unit_price': float(item.unit_price),
                'total_price': float(item.total_price)
            })
        
        return JsonResponse({
            'success': True,
            'items': items,
            'customer_name': sale.customer.name,
            'invoice_number': sale.invoice_number
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# API لطباعة فاتورة المرتجع
@login_required
def print_sale_return(request, return_id):
    sale_return = get_object_or_404(SaleReturn, id=return_id)
    
    context = {
        'sale_return': sale_return,
        'items': sale_return.items.all(),
        'company_name': get_setting('company_name', 'شركة ERP'),
        'currency_symbol': get_setting('currency_symbol', 'ر.س'),
        **get_user_context(request)
    }
    return render(request, 'print_sale_return.html', context)

# API لتصدير فاتورة المرتجع إلى Excel
@login_required
def export_sale_return_excel(request, return_id):
    try:
        import openpyxl
        from django.http import HttpResponse
        
        sale_return = get_object_or_404(SaleReturn, id=return_id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"مرتجع {sale_return.return_number}"
        
        # رأس الفاتورة
        ws.cell(row=1, column=1, value=f"فاتورة مرتجع بيع - {sale_return.return_number}")
        ws.cell(row=2, column=1, value=f"العميل: {sale_return.customer.name}")
        ws.cell(row=3, column=1, value=f"التاريخ: {sale_return.created_at.strftime('%Y-%m-%d')}")
        
        # رأس الجدول
        headers = ['#', 'اسم الصنف', 'كود الصنف', 'الكمية', 'السعر', 'الخصم %', 'الإجمالي']
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)
        
        # البيانات
        row = 6
        for index, item in enumerate(sale_return.items.all(), 1):
            ws.cell(row=row, column=1, value=index)
            ws.cell(row=row, column=2, value=item.product.name)
            ws.cell(row=row, column=3, value=item.product.barcode or '')
            ws.cell(row=row, column=4, value=float(item.quantity))
            ws.cell(row=row, column=5, value=float(item.unit_price))
            ws.cell(row=row, column=6, value=float(item.discount_percent))
            ws.cell(row=row, column=7, value=float(item.total_price))
            row += 1
        
        # الإجماليات
        row += 1
        ws.cell(row=row, column=6, value="الإجمالي قبل الخصم:")
        ws.cell(row=row, column=7, value=float(sale_return.subtotal))
        row += 1
        ws.cell(row=row, column=6, value="الضريبة:")
        ws.cell(row=row, column=7, value=float(sale_return.tax_amount))
        row += 1
        ws.cell(row=row, column=6, value="الإجمالي النهائي:")
        ws.cell(row=row, column=7, value=float(sale_return.total_amount))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="مرتجع_{sale_return.return_number}.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('sale_returns')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('sale_returns')

# API لتصدير فاتورة المرتجع إلى PDF
@login_required
def export_sale_return_pdf(request, return_id):
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from django.http import HttpResponse
        
        sale_return = get_object_or_404(SaleReturn, id=return_id)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="مرتجع_{sale_return.return_number}.pdf"'
        
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        
        # رأس الفاتورة
        p.drawString(100, height - 100, f"Sale Return Invoice - {sale_return.return_number}")
        p.drawString(100, height - 120, f"Customer: {sale_return.customer.name}")
        p.drawString(100, height - 140, f"Date: {sale_return.created_at.strftime('%Y-%m-%d')}")
        
        # الجدول
        y = height - 200
        p.drawString(50, y, "#")
        p.drawString(100, y, "Product")
        p.drawString(250, y, "Qty")
        p.drawString(300, y, "Price")
        p.drawString(350, y, "Discount")
        p.drawString(400, y, "Total")
        
        y -= 20
        for index, item in enumerate(sale_return.items.all(), 1):
            p.drawString(50, y, str(index))
            p.drawString(100, y, item.product.name[:20])
            p.drawString(250, y, str(item.quantity))
            p.drawString(300, y, str(item.unit_price))
            p.drawString(350, y, f"{item.discount_percent}%")
            p.drawString(400, y, str(item.total_price))
            y -= 20
        
        # الإجماليات
        y -= 20
        p.drawString(300, y, f"Subtotal: {sale_return.subtotal}")
        y -= 20
        p.drawString(300, y, f"Tax: {sale_return.tax_amount}")
        y -= 20
        p.drawString(300, y, f"Total: {sale_return.total_amount}")
        
        p.showPage()
        p.save()
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة reportlab غير مثبتة')
        return redirect('sale_returns')
    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('sale_returns')

@login_required
def get_purchase_items(request, purchase_id):
    try:
        purchase = get_object_or_404(Purchase, id=purchase_id, status='confirmed')
        items = []
        
        for item in purchase.items.all():
            items.append({
                'id': item.id,
                'product_name': item.product.name,
                'quantity': float(item.quantity),
                'unit_price': float(item.unit_price),
                'total_price': float(item.total_price)
            })
        
        return JsonResponse({
            'success': True,
            'items': items,
            'supplier_name': purchase.supplier.name,
            'invoice_number': purchase.invoice_number
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# تقارير المبيعات حسب النوع
@login_required
@subscription_required
# دالة جرد المخزون المتقدمة
@login_required
@subscription_required
@permission_required('stock', 'edit')
def inventory_audit(request):
    if request.method == 'POST':
        try:
            audit_data = json.loads(request.POST.get('audit_data', '[]'))
            total_adjustments = 0
            
            for item in audit_data:
                product_id = item['product_id']
                actual_quantity = float(item['actual_quantity'])
                notes = item.get('notes', 'جرد شامل')
                
                product = get_object_or_404(Product, id=product_id)
                
                # معالجة الجرد مع القيود المحاسبية
                entry = InventoryAccountingManager.process_inventory_adjustment(
                    product, actual_quantity, request.user, notes
                )
                
                if entry:
                    total_adjustments += 1
            
            messages.success(request, f'تم جرد {len(audit_data)} منتج مع {total_adjustments} تسوية محاسبية')
            return JsonResponse({'success': True, 'adjustments': total_adjustments})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # عرض صفحة الجرد
    products = Product.objects.filter(is_active=True).order_by('name')
    products_data = []
    
    for product in products:
        current_stock = getattr(product, 'stock', 0) or 0
        products_data.append({
            'id': product.id,
            'name': product.name,
            'barcode': product.barcode or '',
            'current_stock': float(current_stock),
            'unit': product.unit or 'قطعة'
        })
    
    context = {
        'products_data': products_data,
        'currency_symbol': get_setting('currency_symbol', 'د.ك'),
        **get_user_context(request)
    }
    return render(request, 'inventory_audit.html', context)

def sales_by_type_report(request):
    from django.db.models import Sum, Count
    
    # إحصائيات المبيعات
    try:
        invoice_sales = Sale.objects.filter(sale_type='invoice', status='confirmed')
        pos_sales = Sale.objects.filter(sale_type='pos', status='confirmed')
    except:
        # في حالة عدم وجود حقل sale_type
        invoice_sales = Sale.objects.filter(status='confirmed')
        pos_sales = Sale.objects.none()
    
    stats = {
        'invoice_count': invoice_sales.count(),
        'invoice_total': invoice_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'pos_count': pos_sales.count(),
        'pos_total': pos_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
    }
    
    stats['total_count'] = stats['invoice_count'] + stats['pos_count']
    stats['total_amount'] = stats['invoice_total'] + stats['pos_total']
    
    # فلترة حسب التاريخ
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        invoice_sales = invoice_sales.filter(created_at__gte=start_date)
        pos_sales = pos_sales.filter(created_at__gte=start_date)
    if end_date:
        invoice_sales = invoice_sales.filter(created_at__lte=end_date)
        pos_sales = pos_sales.filter(created_at__lte=end_date)
    
    context = {
        'stats': stats,
        'invoice_sales': invoice_sales[:20],
        'pos_sales': pos_sales[:20],
        'start_date': start_date,
        'end_date': end_date,
        **get_user_context(request)
    }
    
    return render(request, 'sales_by_type_report.html', context)
   

@login_required
@subscription_required
@permission_required('permissions', 'view')
def permissions_list(request):
    users = User.objects.all().order_by('username')
    
    # تحضير بيانات الصلاحيات لكل مستخدم
    users_with_permissions = []
    for user in users:
        # جلب صلاحيات المستخدم
        permissions = Permission.objects.filter(user=user)
        permissions_count = permissions.count()
        
        # حساب الشاشات المتاحة
        accessible_screens = permissions.filter(can_view=True).count()
        
        users_with_permissions.append({
            'user': user,
            'permissions_count': permissions_count,
            'accessible_screens': accessible_screens,
            'is_superuser': user.is_superuser,
            'last_login': user.last_login
        })
    
    context = {
        'users_with_permissions': users_with_permissions,
        'available_screens': PermissionSystem.AVAILABLE_SCREENS,
        'available_actions': PermissionSystem.AVAILABLE_ACTIONS,
        'permission_groups': PermissionSystem.PERMISSION_GROUPS,
        **get_user_context(request)
    }
    return render(request, 'permissions.html', context)

@login_required
@subscription_required
@permission_required('permissions', 'view')
def permissions_management(request):
    """إدارة صلاحيات المستخدمين"""
    from .permissions_utils import AVAILABLE_SCREENS, AVAILABLE_ACTIONS
    
    # جلب جميع المستخدمين النشطين
    users = User.objects.filter(is_active=True).order_by('first_name', 'username')
    
    # جلب صلاحيات كل مستخدم
    users_with_permissions = []
    for user in users:
        user_permissions = {}
        permissions = Permission.objects.filter(user=user)
        
        for permission in permissions:
            user_permissions[permission.screen] = {
                'view': permission.can_view,
                'add': permission.can_add,
                'edit': permission.can_edit,
                'delete': permission.can_delete,
                'confirm': permission.can_confirm,
                'print': permission.can_print,
                'export': permission.can_export,
            }
        
        users_with_permissions.append({
            'user': user,
            'permissions': user_permissions
        })
    
    context = {
        'users_with_permissions': users_with_permissions,
        'available_screens': AVAILABLE_SCREENS,
        'available_actions': AVAILABLE_ACTIONS,
        'total_users': users.count(),
        'available_screens_count': len(AVAILABLE_SCREENS),
        **get_user_context(request)
    }
    return render(request, 'permissions.html', context)

@login_required
@subscription_required
@permission_required('permissions', 'edit')
@csrf_exempt
def update_permission(request):
    """تحديث صلاحية مستخدم"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        user_id = request.POST.get('user_id')
        screen = request.POST.get('screen')
        action = request.POST.get('action')
        value = request.POST.get('value') == 'true'
        
        if not all([user_id, screen, action]):
            return JsonResponse({'success': False, 'error': 'بيانات ناقصة'})
        
        user = get_object_or_404(User, id=user_id)
        
        # منع تعديل صلاحيات المدير العام
        if user.is_superuser:
            return JsonResponse({'success': False, 'error': 'لا يمكن تعديل صلاحيات المدير العام'})
        
        # الحصول على أو إنشاء الصلاحية
        company = getattr(request, 'company', None)
        permission, created = Permission.objects.get_or_create(
            user=user,
            screen=screen,
            company=company,
            defaults={'created_by': request.user}
        )
        
        # تحديث الصلاحية
        setattr(permission, f'can_{action}', value)
        permission.save()
        
        return JsonResponse({
            'success': True,
            'message': f'تم تحديث صلاحية {action} في {screen} للمستخدم {user.get_full_name() or user.username}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# إضافة alias للدالة الحالية
permissions = permissions_management