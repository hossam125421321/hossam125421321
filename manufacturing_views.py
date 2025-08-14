from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import date, timedelta
import json
from .models import *
from .decorators import permission_required, subscription_required

@login_required
@subscription_required
def manufacturing_list(request):
    """قائمة أوامر التصنيع"""
    try:
        orders = ManufacturingOrder.objects.select_related('product', 'warehouse', 'created_by').order_by('-created_at')
    except:
        orders = []
    
    search = request.GET.get('search', '')
    if search:
        orders = orders.filter(
            Q(order_number__icontains=search) |
            Q(product__name__icontains=search)
        )
    
    context = {
        'orders': orders,
        'search': search,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'manufacturing.html', context)

@login_required
@subscription_required
def add_manufacturing_order(request):
    """إضافة أمر تصنيع جديد"""
    if request.method == 'POST':
        try:
            product_id = request.POST.get('product_id')
            quantity = float(request.POST.get('quantity', 0))
            warehouse_id = request.POST.get('warehouse_id')
            start_date = request.POST.get('start_date')
            notes = request.POST.get('notes', '')
            
            if quantity <= 0:
                messages.error(request, 'الكمية يجب أن تكون أكبر من صفر')
                return redirect('add_manufacturing_order')
            
            # الحصول على الشركة
            company = getattr(request, 'company', None)
            if not company:
                company, created = Company.objects.get_or_create(
                    code='DEFAULT',
                    defaults={
                        'name': 'الشركة الافتراضية',
                        'database_name': 'erp_default',
                        'subscription_end': date.today() + timedelta(days=365)
                    }
                )
            
            from datetime import datetime
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else date.today()
            
            order = ManufacturingOrder.objects.create(
                company=company,
                product_id=product_id,
                quantity=quantity,
                warehouse_id=warehouse_id,
                start_date=start_date,
                notes=notes,
                status='draft',
                created_by=request.user
            )
            
            messages.success(request, f'تم إنشاء أمر التصنيع #{order.order_number} بنجاح')
            return redirect('manufacturing')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'products': Product.objects.filter(is_active=True, bom=True),
        'warehouses': Warehouse.objects.filter(is_active=True),
        'today': date.today().strftime('%Y-%m-%d'),
    }
    return render(request, 'add_manufacturing.html', context)

@login_required
@subscription_required
def manufacturing_order_detail(request, order_id):
    """تفاصيل أمر التصنيع"""
    order = get_object_or_404(ManufacturingOrder, id=order_id)
    
    context = {
        'order': order,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'manufacturing_detail.html', context)

@login_required
@subscription_required
def edit_manufacturing_order(request, order_id):
    """تعديل أمر التصنيع"""
    order = get_object_or_404(ManufacturingOrder, id=order_id)
    
    if request.method == 'POST':
        try:
            if order.status != 'draft':
                messages.error(request, 'لا يمكن تعديل أمر مؤكد')
                return redirect('manufacturing_order_detail', order_id=order.id)
            
            order.quantity = float(request.POST.get('quantity', 0))
            order.warehouse_id = request.POST.get('warehouse_id')
            order.notes = request.POST.get('notes', '')
            
            start_date = request.POST.get('start_date')
            if start_date:
                from datetime import datetime
                order.start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            
            order.save()
            
            messages.success(request, 'تم تحديث أمر التصنيع بنجاح')
            return redirect('manufacturing_order_detail', order_id=order.id)
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    context = {
        'order': order,
        'warehouses': Warehouse.objects.filter(is_active=True),
    }
    return render(request, 'edit_manufacturing.html', context)

@login_required
@subscription_required
@csrf_exempt
def confirm_manufacturing_order(request, order_id):
    """تأكيد أمر التصنيع"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        order = get_object_or_404(ManufacturingOrder, id=order_id)
        
        if order.status != 'draft':
            return JsonResponse({'success': False, 'error': 'الأمر مؤكد بالفعل'})
        
        order.status = 'confirmed'
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': f'تم تأكيد أمر التصنيع #{order.order_number} بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@csrf_exempt
def start_manufacturing_order(request, order_id):
    """بدء تنفيذ أمر التصنيع"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        order = get_object_or_404(ManufacturingOrder, id=order_id)
        
        if order.status != 'confirmed':
            return JsonResponse({'success': False, 'error': 'يجب تأكيد الأمر أولاً'})
        
        order.status = 'in_progress'
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': f'تم بدء تنفيذ أمر التصنيع #{order.order_number}'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
@csrf_exempt
def complete_manufacturing_order(request, order_id):
    """إكمال أمر التصنيع"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'})
    
    try:
        order = get_object_or_404(ManufacturingOrder, id=order_id)
        
        if order.status != 'in_progress':
            return JsonResponse({'success': False, 'error': 'الأمر ليس قيد التنفيذ'})
        
        # تحديث المخزون
        product = order.product
        if hasattr(product, 'stock'):
            product.stock = (product.stock or 0) + order.quantity
            product.save()
        
        # تسجيل حركة مخزون
        StockMovement.objects.create(
            product=product,
            warehouse=order.warehouse,
            movement_type='in',
            quantity=order.quantity,
            reference=f'إنتاج #{order.order_number}',
            created_by=request.user
        )
        
        order.status = 'completed'
        order.produced_quantity = order.quantity
        order.end_date = date.today()
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': f'تم إكمال أمر التصنيع #{order.order_number} بنجاح'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@subscription_required
def manufacturing_materials_report(request):
    """تقرير المواد الخام"""
    context = {
        'materials': [],  # سيتم تطويرها لاحقاً
        'currency_symbol': 'د.ك',
    }
    return render(request, 'manufacturing_materials_report.html', context)

@login_required
@subscription_required
def manufacturing_production_report(request):
    """تقرير الإنتاج"""
    try:
        orders = ManufacturingOrder.objects.filter(status='completed').order_by('-end_date')[:50]
    except:
        orders = []
    
    context = {
        'orders': orders,
        'currency_symbol': 'د.ك',
    }
    return render(request, 'manufacturing_production_report.html', context)

@login_required
@subscription_required
def manufacturing_cost_report(request):
    """تقرير تكاليف التصنيع"""
    context = {
        'costs': [],  # سيتم تطويرها لاحقاً
        'currency_symbol': 'د.ك',
    }
    return render(request, 'manufacturing_cost_report.html', context)

@login_required
@subscription_required
def update_manufacturing_inventory(request):
    """تحديث مخزون التصنيع"""
    messages.success(request, 'سيتم إضافة هذه الميزة قريباً')
    return redirect('manufacturing')

@login_required
@subscription_required
def export_manufacturing_data(request):
    """تصدير بيانات التصنيع"""
    try:
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "أوامر التصنيع"
        
        headers = ['رقم الأمر', 'المنتج', 'الكمية المطلوبة', 'الكمية المنتجة', 'الحالة', 'تاريخ البدء', 'تاريخ الانتهاء']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        try:
            orders = ManufacturingOrder.objects.all()[:100]
            for row, order in enumerate(orders, 2):
                ws.cell(row=row, column=1, value=order.order_number)
                ws.cell(row=row, column=2, value=order.product.name)
                ws.cell(row=row, column=3, value=float(order.quantity))
                ws.cell(row=row, column=4, value=float(order.produced_quantity))
                ws.cell(row=row, column=5, value=dict(ManufacturingOrder.ORDER_STATUS).get(order.status, order.status))
                ws.cell(row=row, column=6, value=order.start_date.strftime('%Y-%m-%d') if order.start_date else '')
                ws.cell(row=row, column=7, value=order.end_date.strftime('%Y-%m-%d') if order.end_date else '')
        except:
            pass
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="manufacturing_orders.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('manufacturing')

@login_required
@subscription_required
def manufacturing_settings(request):
    """إعدادات التصنيع"""
    if request.method == 'POST':
        messages.success(request, 'سيتم إضافة هذه الميزة قريباً')
        return redirect('manufacturing_settings')
    
    context = {
        'settings': {},  # سيتم تطويرها لاحقاً
    }
    return render(request, 'manufacturing_settings.html', context)