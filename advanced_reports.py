# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Avg, Q, F
from django.utils import timezone
from datetime import date, datetime, timedelta
import json

from .models import *
from .decorators import permission_required, subscription_required

@login_required
@subscription_required
@permission_required('reports', 'view')
def advanced_reports_dashboard(request):
    """لوحة تحكم التقارير المتقدمة"""
    try:
        # إحصائيات سريعة
        today = date.today()
        this_month = today.replace(day=1)
        last_month = (this_month - timedelta(days=1)).replace(day=1)
        
        stats = {
            # مبيعات اليوم
            'today_sales': Sale.objects.filter(
                created_at__date=today,
                status='confirmed'
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
            
            # مبيعات الشهر
            'month_sales': Sale.objects.filter(
                created_at__gte=this_month,
                status='confirmed'
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
            
            # مبيعات الشهر الماضي
            'last_month_sales': Sale.objects.filter(
                created_at__gte=last_month,
                created_at__lt=this_month,
                status='confirmed'
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
            
            # عدد العملاء النشطين
            'active_customers': Customer.objects.filter(is_active=True).count(),
            
            # عدد المنتجات النشطة
            'active_products': Product.objects.filter(is_active=True).count(),
            
            # المنتجات منخفضة المخزون
            'low_stock_count': get_low_stock_count(),
        }
        
        # حساب نسبة النمو
        if stats['last_month_sales'] > 0:
            growth_rate = ((stats['month_sales'] - stats['last_month_sales']) / stats['last_month_sales']) * 100
        else:
            growth_rate = 0
        
        stats['growth_rate'] = round(growth_rate, 2)
        
        # أفضل المنتجات مبيعاً
        top_products = get_top_selling_products(limit=5)
        
        # أفضل العملاء
        top_customers = get_top_customers(limit=5)
        
        context = {
            'stats': stats,
            'top_products': top_products,
            'top_customers': top_customers,
            'currency_symbol': 'د.ك',
        }
        
        return render(request, 'reports/advanced_dashboard.html', context)
        
    except Exception as e:
        return render(request, 'reports/advanced_dashboard.html', {
            'stats': {},
            'error': str(e)
        })

def get_low_stock_count():
    """حساب عدد المنتجات منخفضة المخزون"""
    try:
        count = 0
        for product in Product.objects.filter(is_active=True):
            current_stock = float(getattr(product, 'stock', 0) or 0)
            min_stock = 10  # الحد الأدنى الافتراضي
            if current_stock <= min_stock:
                count += 1
        return count
    except:
        return 0

def get_top_selling_products(limit=10, days=30):
    """أفضل المنتجات مبيعاً"""
    try:
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        # جمع المبيعات من الفواتير العادية
        sale_items = SaleItem.objects.filter(
            sale__status='confirmed',
            sale__created_at__gte=start_date
        ).values('product__name').annotate(
            total_quantity=Sum('quantity'),
            total_amount=Sum('total_price')
        ).order_by('-total_amount')[:limit]
        
        # جمع المبيعات من نقاط البيع
        pos_items = POSSaleItem.objects.filter(
            pos_sale__created_at__gte=start_date
        ).values('product__name').annotate(
            total_quantity=Sum('quantity'),
            total_amount=Sum('total_price')
        ).order_by('-total_amount')[:limit]
        
        # دمج النتائج (مبسط)
        products = []
        for item in sale_items:
            products.append({
                'name': item['product__name'],
                'quantity': float(item['total_quantity']),
                'amount': float(item['total_amount'])
            })
        
        return products
        
    except Exception as e:
        return []

def get_top_customers(limit=10, days=30):
    """أفضل العملاء"""
    try:
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        customers = Customer.objects.annotate(
            total_purchases=Sum(
                'sale__total_amount',
                filter=Q(
                    sale__status='confirmed',
                    sale__created_at__gte=start_date
                )
            ),
            orders_count=Count(
                'sale',
                filter=Q(
                    sale__status='confirmed',
                    sale__created_at__gte=start_date
                )
            )
        ).filter(
            total_purchases__gt=0
        ).order_by('-total_purchases')[:limit]
        
        customers_data = []
        for customer in customers:
            customers_data.append({
                'name': customer.name,
                'total_purchases': float(customer.total_purchases or 0),
                'orders_count': customer.orders_count or 0,
                'phone': customer.phone
            })
        
        return customers_data
        
    except Exception as e:
        return []

@login_required
@subscription_required
@permission_required('reports', 'view')
def sales_analysis_report(request):
    """تقرير تحليل المبيعات"""
    try:
        # فلترة حسب التاريخ
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        customer_id = request.GET.get('customer_id')
        product_id = request.GET.get('product_id')
        
        # تحديد التاريخ الافتراضي (آخر 30 يوم)
        if not start_date:
            start_date = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = date.today().strftime('%Y-%m-%d')
        
        # بناء الاستعلام
        sales_query = Sale.objects.filter(
            status='confirmed',
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        if customer_id:
            sales_query = sales_query.filter(customer_id=customer_id)
        
        # إحصائيات عامة
        total_sales = sales_query.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_orders = sales_query.count()
        avg_order_value = sales_query.aggregate(Avg('total_amount'))['total_amount__avg'] or 0
        
        # المبيعات حسب اليوم
        daily_sales = sales_query.extra(
            select={'day': 'date(created_at)'}
        ).values('day').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('day')
        
        # المبيعات حسب العميل
        customer_sales = sales_query.values(
            'customer__name'
        ).annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('-total')[:10]
        
        # المبيعات حسب المنتج
        product_sales = SaleItem.objects.filter(
            sale__in=sales_query
        ).values(
            'product__name'
        ).annotate(
            quantity=Sum('quantity'),
            total=Sum('total_price')
        ).order_by('-total')[:10]
        
        # تحضير البيانات للرسوم البيانية
        chart_data = {
            'daily_labels': [item['day'] for item in daily_sales],
            'daily_values': [float(item['total']) for item in daily_sales],
            'customer_labels': [item['customer__name'] for item in customer_sales],
            'customer_values': [float(item['total']) for item in customer_sales],
            'product_labels': [item['product__name'] for item in product_sales],
            'product_values': [float(item['total']) for item in product_sales],
        }
        
        context = {
            'total_sales': total_sales,
            'total_orders': total_orders,
            'avg_order_value': avg_order_value,
            'daily_sales': daily_sales,
            'customer_sales': customer_sales,
            'product_sales': product_sales,
            'chart_data': json.dumps(chart_data),
            'start_date': start_date,
            'end_date': end_date,
            'customer_id': customer_id,
            'product_id': product_id,
            'customers': Customer.objects.all().order_by('name'),
            'products': Product.objects.all().order_by('name'),
            'currency_symbol': 'د.ك',
        }
        
        return render(request, 'reports/sales_analysis.html', context)
        
    except Exception as e:
        return render(request, 'reports/sales_analysis.html', {
            'error': str(e),
            'customers': Customer.objects.all(),
            'products': Product.objects.all(),
        })

@login_required
@subscription_required
@permission_required('reports', 'view')
def inventory_analysis_report(request):
    """تقرير تحليل المخزون"""
    try:
        # جمع بيانات المخزون
        products = Product.objects.filter(is_active=True)
        
        inventory_data = []
        total_value = 0
        low_stock_count = 0
        out_of_stock_count = 0
        
        for product in products:
            current_stock = float(getattr(product, 'stock', 0) or 0)
            cost_price = 0
            selling_price = 0
            
            # الحصول على الأسعار
            try:
                price_obj = ProductPrice.objects.filter(product=product).first()
                if price_obj:
                    cost_price = float(price_obj.cost_price or 0)
                    selling_price = float(price_obj.selling_price or 0)
                elif hasattr(product, 'price'):
                    selling_price = float(product.price or 0)
                    cost_price = float(getattr(product, 'cost_price', 0) or 0)
            except:
                pass
            
            stock_value = current_stock * cost_price
            total_value += stock_value
            
            # تحديد حالة المخزون
            min_stock = 10  # يمكن جعلها قابلة للتخصيص
            if current_stock <= 0:
                stock_status = 'نفد'
                out_of_stock_count += 1
            elif current_stock <= min_stock:
                stock_status = 'منخفض'
                low_stock_count += 1
            else:
                stock_status = 'جيد'
            
            # حساب معدل الدوران (مبسط)
            try:
                # مبيعات آخر 30 يوم
                recent_sales = SaleItem.objects.filter(
                    product=product,
                    sale__status='confirmed',
                    sale__created_at__gte=timezone.now() - timedelta(days=30)
                ).aggregate(Sum('quantity'))['quantity__sum'] or 0
                
                turnover_rate = float(recent_sales) / max(current_stock, 1)
            except:
                turnover_rate = 0
            
            inventory_data.append({
                'product': product,
                'current_stock': current_stock,
                'cost_price': cost_price,
                'selling_price': selling_price,
                'stock_value': stock_value,
                'stock_status': stock_status,
                'turnover_rate': round(turnover_rate, 2),
                'recent_sales': recent_sales,
            })
        
        # ترتيب حسب قيمة المخزون
        inventory_data.sort(key=lambda x: x['stock_value'], reverse=True)
        
        # إحصائيات عامة
        stats = {
            'total_products': len(inventory_data),
            'total_value': total_value,
            'low_stock_count': low_stock_count,
            'out_of_stock_count': out_of_stock_count,
            'good_stock_count': len(inventory_data) - low_stock_count - out_of_stock_count,
        }
        
        # تحضير بيانات الرسم البياني
        status_chart = {
            'labels': ['جيد', 'منخفض', 'نفد'],
            'values': [stats['good_stock_count'], stats['low_stock_count'], stats['out_of_stock_count']],
            'colors': ['#28a745', '#ffc107', '#dc3545']
        }
        
        context = {
            'inventory_data': inventory_data,
            'stats': stats,
            'status_chart': json.dumps(status_chart),
            'currency_symbol': 'د.ك',
        }
        
        return render(request, 'reports/inventory_analysis.html', context)
        
    except Exception as e:
        return render(request, 'reports/inventory_analysis.html', {
            'error': str(e),
            'inventory_data': [],
            'stats': {}
        })

@login_required
@subscription_required
@permission_required('reports', 'view')
def financial_summary_report(request):
    """تقرير الملخص المالي"""
    try:
        # فلترة حسب التاريخ
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # تحديد التاريخ الافتراضي (الشهر الحالي)
        if not start_date:
            start_date = date.today().replace(day=1).strftime('%Y-%m-%d')
        if not end_date:
            end_date = date.today().strftime('%Y-%m-%d')
        
        # المبيعات
        sales_data = Sale.objects.filter(
            status='confirmed',
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).aggregate(
            total_sales=Sum('total_amount'),
            total_discount=Sum('discount_amount'),
            total_tax=Sum('tax_amount'),
            count=Count('id')
        )
        
        # المشتريات
        purchases_data = Purchase.objects.filter(
            status='confirmed',
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).aggregate(
            total_purchases=Sum('total_amount'),
            total_discount=Sum('discount_amount'),
            total_tax=Sum('tax_amount'),
            count=Count('id')
        )
        
        # دفعات العملاء
        customer_payments = CustomerPayment.objects.filter(
            payment_date__gte=start_date,
            payment_date__lte=end_date
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # دفعات الموردين
        supplier_payments = SupplierPayment.objects.filter(
            payment_date__gte=start_date,
            payment_date__lte=end_date
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # حساب الأرباح المبدئية
        total_sales = sales_data['total_sales'] or 0
        total_purchases = purchases_data['total_purchases'] or 0
        gross_profit = total_sales - total_purchases
        
        # حساب صافي التدفق النقدي
        cash_inflow = customer_payments
        cash_outflow = supplier_payments
        net_cash_flow = cash_inflow - cash_outflow
        
        # إحصائيات مفصلة
        financial_summary = {
            'sales': {
                'total': total_sales,
                'discount': sales_data['total_discount'] or 0,
                'tax': sales_data['total_tax'] or 0,
                'count': sales_data['count'] or 0,
                'average': (total_sales / max(sales_data['count'], 1)) if sales_data['count'] else 0
            },
            'purchases': {
                'total': total_purchases,
                'discount': purchases_data['total_discount'] or 0,
                'tax': purchases_data['total_tax'] or 0,
                'count': purchases_data['count'] or 0,
                'average': (total_purchases / max(purchases_data['count'], 1)) if purchases_data['count'] else 0
            },
            'payments': {
                'customer_payments': customer_payments,
                'supplier_payments': supplier_payments,
                'net_cash_flow': net_cash_flow
            },
            'profitability': {
                'gross_profit': gross_profit,
                'profit_margin': (gross_profit / max(total_sales, 1)) * 100 if total_sales else 0
            }
        }
        
        # المبيعات والمشتريات حسب اليوم
        daily_data = []
        current_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        while current_date <= end_date_obj:
            day_sales = Sale.objects.filter(
                status='confirmed',
                created_at__date=current_date
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            
            day_purchases = Purchase.objects.filter(
                status='confirmed',
                created_at__date=current_date
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            
            daily_data.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'sales': float(day_sales),
                'purchases': float(day_purchases),
                'profit': float(day_sales - day_purchases)
            })
            
            current_date += timedelta(days=1)
        
        # تحضير بيانات الرسم البياني
        chart_data = {
            'dates': [item['date'] for item in daily_data],
            'sales': [item['sales'] for item in daily_data],
            'purchases': [item['purchases'] for item in daily_data],
            'profit': [item['profit'] for item in daily_data]
        }
        
        context = {
            'financial_summary': financial_summary,
            'daily_data': daily_data,
            'chart_data': json.dumps(chart_data),
            'start_date': start_date,
            'end_date': end_date,
            'currency_symbol': 'د.ك',
        }
        
        return render(request, 'reports/financial_summary.html', context)
        
    except Exception as e:
        return render(request, 'reports/financial_summary.html', {
            'error': str(e),
            'financial_summary': {},
            'daily_data': []
        })

@login_required
@subscription_required
@permission_required('reports', 'export')
def export_advanced_report(request):
    """تصدير التقارير المتقدمة"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        report_type = request.GET.get('type', 'sales')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        wb = openpyxl.Workbook()
        
        if report_type == 'sales':
            ws = wb.active
            ws.title = "تحليل المبيعات"
            
            # إعداد الرؤوس
            headers = ['التاريخ', 'رقم الفاتورة', 'العميل', 'المبلغ الإجمالي', 'الخصم', 'الضريبة', 'الحالة']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # البيانات
            sales = Sale.objects.filter(status='confirmed').order_by('-created_at')
            if start_date:
                sales = sales.filter(created_at__date__gte=start_date)
            if end_date:
                sales = sales.filter(created_at__date__lte=end_date)
            
            for row, sale in enumerate(sales, 2):
                ws.cell(row=row, column=1, value=sale.created_at.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=sale.invoice_number)
                ws.cell(row=row, column=3, value=sale.customer.name)
                ws.cell(row=row, column=4, value=float(sale.total_amount))
                ws.cell(row=row, column=5, value=float(sale.discount_amount))
                ws.cell(row=row, column=6, value=float(sale.tax_amount))
                ws.cell(row=row, column=7, value=dict(Sale.INVOICE_STATUS).get(sale.status, sale.status))
        
        elif report_type == 'inventory':
            ws = wb.active
            ws.title = "تحليل المخزون"
            
            headers = ['المنتج', 'الباركود', 'المخزون الحالي', 'سعر التكلفة', 'سعر البيع', 'قيمة المخزون', 'الحالة']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            products = Product.objects.filter(is_active=True)
            for row, product in enumerate(products, 2):
                current_stock = float(getattr(product, 'stock', 0) or 0)
                
                # الحصول على الأسعار
                cost_price = 0
                selling_price = 0
                try:
                    price_obj = ProductPrice.objects.filter(product=product).first()
                    if price_obj:
                        cost_price = float(price_obj.cost_price or 0)
                        selling_price = float(price_obj.selling_price or 0)
                except:
                    pass
                
                stock_value = current_stock * cost_price
                
                # تحديد حالة المخزون
                if current_stock <= 0:
                    status = 'نفد'
                elif current_stock <= 10:
                    status = 'منخفض'
                else:
                    status = 'جيد'
                
                ws.cell(row=row, column=1, value=product.name)
                ws.cell(row=row, column=2, value=product.barcode or '')
                ws.cell(row=row, column=3, value=current_stock)
                ws.cell(row=row, column=4, value=cost_price)
                ws.cell(row=row, column=5, value=selling_price)
                ws.cell(row=row, column=6, value=stock_value)
                ws.cell(row=row, column=7, value=status)
        
        # تنسيق الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        return JsonResponse({'error': 'مكتبة openpyxl غير مثبتة'})
    except Exception as e:
        return JsonResponse({'error': str(e)})

# API للحصول على بيانات الرسوم البيانية
@login_required
def get_chart_data_api(request):
    """API للحصول على بيانات الرسوم البيانية"""
    try:
        chart_type = request.GET.get('type', 'sales')
        days = int(request.GET.get('days', 30))
        
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        if chart_type == 'sales':
            # مبيعات يومية
            daily_sales = Sale.objects.filter(
                status='confirmed',
                created_at__gte=start_date
            ).extra(
                select={'day': 'date(created_at)'}
            ).values('day').annotate(
                total=Sum('total_amount')
            ).order_by('day')
            
            data = {
                'labels': [item['day'] for item in daily_sales],
                'values': [float(item['total']) for item in daily_sales]
            }
            
        elif chart_type == 'products':
            # أفضل المنتجات
            top_products = SaleItem.objects.filter(
                sale__status='confirmed',
                sale__created_at__gte=start_date
            ).values('product__name').annotate(
                total=Sum('total_price')
            ).order_by('-total')[:10]
            
            data = {
                'labels': [item['product__name'] for item in top_products],
                'values': [float(item['total']) for item in top_products]
            }
            
        elif chart_type == 'customers':
            # أفضل العملاء
            top_customers = Sale.objects.filter(
                status='confirmed',
                created_at__gte=start_date
            ).values('customer__name').annotate(
                total=Sum('total_amount')
            ).order_by('-total')[:10]
            
            data = {
                'labels': [item['customer__name'] for item in top_customers],
                'values': [float(item['total']) for item in top_customers]
            }
            
        else:
            data = {'labels': [], 'values': []}
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})